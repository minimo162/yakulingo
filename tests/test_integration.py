# tests/test_integration.py
"""
Integration tests for YakuLingo translation workflows.
Tests end-to-end scenarios from input to output.
Bidirectional translation: Japanese → English, Other → Japanese.
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch
import openpyxl
from openpyxl.styles import Font

from yakulingo.models.types import (
    TranslationStatus,
    TranslationProgress,
    TextBlock,
    FileType,
)
from yakulingo.config.settings import AppSettings
from yakulingo.services.translation_service import TranslationService, BatchTranslator
from yakulingo.services.prompt_builder import PromptBuilder
from yakulingo.processors.excel_processor import ExcelProcessor
from yakulingo.processors.word_processor import WordProcessor
from yakulingo.processors.pptx_processor import PptxProcessor


# --- Fixtures ---

@pytest.fixture
def mock_copilot_with_translation():
    """
    Mock CopilotHandler that returns predictable translations.
    Bidirectional - returns translations based on input.
    """
    mock = MagicMock()
    mock.is_connected = True

    def translate_single(text, prompt, reference_files=None, char_limit=None):
        # Bidirectional translation mock
        return f"Translated: {text}"

    def translate_sync(texts, prompt, reference_files=None, char_limit=None):
        # Bidirectional batch translation mock
        return [f"Translated: {t}" for t in texts]

    mock.translate_single.side_effect = translate_single
    mock.translate_sync.side_effect = translate_sync
    return mock


@pytest.fixture
def settings():
    """Default AppSettings"""
    return AppSettings()


@pytest.fixture
def excel_file_jp(tmp_path):
    """Create Excel file with Japanese content"""
    file_path = tmp_path / "japanese_doc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "データ"

    # Japanese content
    ws["A1"] = "製品名"
    ws["B1"] = "価格"
    ws["C1"] = "説明"

    ws["A2"] = "りんご"
    ws["B2"] = "100"  # Number - should skip
    ws["C2"] = "新鮮な果物です"

    ws["A3"] = "バナナ"
    ws["B3"] = "80"
    ws["C3"] = "南国産のバナナです"

    # Apply Japanese fonts
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
        for cell in row:
            if cell.value and not str(cell.value).isdigit():
                cell.font = Font(name="MS Gothic", size=11)

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_file_en(tmp_path):
    """Create Excel file with English content"""
    file_path = tmp_path / "english_doc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "Product Name"
    ws["B1"] = "Price"
    ws["C1"] = "Description"

    ws["A2"] = "Apple"
    ws["B2"] = "100"
    ws["C2"] = "Fresh fruit from local farms"

    ws["A3"] = "Banana"
    ws["B3"] = "80"
    ws["C3"] = "Tropical banana imported"

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_file_multi_sheet(tmp_path):
    """Create Excel file with multiple sheets"""
    file_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()

    # Sheet1 - Summary
    ws1 = wb.active
    ws1.title = "概要"
    ws1["A1"] = "月次レポート"
    ws1["A2"] = "2024年1月"  # Date pattern - should skip

    # Sheet2 - Details
    ws2 = wb.create_sheet("詳細")
    ws2["A1"] = "売上データ"
    ws2["A2"] = "test@example.com"  # Email - should skip
    ws2["A3"] = "顧客情報"

    wb.save(file_path)
    return file_path


# --- Integration Tests: Text Translation Workflow ---

class TestTextTranslationWorkflow:
    """Test complete text translation workflow - bidirectional"""

    def test_japanese_text_translation(self, mock_copilot_with_translation, settings):
        """Complete Japanese text translation (should auto-detect and translate to English)"""
        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_text("こんにちは世界")

        assert result.status == TranslationStatus.COMPLETED
        assert result.output_text is not None
        assert result.blocks_translated == 1
        assert result.duration_seconds >= 0

    def test_english_text_translation(self, mock_copilot_with_translation, settings):
        """Complete English text translation (should auto-detect and translate to Japanese)"""
        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_text("Hello World")

        assert result.status == TranslationStatus.COMPLETED
        assert result.output_text is not None
        assert result.blocks_translated == 1

    def test_text_translation_with_reference_files(
        self, mock_copilot_with_translation, settings, tmp_path
    ):
        """Text translation with glossary reference file"""
        # Create a glossary file
        glossary_path = tmp_path / "glossary.csv"
        glossary_path.write_text("Japanese,English\nりんご,Apple\n", encoding="utf-8")

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_text(
            "りんごを食べる",
            reference_files=[glossary_path],
        )

        assert result.status == TranslationStatus.COMPLETED
        # Verify reference files were passed to copilot
        mock_copilot_with_translation.translate_single.assert_called_once()
        call_args = mock_copilot_with_translation.translate_single.call_args
        assert call_args[0][2] == [glossary_path]  # reference_files


# --- Integration Tests: File Translation Workflow ---

class TestFileTranslationWorkflow:
    """Test complete file translation workflow - bidirectional"""

    def test_excel_jp_workflow(
        self, mock_copilot_with_translation, settings, excel_file_jp, tmp_path
    ):
        """Complete Excel Japanese file translation workflow"""
        settings.output_directory = str(tmp_path / "output")
        Path(settings.output_directory).mkdir()

        service = TranslationService(mock_copilot_with_translation, settings)

        # Track progress
        progress_updates = []

        def on_progress(progress: TranslationProgress):
            progress_updates.append(progress)

        result = service.translate_file(
            excel_file_jp,
            on_progress=on_progress,
        )

        # Verify result
        assert result.status == TranslationStatus.COMPLETED
        assert result.output_path is not None
        assert result.output_path.exists()
        assert "_translated" in result.output_path.name
        assert result.blocks_translated > 0
        assert result.blocks_total > 0

        # Verify progress was reported
        assert len(progress_updates) > 0
        assert progress_updates[-1].current == 100

        # Verify output content
        wb = openpyxl.load_workbook(result.output_path)
        ws = wb.active

        # Numbers should be unchanged
        assert ws["B2"].value == "100"
        assert ws["B3"].value == "80"

    def test_excel_en_workflow(
        self, mock_copilot_with_translation, settings, excel_file_en, tmp_path
    ):
        """Complete Excel English file translation workflow"""
        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(excel_file_en)

        assert result.status == TranslationStatus.COMPLETED
        assert result.output_path is not None
        assert "_translated" in result.output_path.name

    def test_multi_sheet_excel_workflow(
        self, mock_copilot_with_translation, settings, excel_file_multi_sheet
    ):
        """Translation workflow with multiple sheets"""
        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(excel_file_multi_sheet)

        assert result.status == TranslationStatus.COMPLETED
        # Should have translated from both sheets
        # Sheet1: "月次レポート" (skips "2024年1月")
        # Sheet2: "売上データ", "顧客情報" (skips email)
        assert result.blocks_translated >= 3

    def test_file_translation_with_reference(
        self, mock_copilot_with_translation, settings, excel_file_jp, tmp_path
    ):
        """File translation with reference glossary"""
        glossary_path = tmp_path / "glossary.csv"
        glossary_path.write_text("JP,EN\nりんご,Apple\n", encoding="utf-8")

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(
            excel_file_jp,
            reference_files=[glossary_path],
        )

        assert result.status == TranslationStatus.COMPLETED


# --- Integration Tests: Batch Processing ---

class TestBatchProcessingWorkflow:
    """Test batch translation with various sizes"""

    def test_single_batch_processing(self, mock_copilot_with_translation, settings):
        """Process blocks that fit in single batch"""
        service = TranslationService(mock_copilot_with_translation, settings)

        # Create 10 blocks (well under batch limit)
        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}", location=f"A{i}")
            for i in range(10)
        ]

        results = service.batch_translator.translate_blocks(blocks)

        assert len(results) == 10
        assert mock_copilot_with_translation.translate_sync.call_count == 1

    def test_multiple_batch_processing(self, mock_copilot_with_translation, settings):
        """Process blocks spanning multiple batches"""
        service = TranslationService(mock_copilot_with_translation, settings)

        # Create 75 blocks (exceeds batch limit of 50)
        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}", location=f"A{i}")
            for i in range(75)
        ]

        progress_updates = []

        def on_progress(progress):
            progress_updates.append(progress)

        results = service.batch_translator.translate_blocks(
            blocks, on_progress=on_progress
        )

        assert len(results) == 75
        assert mock_copilot_with_translation.translate_sync.call_count == 2
        assert len(progress_updates) == 2

    def test_large_text_batch_splitting(self, mock_copilot_with_translation, settings):
        """Batches split when character limit exceeded"""
        service = TranslationService(mock_copilot_with_translation, settings)

        # Create blocks with large text (5000 chars each)
        large_text = "あ" * 5000
        blocks = [
            TextBlock(id=str(i), text=large_text, location=f"A{i}")
            for i in range(3)
        ]

        results = service.batch_translator.translate_blocks(blocks)

        assert len(results) == 3
        # 5000 + 5000 > 10000, so should split
        assert mock_copilot_with_translation.translate_sync.call_count >= 2


# --- Integration Tests: Progress Callback ---

class TestProgressCallbackIntegration:
    """Test progress callback throughout workflow"""

    def test_progress_stages(
        self, mock_copilot_with_translation, settings, excel_file_jp
    ):
        """Verify all progress stages are reported"""
        service = TranslationService(mock_copilot_with_translation, settings)

        progress_statuses = []

        def on_progress(progress: TranslationProgress):
            progress_statuses.append(progress.status)

        service.translate_file(
            excel_file_jp,
            on_progress=on_progress,
        )

        # Should have extraction, translation, application, and complete stages
        assert any("Extracting" in s for s in progress_statuses)
        assert any("Translating" in s or "Batch" in s for s in progress_statuses)
        assert any("Applying" in s for s in progress_statuses)
        assert any("Complete" in s for s in progress_statuses)

    def test_progress_percentage_increases(
        self, mock_copilot_with_translation, settings, excel_file_jp
    ):
        """Progress percentage increases monotonically"""
        service = TranslationService(mock_copilot_with_translation, settings)

        percentages = []

        def on_progress(progress: TranslationProgress):
            percentages.append(progress.current)

        service.translate_file(
            excel_file_jp,
            on_progress=on_progress,
        )

        # Percentages should generally increase
        assert percentages[-1] == 100
        # First should be less than last
        assert percentages[0] < percentages[-1]


# --- Integration Tests: Empty/Edge Cases ---

class TestEdgeCaseWorkflows:
    """Test edge case workflows"""

    def test_empty_file_workflow(
        self, mock_copilot_with_translation, settings, tmp_path
    ):
        """Handle file with no translatable content"""
        # Create file with only numbers
        file_path = tmp_path / "numbers_only.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "12345"
        ws["A2"] = "67890"
        ws["B1"] = "test@example.com"
        wb.save(file_path)

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(file_path)

        assert result.status == TranslationStatus.COMPLETED
        assert result.blocks_total == 0
        assert result.warnings is not None
        assert any("No translatable" in w for w in result.warnings)

    def test_cancellation_workflow(
        self, mock_copilot_with_translation, settings, excel_file_jp
    ):
        """Test cancellation flag behavior"""
        service = TranslationService(mock_copilot_with_translation, settings)

        # Note: _cancel_requested is reset at start of translate_file()
        # So calling cancel() before translate_file() has no effect
        service.cancel()
        assert service._cancel_requested is True

        result = service.translate_file(excel_file_jp)

        # Translation completes because flag is reset at start
        assert result.status == TranslationStatus.COMPLETED

    def test_output_path_conflict_resolution(
        self, mock_copilot_with_translation, settings, excel_file_jp, tmp_path
    ):
        """Handle output path conflicts"""
        # Create existing output file
        existing_output = excel_file_jp.parent / "japanese_doc_translated.xlsx"
        existing_output.touch()

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(excel_file_jp)

        assert result.status == TranslationStatus.COMPLETED
        # Should create _translated_2.xlsx
        assert "_translated_2" in result.output_path.name


# --- Integration Tests: Font Handling ---

class TestFontHandlingWorkflow:
    """Test font conversion during translation"""

    def test_japanese_font_to_english(
        self, mock_copilot_with_translation, settings, tmp_path
    ):
        """Japanese fonts converted to English equivalents"""
        # Create file with Japanese fonts
        file_path = tmp_path / "jp_fonts.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "明朝体テキスト"
        ws["A1"].font = Font(name="MS Mincho", size=12)

        ws["A2"] = "ゴシック体テキスト"
        ws["A2"].font = Font(name="MS Gothic", size=12)

        wb.save(file_path)

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(file_path)

        assert result.status == TranslationStatus.COMPLETED

        # Check font conversions
        wb_out = openpyxl.load_workbook(result.output_path)
        ws_out = wb_out.active

        # Mincho -> Arial
        assert ws_out["A1"].font.name == "Arial"
        # Gothic -> Calibri
        assert ws_out["A2"].font.name == "Calibri"

    def test_font_size_adjustment(
        self, mock_copilot_with_translation, settings, tmp_path
    ):
        """Font size adjusted during translation"""
        file_path = tmp_path / "font_size.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "テスト"
        ws["A1"].font = Font(name="MS Gothic", size=14)

        wb.save(file_path)

        service = TranslationService(mock_copilot_with_translation, settings)

        result = service.translate_file(file_path)

        wb_out = openpyxl.load_workbook(result.output_path)
        ws_out = wb_out.active

        # Size reduced by 2 for JP to EN
        assert ws_out["A1"].font.size == 12


# --- Integration Tests: All File Types ---

class TestAllFileTypesWorkflow:
    """Test workflow with different file types"""

    def test_processor_selection_xlsx(self, mock_copilot_with_translation, settings):
        """Correct processor selected for xlsx"""
        service = TranslationService(mock_copilot_with_translation, settings)
        processor = service._get_processor(Path("test.xlsx"))
        assert isinstance(processor, ExcelProcessor)

    def test_processor_selection_docx(self, mock_copilot_with_translation, settings):
        """Correct processor selected for docx"""
        service = TranslationService(mock_copilot_with_translation, settings)
        processor = service._get_processor(Path("test.docx"))
        assert isinstance(processor, WordProcessor)

    def test_processor_selection_pptx(self, mock_copilot_with_translation, settings):
        """Correct processor selected for pptx"""
        service = TranslationService(mock_copilot_with_translation, settings)
        processor = service._get_processor(Path("test.pptx"))
        assert isinstance(processor, PptxProcessor)

    def test_unsupported_file_type(self, mock_copilot_with_translation, settings):
        """Unsupported file type raises error"""
        service = TranslationService(mock_copilot_with_translation, settings)

        with pytest.raises(ValueError) as exc:
            service._get_processor(Path("test.txt"))

        assert "Unsupported" in str(exc.value)
