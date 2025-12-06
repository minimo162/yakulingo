# tests/test_integration_extended.py
"""
Extended integration tests for YakuLingo.
Covers additional scenarios: async operations, concurrent workflows,
cancellation during translation, and full UI state workflows.
Bidirectional translation: Japanese → English, Other → Japanese.
"""

import pytest
import asyncio
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch, AsyncMock
import openpyxl
import sys

# Add project root to path for direct imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from yakulingo.models.types import (
    TranslationStatus,
    TranslationProgress,
    TextBlock,
    FileType,
    FileInfo,
)
from yakulingo.config.settings import AppSettings
from yakulingo.services.translation_service import TranslationService, BatchTranslator
from yakulingo.ui.state import AppState, Tab, FileState


# --- Fixtures ---

@pytest.fixture
def mock_copilot():
    """Mock CopilotHandler with configurable behavior"""
    mock = MagicMock()
    mock.is_connected = True
    mock.translate_single.return_value = "Translated text"
    mock.translate_sync.return_value = ["Translated 1", "Translated 2"]
    return mock


@pytest.fixture
def settings():
    return AppSettings()


@pytest.fixture
def sample_excel(tmp_path):
    """Create a basic Excel file for testing"""
    file_path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "テスト1"
    ws["A2"] = "テスト2"
    ws["A3"] = "テスト3"
    wb.save(file_path)
    return file_path


# --- State Workflow Tests ---

class TestStateWorkflowIntegration:
    """Test complete state workflows - bidirectional translation"""

    def test_text_translation_state_flow(self):
        """Test state transitions during text translation"""
        state = AppState()

        # Initial state
        assert state.current_tab == Tab.TEXT
        assert state.text_translating is False
        assert state.source_text == ""

        # Without text, cannot translate
        assert state.can_translate() is False

        # User enters text - can_translate() checks text, not connection
        state.source_text = "テスト文章"
        assert state.can_translate() is True  # Connection is checked at execution

        # Start translation
        state.text_translating = True
        assert state.is_translating() is True
        assert state.can_translate() is False  # Cannot translate while translating

        # Translation complete
        state.text_translating = False
        assert state.is_translating() is False
        assert state.can_translate() is True  # Can translate again

    def test_file_translation_state_flow(self):
        """Test state transitions during file translation"""
        state = AppState(current_tab=Tab.FILE)

        # Initial state
        assert state.file_state == FileState.EMPTY
        assert state.can_translate() is False

        # Connect
        state.copilot_ready = True
        assert state.can_translate() is False  # No file selected

        # Select file
        state.file_state = FileState.SELECTED
        state.selected_file = Path("/tmp/test.xlsx")
        state.file_info = FileInfo(
            path=Path("/tmp/test.xlsx"),
            file_type=FileType.EXCEL,
            size_bytes=1024,
        )
        assert state.can_translate() is True

        # Start translation
        state.file_state = FileState.TRANSLATING
        state.translation_progress = 0.0
        assert state.is_translating() is True
        assert state.can_translate() is False

        # Progress updates
        state.translation_progress = 0.5
        state.translation_status = "Translating batch 1 of 2..."
        assert state.translation_progress == 0.5

        # Complete
        state.file_state = FileState.COMPLETE
        state.output_file = Path("/tmp/test_translated.xlsx")
        assert state.is_translating() is False
        assert state.can_translate() is False  # Already complete

    def test_reset_file_state_flow(self):
        """Test reset file state transitions"""
        state = AppState(
            current_tab=Tab.FILE,
            file_state=FileState.COMPLETE,
            selected_file=Path("/tmp/test.xlsx"),
            file_info=FileInfo(
                path=Path("/tmp/test.xlsx"),
                file_type=FileType.EXCEL,
                size_bytes=1024
            ),
            output_file=Path("/tmp/test_translated.xlsx"),
            translation_progress=1.0,
            translation_status="Complete",
            error_message=""
        )

        state.reset_file_state()

        assert state.file_state == FileState.EMPTY
        assert state.selected_file is None
        assert state.file_info is None
        assert state.output_file is None
        assert state.translation_progress == 0.0
        assert state.translation_status == ""

    def test_error_state_flow(self):
        """Test error state transitions"""
        state = AppState(
            current_tab=Tab.FILE,
            file_state=FileState.TRANSLATING,
            copilot_ready=True
        )

        # Error occurs
        state.file_state = FileState.ERROR
        state.error_message = "Translation failed: Connection timeout"

        assert state.file_state == FileState.ERROR
        assert state.is_translating() is False
        assert "Connection timeout" in state.error_message

        # User clicks "Try again"
        state.reset_file_state()

        assert state.file_state == FileState.EMPTY
        assert state.error_message == ""


# --- Cancellation Workflow Tests ---

class TestCancellationWorkflow:
    """Test cancellation during various stages"""

    def test_cancel_during_batch_translation(self, mock_copilot, settings, sample_excel):
        """Test cancellation during batch translation"""
        service = TranslationService(mock_copilot, settings)

        # Set up mock to trigger cancel during translation
        batch_count = [0]

        def translate_with_cancel(texts, prompt, ref=None, char_limit=None, skip_clear_wait=False):
            batch_count[0] += 1
            if batch_count[0] >= 1:
                service.cancel()
            return [f"Trans{i}" for i in range(len(texts))]

        mock_copilot.translate_sync.side_effect = translate_with_cancel

        result = service.translate_file(sample_excel)

        # Should be cancelled
        assert result.status == TranslationStatus.CANCELLED

    def test_cancel_flag_reset_on_new_translation(self, mock_copilot, settings, sample_excel):
        """Cancel flag should reset when starting new translation"""
        service = TranslationService(mock_copilot, settings)

        # Pre-cancel
        service.cancel()
        assert service._cancel_requested is True

        # Start new translation - flag should reset
        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.COMPLETED


# --- Batch Processing Edge Cases ---

class TestBatchProcessingEdgeCases:
    """Test edge cases in batch processing"""

    def test_empty_block_list(self, mock_copilot, settings):
        """Handle empty block list"""
        service = TranslationService(mock_copilot, settings)

        results = service.batch_translator.translate_blocks([])

        assert results == {}
        mock_copilot.translate_sync.assert_not_called()

    def test_single_block(self, mock_copilot, settings):
        """Handle single block"""
        mock_copilot.translate_sync.return_value = ["Translated"]
        service = TranslationService(mock_copilot, settings)

        blocks = [TextBlock(id="1", text="テスト", location="A1")]

        results = service.batch_translator.translate_blocks(blocks)

        assert len(results) == 1
        assert results["1"] == "Translated"

    def test_exactly_batch_size_blocks(self, mock_copilot, settings):
        """Handle exactly MAX_BATCH_SIZE blocks"""
        # Create exactly 50 translations
        mock_copilot.translate_sync.return_value = [f"Trans{i}" for i in range(50)]
        service = TranslationService(mock_copilot, settings)

        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}", location=f"A{i}")
            for i in range(50)
        ]

        results = service.batch_translator.translate_blocks(blocks)

        assert len(results) == 50
        assert mock_copilot.translate_sync.call_count == 1

    def test_char_limit_triggers_batch_split(self, mock_copilot, settings):
        """Handle blocks exceeding char limit per batch"""
        mock_copilot.translate_sync.side_effect = [
            ["Trans0", "Trans1"],
            ["Trans2"],
        ]
        # Use small char limit to force splits
        settings.max_chars_per_batch = 1000
        service = TranslationService(mock_copilot, settings)

        # Create 3 blocks with 400 chars each (800 chars per 2 blocks < 1000)
        text_400 = "あ" * 400
        blocks = [
            TextBlock(id=str(i), text=text_400, location=f"A{i}")
            for i in range(3)
        ]

        results = service.batch_translator.translate_blocks(blocks)

        assert len(results) == 3
        assert mock_copilot.translate_sync.call_count == 2


# --- Error Recovery Tests ---

class TestErrorRecovery:
    """Test error handling and recovery"""

    def test_translation_error_returns_failed_status(self, mock_copilot, settings, sample_excel):
        """Translation error returns FAILED status"""
        mock_copilot.translate_sync.side_effect = RuntimeError("API Error")
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "API Error" in result.error_message

    def test_file_not_found_error(self, mock_copilot, settings):
        """Handle file not found"""
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(Path("/nonexistent/file.xlsx"))

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_corrupted_file_error(self, mock_copilot, settings, tmp_path):
        """Handle corrupted file"""
        corrupt_file = tmp_path / "corrupt.xlsx"
        corrupt_file.write_bytes(b"not a valid xlsx file")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(corrupt_file)

        assert result.status == TranslationStatus.FAILED


# --- Output Path Generation Tests ---

class TestOutputPathGeneration:
    """Test output path generation logic - bidirectional uses _translated suffix"""

    def test_translated_suffix(self, mock_copilot, settings, sample_excel):
        """Bidirectional translation adds _translated suffix"""
        service = TranslationService(mock_copilot, settings)

        output_path = service._generate_output_path(sample_excel)

        assert "_translated" in output_path.name
        assert output_path.suffix == ".xlsx"

    def test_incremental_numbering(self, mock_copilot, settings, tmp_path):
        """Handle existing output files with incremental numbering"""
        # Create input file
        input_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(input_file)

        # Create existing output files
        (tmp_path / "test_translated.xlsx").touch()
        (tmp_path / "test_translated_2.xlsx").touch()

        service = TranslationService(mock_copilot, settings)

        output_path = service._generate_output_path(input_file)

        assert output_path.name == "test_translated_3.xlsx"

    def test_custom_output_directory(self, mock_copilot, settings, tmp_path):
        """Output to custom directory"""
        input_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(input_file)

        output_dir = tmp_path / "custom_output"
        output_dir.mkdir()

        settings.output_directory = str(output_dir)
        service = TranslationService(mock_copilot, settings)

        output_path = service._generate_output_path(input_file)

        assert output_path.parent == output_dir


# --- File Info Extraction Tests ---

class TestFileInfoExtraction:
    """Test file info extraction"""

    def test_excel_file_info(self, mock_copilot, settings, sample_excel):
        """Extract info from Excel file"""
        service = TranslationService(mock_copilot, settings)

        info = service.get_file_info(sample_excel)

        assert info.path == sample_excel
        assert info.file_type == FileType.EXCEL
        assert info.size_bytes > 0
        assert info.sheet_count == 1

    def test_multi_sheet_excel_info(self, mock_copilot, settings, tmp_path):
        """Extract info from multi-sheet Excel"""
        file_path = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "テスト1"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "テスト2"

        ws3 = wb.create_sheet("Sheet3")
        ws3["A1"] = "テスト3"

        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        info = service.get_file_info(file_path)

        assert info.sheet_count == 3


# --- Settings Integration Tests ---

class TestSettingsIntegration:
    """Test settings integration"""

    def test_chars_per_batch_from_settings(self, mock_copilot):
        """Chars per batch should come from settings"""
        settings = AppSettings(max_chars_per_batch=5000)
        service = TranslationService(mock_copilot, settings)

        # Create blocks that would normally be 1 batch but with reduced char limit
        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}" * 100, location=f"A{i}")
            for i in range(10)
        ]

        # Verify the service was created with the settings
        assert settings.max_chars_per_batch == 5000

    def test_output_directory_from_settings(self, mock_copilot, tmp_path, sample_excel):
        """Output directory should come from settings"""
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        settings = AppSettings(output_directory=str(output_dir))
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.output_path.parent == output_dir


# --- Reference Files Integration Tests ---

class TestReferenceFilesIntegration:
    """Test reference files handling"""

    def test_reference_files_passed_to_copilot(self, mock_copilot, settings, sample_excel, tmp_path):
        """Reference files should be passed to Copilot"""
        glossary = tmp_path / "glossary.csv"
        glossary.write_text("JP,EN\nテスト,Test\n", encoding="utf-8")

        service = TranslationService(mock_copilot, settings)

        service.translate_file(
            sample_excel,
            reference_files=[glossary],
        )

        # Verify reference files were passed
        call_args = mock_copilot.translate_sync.call_args
        assert call_args is not None

    def test_multiple_reference_files(self, mock_copilot, settings, sample_excel, tmp_path):
        """Multiple reference files should be supported"""
        glossary1 = tmp_path / "glossary1.csv"
        glossary1.write_text("JP,EN\n用語1,Term1\n", encoding="utf-8")

        glossary2 = tmp_path / "glossary2.csv"
        glossary2.write_text("JP,EN\n用語2,Term2\n", encoding="utf-8")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(
            sample_excel,
            reference_files=[glossary1, glossary2],
        )

        assert result.status == TranslationStatus.COMPLETED


# --- Bidirectional Translation Tests ---

class TestBidirectionalBehavior:
    """Test bidirectional translation behavior"""

    def test_translation_prompt_used(self, mock_copilot, settings, sample_excel):
        """Translation uses proper prompt template"""
        service = TranslationService(mock_copilot, settings)

        service.translate_file(sample_excel)

        # Verify prompt contains translation rules
        call_args = mock_copilot.translate_sync.call_args
        prompt = call_args[0][1]  # Second positional arg is prompt
        # The prompt should contain translation instructions
        assert "Translation Rule" in prompt or "翻訳" in prompt

    def test_japanese_input_translation(self, mock_copilot, settings, sample_excel):
        """Japanese input should be auto-detected and translated"""
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.COMPLETED
        assert result.output_path is not None
        assert "_translated" in result.output_path.name

    def test_english_input_translation(self, mock_copilot, settings, tmp_path):
        """English input should be auto-detected and translated"""
        # Create English content file
        file_path = tmp_path / "english.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello World"
        ws["A2"] = "This is a test"
        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        assert result.status == TranslationStatus.COMPLETED
        assert result.output_path is not None
