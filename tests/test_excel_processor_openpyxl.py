from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import yakulingo.processors.excel_processor as excel_processor_module
from yakulingo.processors.excel_processor import ExcelProcessor
from yakulingo.processors.translators import CellTranslator


def _build_sample_xlsx(path: Path) -> None:
    j_hello = "\u3053\u3093\u306b\u3061\u306f"  # こんにちは
    j_thanks = "\u3042\u308a\u304c\u3068\u3046"  # ありがとう
    j_sales = "\u58f2\u4e0a\u3052"  # 売上げ (contains kana)
    j_kanji_only = "\u65e5\u672c\u8a9e"  # 日本語 (kanji-only)
    c_chinese = "\u4f60\u597d\u4e16\u754c"  # 你好世界
    j_mixed = f"Hello {j_hello}"
    jp_month = f"12\u6708"  # 12月
    yen_full = "\uFFE5"  # ￥ (FULLWIDTH YEN SIGN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = j_hello
    ws["B1"] = "35,555"
    ws["C1"] = f'=CONCAT("{j_hello}","\u4e16\u754c")'  # 世界
    ws["D1"] = "USA"
    ws["E1"] = "2025-12-25"
    ws["F1"] = "test@example.com"
    ws["G1"] = "https://example.com"
    ws["H1"] = "ABC-123"
    ws["I1"] = j_kanji_only
    ws["J1"] = c_chinese
    ws["K1"] = j_thanks
    ws["L1"] = j_mixed
    ws["M1"] = "50%"
    ws["O1"] = yen_full + "1,234"

    ws["A2"] = j_sales
    ws["B2"] = "Hello"
    ws["C2"] = "Hello 123"
    ws["D2"] = jp_month

    ws.merge_cells("A4:C4")
    ws["A4"] = "\u30de\u30fc\u30b8\u30c6\u30b9\u30c8"  # マージテスト

    ws2 = wb.create_sheet("my_sheet")
    ws2["A1"] = j_hello
    ws2["B1"] = "Hello"

    wb.save(path)
    wb.close()


@pytest.mark.unit
def test_celltranslator_skips_fullwidth_currency_symbols() -> None:
    translator = CellTranslator()

    # FULLWIDTH YEN SIGN should be skipped for EN→JP too.
    assert translator.should_translate("\uFFE51,234", "jp") is False
    assert translator.should_translate("\uFFE51,234", "en") is False

    # FULLWIDTH DOLLAR SIGN / FULLWIDTH POUND SIGN should also be skipped.
    assert translator.should_translate("\uFF041,234", "jp") is False
    assert translator.should_translate("\uFFE11,234", "jp") is False


@pytest.mark.unit
def test_celltranslator_translates_year_fragments() -> None:
    translator = CellTranslator()

    assert translator.should_translate("6\u5E74", "en") is True
    assert translator.should_translate("\uFF16\u5E74", "en") is True
    assert translator.should_translate("12\u6708", "en") is True


@pytest.mark.unit
def test_excelprocessor_openpyxl_extract_and_apply(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    # Force openpyxl fallback to avoid requiring Microsoft Excel in test environments.
    monkeypatch.setattr(excel_processor_module, "_can_use_xlwings", lambda: False)

    src = tmp_path / "sample.xlsx"
    _build_sample_xlsx(src)

    processor = ExcelProcessor()

    # JP→EN extraction should pick up only cells containing Japanese/CJK text (and skip formulas/numbers/etc.)
    blocks_en = list(processor.extract_text_blocks(src, output_language="en"))
    ids_en = {b.id for b in blocks_en}
    assert ids_en == {
        "Sheet1_A1",
        "Sheet1_A2",
        "Sheet1_A4",
        "Sheet1_D2",
        "Sheet1_I1",
        "Sheet1_J1",
        "Sheet1_K1",
        "Sheet1_L1",
        "my_sheet_A1",
    }

    out_en = tmp_path / "translated_en.xlsx"
    translations_en = {b.id: "EN_" + b.text for b in blocks_en}
    processor.apply_translations(src, out_en, translations_en, direction="jp_to_en", text_blocks=blocks_en)

    wb_en = openpyxl.load_workbook(out_en, data_only=False)
    ws_en = wb_en["Sheet1"]
    assert ws_en["A1"].value.startswith("EN_")
    assert ws_en["B1"].value == "35,555"
    assert isinstance(ws_en["C1"].value, str) and ws_en["C1"].value.startswith("=")  # formula preserved
    assert ws_en["O1"].value == "\uFFE51,234"  # currency preserved (skipped)
    assert ws_en["D2"].value.startswith("EN_")
    assert ws_en["A1"].font.name == "Arial"  # default JP→EN output font
    wb_en.close()

    # EN→JP extraction should include English/Chinese/kanji-only, but skip Japanese-only (kana) and currency.
    blocks_jp = list(processor.extract_text_blocks(src, output_language="jp"))
    ids_jp = {b.id for b in blocks_jp}
    assert ids_jp == {
        "Sheet1_B2",
        "Sheet1_C2",
        "Sheet1_D1",
        "Sheet1_I1",
        "Sheet1_J1",
        "Sheet1_L1",
        "my_sheet_B1",
    }

    out_jp = tmp_path / "translated_jp.xlsx"
    translations_jp = {b.id: "JP_" + b.text for b in blocks_jp}
    processor.apply_translations(src, out_jp, translations_jp, direction="en_to_jp", text_blocks=blocks_jp)

    wb_jp = openpyxl.load_workbook(out_jp, data_only=False)
    ws_jp = wb_jp["Sheet1"]
    assert ws_jp["B2"].value == "JP_Hello"
    assert ws_jp["D1"].value == "JP_USA"
    assert ws_jp["A1"].value == "\u3053\u3093\u306b\u3061\u306f"  # Japanese left untouched
    assert ws_jp["O1"].value == "\uFFE51,234"  # currency preserved (skipped)
    wb_jp.close()
