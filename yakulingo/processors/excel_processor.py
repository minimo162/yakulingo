# yakulingo/processors/excel_processor.py
"""
Processor for Excel files (.xlsx, .xls).

Uses xlwings for full Excel functionality (shapes, charts, textboxes).
Falls back to openpyxl if xlwings/Excel is not available (Linux or no Excel installed).
"""

import gc
import logging
import re
import sys
import threading
import time
import zipfile
from functools import lru_cache
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator, Optional
from xml.etree import ElementTree as ET

from .base import FileProcessor
from .translators import CellTranslator
from .font_manager import FontManager
from yakulingo.models.types import TextBlock, FileInfo, FileType, SectionDetail

# Module logger
logger = logging.getLogger(__name__)


# =============================================================================
# COM initialization for Windows (required for xlwings in worker threads)
# =============================================================================
_pythoncom = None
_pywintypes = None

def _get_pythoncom():
    """Lazy import pythoncom (Windows COM library)."""
    global _pythoncom
    if _pythoncom is None and sys.platform == 'win32':
        try:
            import pythoncom
            _pythoncom = pythoncom
        except ImportError:
            logger.debug("pythoncom not available")
    return _pythoncom


def _get_pywintypes():
    """Lazy import pywintypes (Windows COM error types)."""
    global _pywintypes
    if _pywintypes is None and sys.platform == 'win32':
        try:
            import pywintypes
            _pywintypes = pywintypes
        except ImportError:
            logger.debug("pywintypes not available")
    return _pywintypes


@contextmanager
def com_initialized():
    """
    Context manager to ensure COM is initialized for the current thread.

    Required when xlwings is called from a worker thread (e.g., asyncio.to_thread).
    On non-Windows platforms, this is a no-op.

    Usage:
        with com_initialized():
            app = xw.App(visible=False)
            ...
    """
    pythoncom = _get_pythoncom()
    initialized = False

    if pythoncom is not None:
        try:
            # Try CoInitializeEx with COINIT_APARTMENTTHREADED (STA) first
            # This is the default mode and required by most COM servers including Excel
            # If already initialized, it returns RPC_E_CHANGED_MODE (0x80010106)
            thread_id = threading.current_thread().ident
            try:
                hr = pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
                # S_OK (0) = newly initialized, S_FALSE (1) = already initialized
                # Some pywin32 versions may return None on success
                # Only set initialized=True for S_OK (0) or None to ensure proper cleanup
                initialized = (hr == 0 or hr is None)
                logger.debug("COM initialized (STA) thread=%s hr=%s will_uninit=%s", thread_id, hr, initialized)
            except Exception:
                # Fall back to simple CoInitialize if CoInitializeEx fails
                hr = pythoncom.CoInitialize()
                initialized = (hr == 0 or hr is None)
                logger.debug("COM initialized (fallback) thread=%s hr=%s will_uninit=%s", thread_id, hr, initialized)
        except Exception as e:
            logger.debug("COM initialization skipped: %s", e)

    try:
        yield
    finally:
        if initialized and pythoncom is not None:
            try:
                thread_id = threading.current_thread().ident
                pythoncom.CoUninitialize()
                logger.debug("COM uninitialized thread=%s", thread_id)
            except Exception as e:
                logger.debug("COM uninitialization failed: %s", e)


# =============================================================================
# xlwings detection (requires Excel on Windows/macOS)
# =============================================================================
_xlwings = None
HAS_XLWINGS = False
_EXCEL_COM_REGISTERED: bool | None = None

def _get_xlwings():
    """Lazy import xlwings."""
    global _xlwings, HAS_XLWINGS
    if _xlwings is None:
        try:
            import xlwings as xw
            _xlwings = xw
            HAS_XLWINGS = True
        except ImportError:
            HAS_XLWINGS = False
    return _xlwings


def is_xlwings_available() -> bool:
    """Check if xlwings is available."""
    _get_xlwings()
    return HAS_XLWINGS


def _is_excel_com_registered() -> bool:
    """Return True if Excel's COM ProgID looks registered on Windows."""
    global _EXCEL_COM_REGISTERED
    if sys.platform != "win32":
        return True  # Not applicable (xlwings handles Excel availability on other platforms)
    if _EXCEL_COM_REGISTERED is not None:
        return _EXCEL_COM_REGISTERED

    try:
        import winreg

        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\\CLSID") as key:
            clsid, _ = winreg.QueryValueEx(key, None)
            _EXCEL_COM_REGISTERED = bool(clsid)
    except Exception:
        _EXCEL_COM_REGISTERED = False

    return _EXCEL_COM_REGISTERED


def _can_use_xlwings() -> bool:
    """Return True when xlwings can realistically drive Excel for this runtime."""
    _get_xlwings()
    return HAS_XLWINGS and _is_excel_com_registered()


# Retry settings for Excel COM server connection
_EXCEL_RETRY_COUNT = 3
_EXCEL_RETRY_DELAY = 1.0  # seconds

# Excel cell character limit (32,767 characters per cell)
EXCEL_CELL_CHAR_LIMIT = 32767

# LRU cache size for column letter conversions (avoid memory bloat on wide sheets)
_COLUMN_LETTER_CACHE_SIZE = 1000


def _cleanup_com_before_retry():
    """
    Clean up COM resources before retrying Excel connection.

    This helps resolve CO_E_SERVER_EXEC_FAILURE errors by:
    1. Pumping pending COM messages (Windows only)
    2. Running garbage collection to release COM objects (single pass for performance)
    """
    pythoncom = _get_pythoncom()
    if pythoncom is not None:
        try:
            # Pump any waiting COM messages to allow cleanup to complete
            pythoncom.PumpWaitingMessages()
        except Exception as e:
            logger.debug("PumpWaitingMessages failed: %s", e)

    # Single gc pass after message pump (removed duplicate call for performance)
    gc.collect()


# COM error messages that indicate recoverable errors
_RECOVERABLE_COM_ERROR_MESSAGES = [
    "サーバーの実行に失敗しました",  # Server execution failed
    "RPC server",
    "RPC サーバー",
    "オートメーション",  # Automation error
    "Call was rejected",  # COM call rejected
    "-2147418111",  # RPC_E_CALL_REJECTED
    "-2146959355",  # CO_E_SERVER_EXEC_FAILURE
]

_NON_RECOVERABLE_COM_ERROR_HRESULTS = {
    -2147221005,  # CO_E_CLASSSTRING: invalid class string (Excel not registered)
    -2147221164,  # REGDB_E_CLASSNOTREG: class not registered
}

_NON_RECOVERABLE_COM_ERROR_MESSAGES = [
    "クラス文字列が無効です",
    "Class string is invalid",
    "Invalid class string",
    "クラスが登録されていません",
    "Class not registered",
]


def _is_recoverable_com_error(e: Exception) -> bool:
    """
    Check if an exception is a recoverable COM error.

    Args:
        e: The exception to check

    Returns:
        True if the error is a COM error that might be recoverable with retry
    """
    pywintypes = _get_pywintypes()
    error_str = str(e)

    # Check if this is a pywintypes.com_error
    is_com_error = False
    if pywintypes is not None:
        is_com_error = isinstance(e, pywintypes.com_error)
        if is_com_error:
            try:
                hresult = e.args[0]
            except Exception:
                hresult = None
            if hresult in _NON_RECOVERABLE_COM_ERROR_HRESULTS:
                return False
            if any(msg in error_str for msg in _NON_RECOVERABLE_COM_ERROR_MESSAGES):
                return False

    # Also check for common COM error messages (works on Japanese/English Windows)
    return is_com_error or any(err_msg in error_str for err_msg in _RECOVERABLE_COM_ERROR_MESSAGES)


def _try_create_new_excel_instance(xw, max_attempts: int = 3):
    """
    Try to create a new isolated Excel instance using win32com.DispatchEx.

    Unlike xw.App() which may connect to an existing Excel instance via ROT,
    DispatchEx always creates a new process.

    Args:
        xw: xlwings module
        max_attempts: Maximum number of creation attempts

    Returns:
        xlwings App instance with no pre-existing books, or None on failure

    Note:
        This function does NOT call quit() on any Excel instance to avoid
        closing the user's manually opened Excel files.
    """
    try:
        import win32com.client
    except ImportError:
        logger.warning("win32com not available - falling back to xw.App()")
        # Fall back to simple retry with xw.App()
        for attempt in range(max_attempts):
            # Note: _cleanup_com_before_retry() already calls gc.collect() once
            _cleanup_com_before_retry()
            time.sleep(0.3 * (attempt + 1))
            try:
                app = xw.App(visible=False, add_book=False)
                if len(app.books) == 0:
                    return app
                del app
            except Exception:
                pass
        return None

    for attempt in range(max_attempts):
        # Note: _cleanup_com_before_retry() already calls gc.collect() once
        _cleanup_com_before_retry()
        time.sleep(0.2 * (attempt + 1))

        excel_com = None
        try:
            # DispatchEx ALWAYS creates a new Excel process (not via ROT)
            excel_com = win32com.client.DispatchEx("Excel.Application")
            excel_com.Visible = False
            excel_com.DisplayAlerts = False

            # Get the Hwnd of this new Excel process for identification
            new_hwnd = excel_com.Hwnd
            logger.debug("Created new Excel via DispatchEx (Hwnd=%s)", new_hwnd)

            # Find this instance in xlwings by matching Hwnd
            # This is more reliable than xw.App() which may connect to a different instance
            target_app = None
            for xw_app in xw.apps:
                try:
                    # xlwings App has .hwnd property that matches Excel.Application.Hwnd
                    if hasattr(xw_app, 'hwnd') and xw_app.hwnd == new_hwnd:
                        target_app = xw_app
                        logger.debug("Found xlwings App matching Hwnd=%s", new_hwnd)
                        break
                except Exception:
                    continue

            if target_app is not None:
                # Verify this instance has no pre-existing books
                if len(target_app.books) == 0:
                    logger.info(
                        "Created isolated Excel instance via DispatchEx (PID=%s, Hwnd=%s)",
                        target_app.pid, new_hwnd
                    )
                    return target_app
                else:
                    logger.warning(
                        "DispatchEx instance has books (count=%d), this is unexpected",
                        len(target_app.books)
                    )
                    # Clean up this unexpected instance
                    try:
                        excel_com.Quit()
                    except Exception:
                        pass
                    excel_com = None
                    continue  # Retry

            # Hwnd not found in xw.apps - wait for xlwings to register the instance
            # xlwings monitors ROT asynchronously, may need a short wait
            logger.debug("Hwnd %s not found in xw.apps, waiting for registration...", new_hwnd)
            for wait_attempt in range(5):
                time.sleep(0.1)
                for xw_app in xw.apps:
                    try:
                        if hasattr(xw_app, 'hwnd') and xw_app.hwnd == new_hwnd:
                            if len(xw_app.books) == 0:
                                logger.info(
                                    "Created isolated Excel instance via DispatchEx after wait "
                                    "(PID=%s, Hwnd=%s, wait_attempts=%d)",
                                    xw_app.pid, new_hwnd, wait_attempt + 1
                                )
                                return xw_app
                            break
                    except Exception:
                        continue

            # Still not found - clean up and retry
            logger.debug(
                "Attempt %d: Could not find DispatchEx instance (Hwnd=%s) in xlwings",
                attempt + 1, new_hwnd
            )
            try:
                excel_com.Quit()
            except Exception:
                pass
            excel_com = None

        except Exception as e:
            logger.debug("Attempt %d failed: %s", attempt + 1, e)
            # Clean up on error
            if excel_com is not None:
                try:
                    excel_com.Quit()
                except Exception:
                    pass

    return None


def _copy_sheet_with_retry(
    source_sheet,
    target_wb,
    position: str,
    ref_sheet_index: int,
    new_name: str,
    max_retries: int = _EXCEL_RETRY_COUNT,
    retry_delay: float = _EXCEL_RETRY_DELAY,
) -> None:
    """
    Copy a sheet to target workbook with retry logic for RPC errors.

    Args:
        source_sheet: The xlwings sheet to copy
        target_wb: The target xlwings workbook
        position: "Before" or "After" the reference sheet
        ref_sheet_index: Index of the reference sheet in target_wb
        new_name: Name to give the copied sheet
        max_retries: Maximum retry attempts
        retry_delay: Initial retry delay (exponential backoff)

    Raises:
        Exception: If all retries fail
    """
    last_error = None

    for attempt in range(max_retries):
        try:
            ref_sheet = target_wb.sheets[ref_sheet_index]
            if position == "Before":
                source_sheet.api.Copy(Before=ref_sheet.api)
                # Copied sheet is inserted at ref_sheet_index
                target_wb.sheets[ref_sheet_index].name = new_name
            else:  # "After"
                source_sheet.api.Copy(After=ref_sheet.api)
                # Copied sheet is inserted at ref_sheet_index + 1
                target_wb.sheets[ref_sheet_index + 1].name = new_name
            return  # Success
        except Exception as e:
            last_error = e
            error_str = str(e)

            if _is_recoverable_com_error(e) and attempt < max_retries - 1:
                delay = retry_delay * (2 ** attempt)
                logger.warning(
                    "Sheet copy RPC error (attempt %d/%d) for '%s': %s. Retrying in %.1fs...",
                    attempt + 1, max_retries, new_name, error_str, delay
                )
                _cleanup_com_before_retry()
                time.sleep(delay)
            else:
                logger.error(
                    "Sheet copy failed (attempt %d/%d) for '%s': %s",
                    attempt + 1, max_retries, new_name, error_str
                )
                raise

    if last_error:
        raise last_error


def _create_excel_app_with_retry(xw, max_retries: int = _EXCEL_RETRY_COUNT, retry_delay: float = _EXCEL_RETRY_DELAY):
    """
    Create xlwings App with retry logic.

    Handles COM server errors like "サーバーの実行に失敗しました" (Server execution failed)
    by retrying with exponential backoff and COM cleanup.

    Args:
        xw: xlwings module
        max_retries: Maximum number of retry attempts
        retry_delay: Initial delay between retries (doubled each attempt)

    Returns:
        xlwings App instance

    Raises:
        Exception: If all retry attempts fail
    """
    pywintypes = _get_pywintypes()
    last_error = None

    # Pre-emptive cleanup before first attempt to avoid COM server errors
    # caused by lingering COM objects from previous sessions
    _cleanup_com_before_retry()

    for attempt in range(max_retries):
        try:
            app = xw.App(visible=False, add_book=False)
            # Verify this is a truly isolated instance (no pre-existing books)
            if len(app.books) > 0:
                existing_books = len(app.books)
                existing_pid = app.pid
                # CRITICAL: Do NOT call app.quit() here - it would close user's Excel!
                # Instead, just release this reference and try to create a new instance
                logger.warning(
                    "xlwings connected to existing Excel instance with %d books (PID=%s). "
                    "Releasing reference and creating new instance...",
                    existing_books, existing_pid
                )
                # Release the reference without closing the app
                del app
                # Force COM to release objects
                gc.collect()
                _cleanup_com_before_retry()
                time.sleep(0.5)
                # Try creating a new instance - if we still get the same one,
                # we need to use win32com to force a new process
                app = xw.App(visible=False, add_book=False)
                if len(app.books) > 0:
                    # Still connected to existing instance - try multiple attempts
                    logger.warning(
                        "Still connected to existing Excel (PID=%s). Attempting isolation...",
                        app.pid
                    )
                    del app
                    gc.collect()
                    # Try to create a new isolated instance
                    app = _try_create_new_excel_instance(xw)
                    if app is None:
                        raise RuntimeError(
                            "既存のExcelインスタンスから分離できませんでした。\n"
                            "開いているExcelファイルをすべて閉じてから再試行してください。"
                        )
            logger.debug("Created isolated Excel instance (PID=%s, books=%d)", app.pid, len(app.books))
            return app
        except Exception as e:
            last_error = e
            error_str = str(e)

            if _is_recoverable_com_error(e) and attempt < max_retries - 1:
                delay = retry_delay * (2 ** attempt)  # Exponential backoff
                logger.warning(
                    "Excel COM error (attempt %d/%d): %s. Retrying in %.1fs...",
                    attempt + 1, max_retries, error_str, delay
                )

                # Clean up COM resources before retry
                _cleanup_com_before_retry()

                time.sleep(delay)
            else:
                # Non-recoverable error or last attempt
                logger.error(
                    "Excel COM error (final attempt %d/%d): %s",
                    attempt + 1, max_retries, error_str
                )
                raise

    # All retries exhausted (shouldn't reach here)
    if last_error:
        raise last_error


def _verify_workbook_path(wb, expected_path: Path, operation: str = "open") -> None:
    """Verify that the opened workbook matches the expected path.

    This is a critical safety check to prevent operating on wrong files
    if xlwings accidentally connected to an existing Excel instance.

    Args:
        wb: xlwings Workbook object
        expected_path: Expected file path
        operation: Description of operation for error message

    Raises:
        RuntimeError: If paths don't match
    """
    try:
        opened_path = Path(wb.fullname).resolve()
        expected_resolved = Path(expected_path).resolve()

        if opened_path != expected_resolved:
            logger.error(
                "SAFETY: Workbook path mismatch during %s! Expected: %s, Got: %s. "
                "This may indicate xlwings connected to wrong Excel instance.",
                operation, expected_resolved, opened_path
            )
            raise RuntimeError(
                f"ワークブックパスの不一致を検出しました（{operation}）。"
                f"期待: {expected_resolved}, 実際: {opened_path}。"
                f"Excelが他のファイルを開いている可能性があります。"
            )
        logger.debug("Verified workbook path (%s): %s", operation, opened_path)
    except AttributeError:
        # wb.fullname not available (shouldn't happen with xlwings)
        logger.warning("Could not verify workbook path - fullname attribute not available")


# Excel sheet name forbidden characters: \ / ? * [ ] :
# Maximum length: 31 characters
_EXCEL_SHEET_NAME_FORBIDDEN = re.compile(r'[\\/?*\[\]:]')
_EXCEL_SHEET_NAME_MAX_LENGTH = 31


def sanitize_sheet_name(name: str, max_length: int = _EXCEL_SHEET_NAME_MAX_LENGTH) -> str:
    """
    Sanitize a string to be used as an Excel sheet name.

    Excel sheet names cannot contain: \\ / ? * [ ] :
    Maximum length is 31 characters.

    Args:
        name: The sheet name to sanitize
        max_length: Maximum allowed length (default 31)

    Returns:
        Sanitized sheet name safe for Excel
    """
    # Replace forbidden characters with underscore
    sanitized = _EXCEL_SHEET_NAME_FORBIDDEN.sub('_', name)

    # Truncate if too long
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length - 3] + '...'

    return sanitized


def _ensure_unique_sheet_name(name: str, existing_names: set[str]) -> str:
    """Ensure sheet name is unique within the workbook while respecting Excel limits."""

    # If name already sanitized and unique, return as-is
    if name not in existing_names:
        existing_names.add(name)
        return name

    base_name = name
    counter = 1

    # Append incremental suffixes until unique (truncating to 31 chars if needed)
    while True:
        suffix = f"_{counter}"
        max_base_length = _EXCEL_SHEET_NAME_MAX_LENGTH - len(suffix)
        truncated_base = base_name[:max_base_length]
        candidate = f"{truncated_base}{suffix}"

        if candidate not in existing_names:
            existing_names.add(candidate)
            return candidate

        counter += 1


# =============================================================================
# openpyxl fallback
# =============================================================================
import openpyxl
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.styles import Font


def _detect_formula_cells_via_zipfile(file_path: Path) -> set[tuple[str, int, int]]:
    """
    Detect formula cells by parsing the XLSX file directly via zipfile.

    This is much more memory-efficient than loading the workbook with openpyxl,
    as it only reads and parses the necessary XML portions.

    XLSX files are ZIP archives containing XML files:
    - xl/workbook.xml: Contains sheet definitions
    - xl/worksheets/sheet1.xml, sheet2.xml, etc.: Contains cell data

    Formula cells are identified by the presence of <f> elements within <c> (cell) elements.

    Args:
        file_path: Path to the XLSX file

    Returns:
        Set of (sheet_name, row, col) tuples for formula cells
    """
    formula_cells: set[tuple[str, int, int]] = set()

    try:
        with zipfile.ZipFile(file_path, 'r') as xlsx:
            # Parse workbook.xml to get sheet name -> relationship mapping
            sheet_names: dict[str, str] = {}  # rId -> sheet_name
            sheet_order: list[str] = []  # Ordered list of rIds

            try:
                with xlsx.open('xl/workbook.xml') as workbook_xml:
                    # Parse XML incrementally to reduce memory
                    for event, elem in ET.iterparse(workbook_xml, events=['end']):
                        # Sheet elements: <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
                        if elem.tag.endswith('}sheet') or elem.tag == 'sheet':
                            sheet_name = elem.get('name', '')
                            # Get r:id attribute (namespace might vary)
                            rid = None
                            for attr_name, attr_value in elem.attrib.items():
                                if attr_name.endswith('}id') or attr_name == 'id':
                                    rid = attr_value
                                    break
                            if rid and sheet_name:
                                sheet_names[rid] = sheet_name
                                sheet_order.append(rid)
                            elem.clear()  # Free memory
            except KeyError:
                logger.debug("workbook.xml not found in XLSX")
                return formula_cells

            # Parse relationships to map rId to sheet file paths
            rid_to_file: dict[str, str] = {}
            try:
                with xlsx.open('xl/_rels/workbook.xml.rels') as rels_xml:
                    for event, elem in ET.iterparse(rels_xml, events=['end']):
                        if elem.tag.endswith('}Relationship') or elem.tag == 'Relationship':
                            rid = elem.get('Id', '')
                            target = elem.get('Target', '')
                            if rid and target:
                                # Target can be:
                                # - Relative: "worksheets/sheet1.xml"
                                # - Absolute from xl/: "/xl/worksheets/sheet1.xml"
                                # Normalize to full path within ZIP
                                target = target.lstrip('/')
                                if target.startswith('xl/'):
                                    rid_to_file[rid] = target
                                else:
                                    rid_to_file[rid] = f"xl/{target}"
                            elem.clear()
            except KeyError:
                logger.debug("workbook.xml.rels not found in XLSX")
                # Fall back to sequential sheet file names
                for i, rid in enumerate(sheet_order, 1):
                    rid_to_file[rid] = f"xl/worksheets/sheet{i}.xml"

            # Parse each sheet XML to find formula cells
            cell_ref_pattern = re.compile(r'^([A-Z]+)(\d+)$')

            for rid in sheet_order:
                sheet_name = sheet_names.get(rid, '')
                sheet_file = rid_to_file.get(rid, '')
                if not sheet_name or not sheet_file:
                    continue

                try:
                    with xlsx.open(sheet_file) as sheet_xml:
                        current_row = 0
                        for event, elem in ET.iterparse(sheet_xml, events=['end']):
                            # Row element: <row r="1">
                            if elem.tag.endswith('}row') or elem.tag == 'row':
                                current_row = int(elem.get('r', 0))
                                elem.clear()
                            # Cell element: <c r="A1"><f>SUM(B1:B10)</f><v>100</v></c>
                            elif elem.tag.endswith('}c') or elem.tag == 'c':
                                # Check if cell has a formula child element
                                has_formula = False
                                for child in elem:
                                    if child.tag.endswith('}f') or child.tag == 'f':
                                        has_formula = True
                                        break

                                if has_formula:
                                    cell_ref = elem.get('r', '')
                                    match = cell_ref_pattern.match(cell_ref)
                                    if match:
                                        col_letter = match.group(1)
                                        row_num = int(match.group(2))
                                        try:
                                            col_idx = column_index_from_string(col_letter)
                                            formula_cells.add((sheet_name, row_num, col_idx))
                                        except (ValueError, TypeError):
                                            pass
                                elem.clear()
                except KeyError:
                    logger.debug("Sheet file not found: %s", sheet_file)
                    continue

    except zipfile.BadZipFile:
        logger.warning("Invalid XLSX file (not a valid ZIP): %s", file_path)
    except Exception as e:
        logger.warning("Error detecting formula cells via zipfile: %s", e)

    return formula_cells


class ExcelProcessor(FileProcessor):
    """
    Processor for Excel files (.xlsx, .xls).

    Translation targets:
    - Cell values (text only)
    - Shape text (TextBox, etc.) - xlwings only
    - Chart titles and labels - xlwings only

    Preserved:
    - Formulas (not translated)
    - Cell formatting (font, color, borders)
    - Column widths, row heights
    - Merged cells
    - Images
    - Charts (structure)

    Not translated:
    - Sheet names
    - Named ranges
    - Comments
    - Header/Footer text
    """

    def __init__(self):
        self.cell_translator = CellTranslator()
        # Warning messages for user feedback
        self._warnings: list[str] = []
        # Flag to indicate openpyxl fallback mode
        self._using_openpyxl_fallback = False
        # Font info cache: block_id -> (font_name, font_size)
        # Populated during extract, used during apply to avoid double COM calls
        self._font_cache: dict[str, tuple[Optional[str], float]] = {}
        # Merged cells cache: sheet_name -> {merge_address: (row1, col1, row2, col2)}
        # Built once per sheet during apply_translations to avoid repeated COM calls
        self._merged_cells_cache: dict[str, dict[str, tuple[int, int, int, int]]] = {}
        # XLSX worksheet XML cache for fast merged-cell detection (avoids slow COM scans on merge-heavy sheets)
        self._xlsx_sheet_xml_paths: dict[str, str] = {}
        self._xlsx_sheet_xml_paths_source: Optional[Path] = None

    def clear_warnings(self) -> None:
        """Clear accumulated warnings."""
        self._warnings.clear()
        self._using_openpyxl_fallback = False

    def clear_font_cache(self) -> None:
        """Clear font info cache."""
        self._font_cache.clear()

    def clear_merged_cells_cache(self) -> None:
        """Clear merged cells cache."""
        self._merged_cells_cache.clear()
        self._xlsx_sheet_xml_paths.clear()
        self._xlsx_sheet_xml_paths_source = None

    def _get_xlsx_sheet_xml_paths(self, file_path: Path) -> dict[str, str]:
        """Return a mapping from sheet name to worksheet XML path inside an XLSX ZIP."""
        if self._xlsx_sheet_xml_paths_source == file_path:
            return self._xlsx_sheet_xml_paths

        sheet_to_file: dict[str, str] = {}

        try:
            with zipfile.ZipFile(file_path, 'r') as xlsx:
                sheet_names: dict[str, str] = {}  # rId -> sheet_name
                sheet_order: list[str] = []  # ordered list of rIds

                try:
                    with xlsx.open('xl/workbook.xml') as workbook_xml:
                        for _event, elem in ET.iterparse(workbook_xml, events=['end']):
                            if elem.tag.endswith('}sheet') or elem.tag == 'sheet':
                                sheet_name = elem.get('name', '')
                                rid = None
                                for attr_name, attr_value in elem.attrib.items():
                                    if attr_name.endswith('}id') or attr_name == 'id':
                                        rid = attr_value
                                        break
                                if rid and sheet_name:
                                    sheet_names[rid] = sheet_name
                                    sheet_order.append(rid)
                                elem.clear()
                except KeyError:
                    self._xlsx_sheet_xml_paths_source = file_path
                    self._xlsx_sheet_xml_paths = sheet_to_file
                    return sheet_to_file

                rid_to_file: dict[str, str] = {}
                try:
                    with xlsx.open('xl/_rels/workbook.xml.rels') as rels_xml:
                        for _event, elem in ET.iterparse(rels_xml, events=['end']):
                            if elem.tag.endswith('}Relationship') or elem.tag == 'Relationship':
                                rid = elem.get('Id', '')
                                target = elem.get('Target', '')
                                if rid and target:
                                    target = target.lstrip('/')
                                    if target.startswith('xl/'):
                                        rid_to_file[rid] = target
                                    else:
                                        rid_to_file[rid] = f"xl/{target}"
                                elem.clear()
                except KeyError:
                    rid_to_file = {}

                for i, rid in enumerate(sheet_order, 1):
                    sheet_name = sheet_names.get(rid)
                    if not sheet_name:
                        continue
                    sheet_file = rid_to_file.get(rid) or f"xl/worksheets/sheet{i}.xml"
                    sheet_to_file[sheet_name] = sheet_file
        except (zipfile.BadZipFile, ET.ParseError) as e:
            logger.debug("Failed to parse XLSX sheet XML paths: %s", e)
            sheet_to_file = {}
        except Exception as e:
            logger.debug("Failed to parse XLSX sheet XML paths: %s", e)
            sheet_to_file = {}

        self._xlsx_sheet_xml_paths_source = file_path
        self._xlsx_sheet_xml_paths = sheet_to_file
        return sheet_to_file

    def _get_merged_cells_map(self, sheet) -> dict[str, tuple[int, int, int, int]]:
        """Get merged cells map for a sheet, building cache if needed.

        Uses Excel's FindFormat API to efficiently find all merged cells
        in a single scan, rather than checking each cell individually.

        Args:
            sheet: xlwings Sheet object

        Returns:
            Dict mapping merge_address to (row1, col1, row2, col2) bounds
        """
        sheet_name = sheet.name

        if sheet_name in self._merged_cells_cache:
            return self._merged_cells_cache[sheet_name]

        merged_cells: dict[str, tuple[int, int, int, int]] = {}

        try:
            used_range = sheet.used_range
            if used_range is None:
                self._merged_cells_cache[sheet_name] = merged_cells
                return merged_cells

            # Fast path: most sheets have no merged cells.
            # Excel may return:
            # - False: no merged cells in the range
            # - True / None (Variant Null): some merged cells exist
            try:
                merge_flag = used_range.api.MergeCells
                if merge_flag is False:
                    self._merged_cells_cache[sheet_name] = merged_cells
                    return merged_cells
            except Exception:
                pass

            # XLSX fast path: parse merged cells directly from the worksheet XML.
            # This is significantly faster than COM FindFormat scans on merge-heavy sheets.
            workbook_path = None
            try:
                workbook_fullname = getattr(getattr(sheet, "book", None), "fullname", None)
                if workbook_fullname:
                    workbook_path = Path(workbook_fullname)
            except Exception:
                workbook_path = None

            if workbook_path and workbook_path.suffix.lower() == ".xlsx" and workbook_path.exists():
                sheet_xml_paths = self._get_xlsx_sheet_xml_paths(workbook_path)
                sheet_xml_path = sheet_xml_paths.get(sheet_name)
                if sheet_xml_path:
                    t0 = time.perf_counter()
                    try:
                        with zipfile.ZipFile(workbook_path, 'r') as xlsx:
                            with xlsx.open(sheet_xml_path) as sheet_xml:
                                for _event, elem in ET.iterparse(sheet_xml, events=['end']):
                                    if elem.tag.endswith('}mergeCell') or elem.tag == 'mergeCell':
                                        ref = elem.get('ref')
                                        if ref:
                                            try:
                                                min_col, min_row, max_col, max_row = range_boundaries(ref)
                                                merged_cells[ref] = (min_row, min_col, max_row, max_col)
                                            except Exception:
                                                pass
                                    elem.clear()

                        logger.debug(
                            "Built merged cells map for sheet '%s' via XLSX parse: %d merged areas (%.2fs)",
                            sheet_name, len(merged_cells), time.perf_counter() - t0,
                        )
                        self._merged_cells_cache[sheet_name] = merged_cells
                        return merged_cells
                    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
                        logger.debug(
                            "XLSX merged-cell parse failed for sheet '%s' (falling back to COM scan): %s",
                            sheet_name, e,
                        )
                    except Exception as e:
                        logger.debug(
                            "XLSX merged-cell parse failed for sheet '%s' (falling back to COM scan): %s",
                            sheet_name, e,
                        )

            app = sheet.book.app

            # Clear FindFormat and set to search for merged cells
            app.api.FindFormat.Clear()
            app.api.FindFormat.MergeCells = True

            # Find first merged cell
            # LookAt=2 is xlPart, SearchFormat=True enables format-based search
            first_cell = used_range.api.Find(
                What="",
                LookAt=2,
                SearchFormat=True
            )

            if first_cell is None:
                self._merged_cells_cache[sheet_name] = merged_cells
                return merged_cells

            first_address = first_cell.Address
            current_cell = first_cell

            # Track visited addresses to prevent infinite loop
            visited_addresses: set[str] = set()
            # Safety limit to prevent infinite loop
            MAX_ITERATIONS = 10000

            # Loop through all merged cells
            for _ in range(MAX_ITERATIONS):
                current_address = current_cell.Address

                # Check if we've seen this cell before (prevents infinite loop)
                if current_address in visited_addresses:
                    break
                visited_addresses.add(current_address)

                merge_area = current_cell.MergeArea
                merge_address = merge_area.Address

                if merge_address not in merged_cells:
                    merged_cells[merge_address] = (
                        merge_area.Row,
                        merge_area.Column,
                        merge_area.Row + merge_area.Rows.Count - 1,
                        merge_area.Column + merge_area.Columns.Count - 1
                    )

                # Find next merged cell
                current_cell = used_range.api.FindNext(current_cell)

                # Stop when we've looped back to the first cell or no more cells
                if current_cell is None or current_cell.Address == first_address:
                    break

            logger.debug(
                "Built merged cells map for sheet '%s': %d merged areas",
                sheet_name, len(merged_cells)
            )

        except Exception as e:
            logger.debug("Error building merged cells map for '%s': %s", sheet_name, e)

        self._merged_cells_cache[sheet_name] = merged_cells
        return merged_cells

    def _is_cell_in_merged_area(
        self,
        row: int,
        col: int,
        merged_map: dict[str, tuple[int, int, int, int]]
    ) -> Optional[str]:
        """Check if a cell is within a merged area.

        Args:
            row: Cell row (1-indexed)
            col: Cell column (1-indexed)
            merged_map: Dict from _get_merged_cells_map()

        Returns:
            Merge address string if cell is in a merged area, None otherwise
        """
        for merge_address, (r1, c1, r2, c2) in merged_map.items():
            if r1 <= row <= r2 and c1 <= col <= c2:
                return merge_address
        return None

    @property
    def warnings(self) -> list[str]:
        """Get accumulated warnings."""
        return self._warnings.copy()

    @property
    def file_type(self) -> FileType:
        return FileType.EXCEL

    @property
    def supported_extensions(self) -> list[str]:
        return ['.xlsx', '.xls']

    def get_file_info(self, file_path: Path) -> FileInfo:
        """Get Excel file info.

        Uses fast ZIP parsing (openpyxl) for `.xlsx` because it's much faster than
        starting an Excel COM server via xlwings (often 3-15 seconds).

        For legacy `.xls`, openpyxl cannot read the format, so we require Microsoft
        Excel via xlwings.
        """
        self._ensure_xls_supported(file_path)
        if file_path.suffix.lower() == ".xls":
            xw = _get_xlwings()
            return self._get_file_info_xlwings(file_path, xw)
        return self._get_file_info_openpyxl(file_path)

    def _get_file_info_xlwings(self, file_path: Path, xw) -> FileInfo:
        """Get file info using xlwings (fast: sheet names only, no cell scanning)"""
        with com_initialized():
            # Use isolated app creation to prevent connecting to existing Excel instances
            app = _create_excel_app_with_retry(xw)
            try:
                wb = app.books.open(str(file_path), ignore_read_only_recommended=True)
                # SAFETY: Verify we opened the correct workbook (even for read-only)
                _verify_workbook_path(wb, file_path, "get_file_info")
                try:
                    sheet_count = len(wb.sheets)
                    section_details = [
                        SectionDetail(index=idx, name=sheet.name)
                        for idx, sheet in enumerate(wb.sheets)
                    ]

                    return FileInfo(
                        path=file_path,
                        file_type=FileType.EXCEL,
                        size_bytes=file_path.stat().st_size,
                        sheet_count=sheet_count,
                        section_details=section_details,
                    )
                finally:
                    wb.close()
            finally:
                app.quit()

    def _get_file_info_openpyxl(self, file_path: Path) -> FileInfo:
        """Get file info using fast ZIP parsing, falling back to openpyxl."""
        # Fast path: parse workbook.xml directly for sheet names (avoids full workbook load)
        try:
            fast_details = self._get_sheet_details_fast(file_path)
            if fast_details is not None:
                return FileInfo(
                    path=file_path,
                    file_type=FileType.EXCEL,
                    size_bytes=file_path.stat().st_size,
                    sheet_count=len(fast_details),
                    section_details=fast_details,
                )
        except Exception as e:
            # If fast path fails for any reason, fall back to openpyxl
            logger.debug("Fast sheet parse failed, falling back to openpyxl: %s", e)

        # Fallback: use openpyxl (read_only) to obtain sheet names
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                sheet_count = len(wb.sheetnames)
                section_details = [
                    SectionDetail(index=idx, name=name)
                    for idx, name in enumerate(wb.sheetnames)
                ]
            finally:
                wb.close()
        except (OSError, zipfile.BadZipFile, KeyError) as e:
            logger.warning("Error reading Excel file info: %s", e)
            raise

        return FileInfo(
            path=file_path,
            file_type=FileType.EXCEL,
            size_bytes=file_path.stat().st_size,
            sheet_count=sheet_count,
            section_details=section_details,
        )

    def _get_sheet_details_fast(self, file_path: Path) -> Optional[list[SectionDetail]]:
        """Extract sheet metadata by reading workbook.xml directly.

        This avoids fully loading the workbook with openpyxl, which can be slow for
        large files or when many uploads are processed in succession.
        """
        if file_path.suffix.lower() != ".xlsx":
            return None

        with zipfile.ZipFile(file_path) as zf:
            workbook_xml = zf.read("xl/workbook.xml")

        root = ET.fromstring(workbook_xml)
        namespace = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = root.findall("ns:sheets/ns:sheet", namespaces=namespace)

        section_details = [
            SectionDetail(index=idx, name=sheet.attrib.get("name", f"Sheet{idx+1}"))
            for idx, sheet in enumerate(sheets)
        ]

        return section_details if section_details else None

    def extract_sample_text_fast(
        self, file_path: Path, max_chars: int = 500
    ) -> Optional[str]:
        """Extract a text sample for language detection without full workbook load.

        This method uses iterparse to stream sharedStrings.xml directly from the xlsx
        archive, avoiding the overhead of loading the entire XML into memory.
        Large Excel files can have sharedStrings.xml of tens of MB, so streaming
        is essential for fast language detection.

        Args:
            file_path: Path to the Excel file
            max_chars: Maximum characters to extract (default 500)

        Returns:
            Sample text string or None if extraction fails
        """
        if file_path.suffix.lower() != ".xlsx":
            # Fall back to None for .xls files (require full load)
            return None

        try:
            texts = []
            total_chars = 0

            with zipfile.ZipFile(file_path, 'r') as zf:
                # sharedStrings.xml contains all unique text strings in the workbook
                if 'xl/sharedStrings.xml' not in zf.namelist():
                    logger.debug("No sharedStrings.xml found in xlsx")
                    return None

                # Use iterparse for streaming XML parsing (avoids loading entire XML into memory)
                # This is critical for large Excel files where sharedStrings.xml can be huge
                with zf.open('xl/sharedStrings.xml') as xml_file:
                    # Excel namespace
                    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    si_tag = f'{{{ns}}}si'
                    t_tag = f'{{{ns}}}t'
                    r_tag = f'{{{ns}}}r'

                    # Track current <si> element's text parts
                    current_parts = []
                    in_si = False

                    for event, elem in ET.iterparse(xml_file, events=('start', 'end')):
                        if event == 'start' and elem.tag == si_tag:
                            in_si = True
                            current_parts = []
                        elif event == 'end':
                            if elem.tag == t_tag and in_si:
                                # Collect text from <t> elements
                                if elem.text:
                                    current_parts.append(elem.text)
                            elif elem.tag == si_tag:
                                # End of <si> element - process collected text
                                in_si = False
                                if current_parts:
                                    text = ''.join(current_parts).strip()
                                    if text and len(text) > 1:  # Skip single chars
                                        texts.append(text)
                                        total_chars += len(text)
                                        if total_chars >= max_chars:
                                            break
                                # Clear element to free memory
                                elem.clear()

            if texts:
                result = ' '.join(texts)[:max_chars]
                logger.debug(
                    "Excel fast sample extraction: %d chars from %d strings",
                    len(result), len(texts)
                )
                return result

            return None

        except (zipfile.BadZipFile, ET.ParseError, KeyError) as e:
            logger.debug("Excel fast sample extraction failed: %s", e)
            return None

    def extract_text_blocks(
        self,
        file_path: Path,
        output_language: str = "en",
        selected_sections: Optional[list[int]] = None,
    ) -> Iterator[TextBlock]:
        """Extract text from cells, shapes, and charts

        Args:
            file_path: Path to the Excel file
            output_language: "en" for JP→EN, "jp" for EN→JP translation
            selected_sections: 処理対象のシート index (0-based)。None の場合は全シート。
        """
        # Clear warnings and caches at start of extraction
        self.clear_warnings()
        self.clear_font_cache()
        self.clear_merged_cells_cache()

        xw = _get_xlwings()

        self._ensure_xls_supported(file_path)

        if _can_use_xlwings():
            self._using_openpyxl_fallback = False
            yield from self._extract_text_blocks_xlwings(
                file_path, xw, output_language, selected_sections
            )
        else:
            # Add warning for openpyxl fallback mode
            self._using_openpyxl_fallback = True
            self._warnings.append(
                "xlwingsが利用できないため、シェイプ（テキストボックス等）と"
                "グラフのタイトル/ラベルは翻訳対象外です。"
                "フル機能を使用するにはMicrosoft Excelをインストールしてください。"
            )
            logger.warning(
                "xlwings/Excel not available, falling back to openpyxl. "
                "Shapes and charts will not be translated."
            )
            yield from self._extract_text_blocks_openpyxl(
                file_path, output_language, selected_sections
            )

    def _read_used_range_values_2d(self, used_range):
        """Read an xlwings range value as a 2D sequence.

        xlwings returns a 1D list for single-row/column ranges. If we misinterpret the
        orientation, extracted (row, col) metadata can be wrong, which later causes
        translations to be applied to the wrong cells.
        """
        try:
            values = used_range.options(ndim=2).value
        except Exception:
            values = used_range.value

        if values is None:
            return None

        if not isinstance(values, (list, tuple)):
            return [[values]]

        if values and not isinstance(values[0], (list, tuple)):
            # 1D list: distinguish single row vs single column.
            row_count = None
            col_count = None
            try:
                row_count = used_range.rows.count
                col_count = used_range.columns.count
            except Exception:
                try:
                    row_count = used_range.api.Rows.Count
                    col_count = used_range.api.Columns.Count
                except Exception:
                    row_count = None
                    col_count = None

            if row_count is None or col_count is None:
                addr = None
                try:
                    addr = used_range.address
                except Exception:
                    try:
                        addr = used_range.api.Address
                    except Exception:
                        addr = None
                if addr:
                    addr = addr.replace('$', '')
                    if '!' in addr:
                        addr = addr.split('!')[-1]
                    try:
                        if ':' in addr:
                            start_addr, end_addr = addr.split(':', 1)
                        else:
                            start_addr = addr
                            end_addr = addr
                        start_col_letters, start_row_str = coordinate_from_string(start_addr)
                        end_col_letters, end_row_str = coordinate_from_string(end_addr)
                        row_count = int(end_row_str) - int(start_row_str) + 1
                        col_count = (
                            column_index_from_string(end_col_letters)
                            - column_index_from_string(start_col_letters)
                            + 1
                        )
                    except Exception:
                        pass

            if row_count == 1 and col_count != 1:
                return [list(values)]
            if col_count == 1 and row_count != 1:
                return [[v] for v in values]

            # Last resort: prefer treating it as a single column, which avoids shifting
            # vertical lists across columns.
            return [[v] for v in values]

        return values

    def _extract_text_blocks_xlwings(
        self,
        file_path: Path,
        xw,
        output_language: str = "en",
        selected_sections: Optional[list[int]] = None,
    ) -> Iterator[TextBlock]:
        """Extract text using xlwings

        Optimized for performance:
        - Bulk read all cell values at once via used_range.value
        - Fetch font info for cells during extraction (cached for apply_translations)

        Args:
            file_path: Path to the Excel file
            xw: xlwings module
            output_language: "en" for JP→EN, "jp" for EN→JP translation

        Thread Safety / COM Constraints:
            COM initialization is required when called from worker threads
            (e.g., asyncio.to_thread). The com_initialized() context manager
            handles CoInitialize/CoUninitialize for the current thread.

            IMPORTANT: All COM operations MUST complete within the com_initialized()
            context. This is why blocks are collected into a list before yielding,
            rather than yielding directly from within the context.

            If you modify this method to yield blocks directly, the generator may
            be consumed in a different thread context, causing COM errors like:
            - "CoInitialize has not been called"
            - "The application called an interface that was marshalled for a different thread"

            DO NOT change to `yield TextBlock(...)` inside the com_initialized() block.

        Memory Considerations:
            All blocks are collected into a list before yielding due to COM constraints.
            For very large files (10,000+ translatable cells), this may consume
            significant memory. Unlike PDF processing which supports streaming,
            Excel COM operations require all work to complete within the same
            thread context, making true streaming impractical.

            If memory becomes an issue for extremely large files, consider:
            1. Processing sheets individually (already somewhat optimized)
            2. Splitting the file into smaller chunks before processing
            3. Using openpyxl fallback which has better streaming support for reads
        """
        import time as _time
        blocks: list[TextBlock] = []
        selected_set = set(selected_sections) if selected_sections is not None else None

        with com_initialized():
            _t_start = _time.perf_counter()
            app = _create_excel_app_with_retry(xw)
            _t_app = _time.perf_counter()
            logger.debug("[TIMING] Excel app creation: %.2fs", _t_app - _t_start)
            try:
                wb = app.books.open(str(file_path), ignore_read_only_recommended=True)
                _t_open = _time.perf_counter()
                logger.debug("[TIMING] Workbook open: %.2fs", _t_open - _t_app)
                # SAFETY: Verify we opened the correct workbook
                _verify_workbook_path(wb, file_path, "extract_text_blocks")
                try:
                    _t_sheets_start = _time.perf_counter()
                    for sheet_idx, sheet in enumerate(wb.sheets):
                        if selected_set is not None and sheet_idx not in selected_set:
                            continue
                        sheet_name = sheet.name

                        # === Cells (bulk read optimization) ===
                        # Wrap in try-except to handle COM errors
                        try:
                            used_range = sheet.used_range
                            if used_range is not None:
                                # Get all values at once (much faster than cell-by-cell)
                                all_values = self._read_used_range_values_2d(used_range)
                                if all_values is not None:

                                    # Get start position of used_range
                                    start_row = used_range.row
                                    start_col = used_range.column

                                    # First pass: identify translatable cells
                                    translatable_cells = []
                                    for row_offset, row_values in enumerate(all_values):
                                        if row_values is None:
                                            continue
                                        for col_offset, value in enumerate(row_values):
                                            if value and isinstance(value, str):
                                                if self.cell_translator.should_translate(str(value), output_language):
                                                    row_idx = start_row + row_offset
                                                    col_idx = start_col + col_offset
                                                    translatable_cells.append((row_idx, col_idx, str(value)))

                                    # Skip if no translatable cells found
                                    if not translatable_cells:
                                        continue

                                    # Merged cells: xlwings/COM may report the merged value on every cell in the merge area.
                                    # Keep only the merge area's top-left cell to avoid duplicate translation blocks and
                                    # misaligned output when applying translations.
                                    try:
                                        merged_map = self._get_merged_cells_map(sheet)
                                    except Exception:
                                        merged_map = {}

                                    if merged_map:
                                        filtered_cells: list[tuple[int, int, str]] = []
                                        for r, c, txt in translatable_cells:
                                            merge_address = self._is_cell_in_merged_area(r, c, merged_map)
                                            if merge_address:
                                                r1, c1, _r2, _c2 = merged_map[merge_address]
                                                if (r, c) != (r1, c1):
                                                    continue
                                            filtered_cells.append((r, c, txt))
                                        translatable_cells = filtered_cells
                                        if not translatable_cells:
                                            continue

                                    # Batch optimization: Get formula cells in one COM call
                                    # Uses SpecialCells(xlCellTypeFormulas) which is much faster
                                    # than checking each cell individually
                                    formula_cells: set[tuple[int, int]] = set()
                                    try:
                                        # xlCellTypeFormulas = -4123
                                        formula_range = used_range.api.SpecialCells(-4123)
                                        # Parse addresses to get formula cell coordinates
                                        # Address can be like "A1,B2:C3,D4" (comma-separated areas)
                                        for area in formula_range.Areas:
                                            addr = area.Address.replace('$', '')
                                            # Handle range like "A1:B3" or single cell "A1"
                                            if ':' in addr:
                                                parts = addr.split(':')
                                                start_cell = parts[0]
                                                end_cell = parts[1]
                                                # Parse start/end
                                                start_match = re.match(r'([A-Z]+)(\d+)', start_cell)
                                                end_match = re.match(r'([A-Z]+)(\d+)', end_cell)
                                                if start_match and end_match:
                                                    start_c = column_index_from_string(start_match.group(1))
                                                    start_r = int(start_match.group(2))
                                                    end_c = column_index_from_string(end_match.group(1))
                                                    end_r = int(end_match.group(2))
                                                    for r in range(start_r, end_r + 1):
                                                        for c in range(start_c, end_c + 1):
                                                            formula_cells.add((r, c))
                                            else:
                                                match = re.match(r'([A-Z]+)(\d+)', addr)
                                                if match:
                                                    c = column_index_from_string(match.group(1))
                                                    r = int(match.group(2))
                                                    formula_cells.add((r, c))
                                    except Exception:
                                        # No formula cells or SpecialCells not supported
                                        # This is expected if sheet has no formulas
                                        pass

                                    # Get sheet's default font size (single COM call)
                                    # Font name is not needed as it's determined by translation direction
                                    default_font_size = 11.0
                                    try:
                                        # Try to get from first cell of used range
                                        first_cell_size = used_range.api.Cells(1, 1).Font.Size
                                        if first_cell_size:
                                            default_font_size = float(first_cell_size)
                                    except Exception:
                                        pass

                                    # Second pass: collect TextBlocks (optimized - no per-cell COM calls)
                                    formula_skipped = 0
                                    for row_idx, col_idx, text in translatable_cells:
                                        # Skip formula cells (already identified in batch)
                                        if (row_idx, col_idx) in formula_cells:
                                            formula_skipped += 1
                                            continue

                                        col_letter = get_column_letter(col_idx)
                                        block_id = f"{sheet_name}_{col_letter}{row_idx}"

                                        # Use default font size (font name not needed - determined by direction)
                                        # Cache font info for use in apply_translations
                                        self._font_cache[block_id] = (None, default_font_size)

                                        blocks.append(TextBlock(
                                            id=block_id,
                                            text=text,
                                            location=f"{sheet_name}, {col_letter}{row_idx}",
                                            metadata={
                                                'sheet': sheet_name,
                                                'sheet_idx': sheet_idx,
                                                'row': row_idx,
                                                'col': col_idx,
                                                'type': 'cell',
                                                'font_name': None,
                                                'font_size': default_font_size,
                                            }
                                        ))

                                    if formula_skipped > 0:
                                        logger.info(
                                            "Skipped %d formula cells in sheet '%s' (formulas preserved)",
                                            formula_skipped, sheet_name
                                        )
                        except Exception as e:
                            logger.warning("Error reading used_range in sheet '%s': %s", sheet_name, e)

                        # === Shapes (TextBox, etc.) ===
                        # Wrap shape iteration in try-except to handle COM errors
                        # Note: Some shape types (images, OLE objects, charts) don't have
                        # a text property and will raise COM errors when accessed
                        try:
                            # Optimization: Skip shape iteration if no shapes exist
                            shape_count = len(sheet.shapes)
                            if shape_count > 0:
                                for shape_idx, shape in enumerate(sheet.shapes):
                                    try:
                                        # First check shape type - some shapes don't support text
                                        # xlwings shape types: 1=msoAutoShape, 17=msoTextBox, etc.
                                        # Skip shapes that typically don't have text (images, OLE, etc.)
                                        shape_type = None
                                        try:
                                            shape_type = shape.type
                                        except Exception:
                                            # If we can't even get the type, skip this shape
                                            continue

                                        # Skip shape types that don't have text:
                                        # 13 = msoPicture (images)
                                        # 3 = msoChart (embedded charts)
                                        # 12 = msoOLEControlObject
                                        # 7 = msoEmbeddedOLEObject
                                        # 10 = msoLinkedOLEObject
                                        # 19 = msoMedia
                                        non_text_types = {3, 7, 10, 12, 13, 19}
                                        if shape_type in non_text_types:
                                            continue

                                        # Try to get text from shape
                                        text = None
                                        try:
                                            if hasattr(shape, 'text'):
                                                raw_text = shape.text  # Access only once
                                                if raw_text:
                                                    text = raw_text.strip()
                                        except Exception:
                                            # COM error accessing text - shape doesn't support text
                                            continue

                                        if text and self.cell_translator.should_translate(text, output_language):
                                            # Get shape name safely
                                            shape_name = None
                                            try:
                                                shape_name = shape.name
                                            except Exception:
                                                shape_name = f"Shape_{shape_idx}"

                                            blocks.append(TextBlock(
                                                id=f"{sheet_name}_shape_{shape_idx}",
                                                text=text,
                                                location=f"{sheet_name}, Shape '{shape_name}'",
                                                metadata={
                                                    'sheet': sheet_name,
                                                    'sheet_idx': sheet_idx,
                                                    'shape': shape_idx,
                                                    'shape_name': shape_name,
                                                    'type': 'shape',
                                                }
                                            ))
                                    except Exception as e:
                                        # Log at debug level - many shapes legitimately don't have text
                                        logger.debug("Skipping shape %d in sheet '%s': %s", shape_idx, sheet_name, e)
                        except Exception as e:
                            logger.warning("Error iterating shapes in sheet '%s': %s", sheet_name, e)

                        # === Chart Titles and Labels ===
                        # Wrap chart iteration in try-except to handle COM errors
                        try:
                            # Optimization: Skip chart iteration if no charts exist
                            chart_count = len(sheet.charts)
                            if chart_count > 0:
                                for chart_idx, chart in enumerate(sheet.charts):
                                    try:
                                        api_chart = chart.api[1]  # xlwings COM object (1-indexed)

                                        # Chart title
                                        if api_chart.HasTitle:
                                            title = api_chart.ChartTitle.Text
                                            if title and self.cell_translator.should_translate(title, output_language):
                                                blocks.append(TextBlock(
                                                    id=f"{sheet_name}_chart_{chart_idx}_title",
                                                    text=title,
                                                    location=f"{sheet_name}, Chart {chart_idx + 1} Title",
                                                    metadata={
                                                        'sheet': sheet_name,
                                                        'sheet_idx': sheet_idx,
                                                        'chart': chart_idx,
                                                        'type': 'chart_title',
                                                    }
                                                ))

                                        # Axis titles
                                        for axis_type, axis_name in [(1, 'category'), (2, 'value')]:
                                            try:
                                                axis = api_chart.Axes(axis_type)
                                                if axis.HasTitle:
                                                    axis_title = axis.AxisTitle.Text
                                                    if axis_title and self.cell_translator.should_translate(axis_title, output_language):
                                                        blocks.append(TextBlock(
                                                            id=f"{sheet_name}_chart_{chart_idx}_axis_{axis_name}",
                                                            text=axis_title,
                                                            location=f"{sheet_name}, Chart {chart_idx + 1} {axis_name.title()} Axis",
                                                            metadata={
                                                                'sheet': sheet_name,
                                                                'sheet_idx': sheet_idx,
                                                                'chart': chart_idx,
                                                                'axis': axis_name,
                                                                'type': 'chart_axis_title',
                                                            }
                                                        ))
                                            except Exception as e:
                                                logger.debug("Error reading %s axis title for chart %d: %s", axis_name, chart_idx, e)

                                    except Exception as e:
                                        logger.debug("Error extracting chart %d in sheet '%s': %s", chart_idx, sheet_name, e)
                        except Exception as e:
                            logger.warning("Error iterating charts in sheet '%s': %s", sheet_name, e)
                    _t_sheets_end = _time.perf_counter()
                    logger.debug("[TIMING] All sheets processed: %.2fs (%d sheets, %d blocks)",
                                _t_sheets_end - _t_sheets_start, len(wb.sheets), len(blocks))
                finally:
                    wb.close()
                    _t_close = _time.perf_counter()
                    logger.debug("[TIMING] Workbook close: %.2fs", _t_close - _t_sheets_end)
            finally:
                _t_before_quit = _time.perf_counter()
                app.quit()
                _t_quit = _time.perf_counter()
                logger.debug("[TIMING] Excel app quit: %.2fs", _t_quit - _t_before_quit)
                logger.debug("[TIMING] Total extraction: %.2fs", _t_quit - _t_start)

        # Log warning for very large files
        block_count = len(blocks)
        if block_count > 10000:
            logger.warning(
                "Large Excel file: %d translatable blocks collected. "
                "This may consume significant memory during translation.",
                block_count
            )
        elif block_count > 5000:
            logger.info(
                "Processing %d translatable blocks from Excel file.",
                block_count
            )

        # Yield blocks after COM operations complete
        yield from blocks

    def _extract_text_blocks_openpyxl(
        self,
        file_path: Path,
        output_language: str = "en",
        selected_sections: Optional[list[int]] = None,
    ) -> Iterator[TextBlock]:
        """Extract text using openpyxl (fallback - cells only)

        Single-pass approach with lightweight formula detection:
        1. Detect formula cells via zipfile XML parsing (memory-efficient)
        2. Extract text blocks with data_only=True to get calculated values

        The zipfile approach is much more memory-efficient than loading
        the workbook twice, as it only parses the necessary XML portions.

        Font info is not available in read_only mode but is fetched
        during apply_translations from the original file.

        Args:
            file_path: Path to the Excel file
            output_language: "en" for JP→EN, "jp" for EN→JP translation
        """
        # Detect formula cells via lightweight zipfile parsing
        # This is much more memory-efficient than loading with data_only=False
        formula_cells = _detect_formula_cells_via_zipfile(file_path)
        formula_count = len(formula_cells)

        if formula_count > 0:
            logger.info("Detected %d formula cells (will be preserved)", formula_count)

        # Extract text blocks with calculated values (data_only=True)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Cache column letters to avoid repeated conversions during large reads
        # Limited to _COLUMN_LETTER_CACHE_SIZE entries to prevent memory bloat on very wide sheets
        # (Excel max is 16,384 columns, but typically only a fraction are used)
        column_letter_cache: dict[int, str] = {}
        selected_set = set(selected_sections) if selected_sections is not None else None

        try:
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                if selected_set is not None and sheet_idx not in selected_set:
                    continue
                sheet = wb[sheet_name]

                # Limit iteration to the used range for the sheet to avoid scanning
                # entire default grids (e.g., 1,048,576 rows).
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(sheet.calculate_dimension())
                except (ValueError, TypeError):
                    continue

                for row_idx, row_values in enumerate(
                    sheet.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    ),
                    start=min_row,
                ):
                    if not row_values:
                        continue

                    for col_idx, value in enumerate(row_values, start=min_col):
                        # Skip formula cells to preserve them
                        if (sheet_name, row_idx, col_idx) in formula_cells:
                            continue

                        if value and isinstance(value, str):
                            value_str = str(value)
                            if self.cell_translator.should_translate(value_str, output_language):
                                col_letter = column_letter_cache.get(col_idx)
                                if col_letter is None:
                                    col_letter = get_column_letter(col_idx)
                                    # Limit cache size to prevent memory bloat
                                    if len(column_letter_cache) < _COLUMN_LETTER_CACHE_SIZE:
                                        column_letter_cache[col_idx] = col_letter

                                # Font info not available in read_only mode
                                # Will be fetched during apply_translations
                                yield TextBlock(
                                    id=f"{sheet_name}_{col_letter}{row_idx}",
                                    text=value_str,
                                    location=f"{sheet_name}, {col_letter}{row_idx}",
                                    metadata={
                                        'sheet': sheet_name,
                                        'sheet_idx': sheet_idx,
                                        'row': row_idx,
                                        'col': col_idx,
                                        'type': 'cell',
                                        'font_name': None,
                                        'font_size': 11.0,
                                    }
                                )
        finally:
            wb.close()

    def _group_translations_by_sheet(
        self,
        translations: dict[str, str],
        sheet_names: set[str],
        text_blocks: Optional[list[TextBlock]] = None,
    ) -> dict[str, dict]:
        """
        Group translations by sheet name for efficient batch processing.

        Args:
            translations: dict mapping block_id to translated text
            sheet_names: set of valid sheet names in the workbook
            text_blocks: Optional list of TextBlock objects with metadata (row/col info)
                        When provided, uses metadata for precise positioning.

        Returns:
            dict mapping sheet_name to {
                'cells': {(row, col): (translated_text, cell_ref)},
                'shapes': {shape_idx: translated_text},
                'charts': {chart_idx: {'title': text, 'category': text, 'value': text}}
            }

        Block ID formats:
            - Cells: "SheetName_A1", "SheetName_AA100"
            - Shapes: "SheetName_shape_0", "SheetName_shape_1"
            - Charts: "SheetName_chart_0_title", "SheetName_chart_0_axis_category"

        Note on sheet names:
            Excel prohibits certain characters in sheet names (\\/?*[]:), so we don't
            need to sanitize them here. The block_id uses the exact sheet name from
            the workbook, ensuring consistent matching between extract and apply.

            Sheet names containing underscores are handled by sorting names by length
            (longest first) to avoid prefix collision during matching.
        """
        result: dict[str, dict] = {}

        # Build a lookup from block_id to TextBlock metadata for precise positioning
        block_metadata: dict[str, dict] = {}
        if text_blocks:
            for block in text_blocks:
                block_metadata[block.id] = block.metadata

        # Sort sheet names by length (longest first, then alphabetically for stability)
        # This ensures consistent matching even when sheet_names comes from a set.
        # e.g., "my_sheet" should match before "my" for block_id "my_sheet_A1"
        sorted_sheet_names = sorted(sheet_names, key=lambda x: (-len(x), x))

        # Pre-compile regex for valid cell reference pattern (A1, AA100, etc.)
        cell_ref_pattern = re.compile(r'^[A-Z]+\d+$')

        for block_id, translated_text in translations.items():
            # First, try to use metadata for precise positioning (most reliable)
            metadata = block_metadata.get(block_id, {})
            if metadata and metadata.get('type') == 'cell':
                sheet_name = metadata.get('sheet')
                row = metadata.get('row')
                col = metadata.get('col')
                if sheet_name and row and col and sheet_name in sheet_names:
                    if sheet_name not in result:
                        result[sheet_name] = {'cells': {}, 'shapes': {}, 'charts': {}}
                    # Store as (row, col) tuple for direct access
                    cell_ref = get_column_letter(col) + str(row)
                    result[sheet_name]['cells'][(row, col)] = (translated_text, cell_ref)
                    continue

            # Fallback: parse block_id (for backward compatibility or missing metadata)
            # Find matching sheet name (handles sheet names with underscores)
            # Validates that suffix is a valid block type to avoid false matches
            sheet_name = None
            for name in sorted_sheet_names:
                if block_id.startswith(f"{name}_"):
                    suffix = block_id[len(name) + 1:]
                    # Validate suffix is a valid block type
                    if (cell_ref_pattern.match(suffix) or
                        suffix.startswith("shape_") or
                        suffix.startswith("chart_")):
                        sheet_name = name
                        break

            if sheet_name is None:
                logger.warning("Could not determine sheet for block_id: %s", block_id)
                continue

            # Initialize sheet entry if needed
            if sheet_name not in result:
                result[sheet_name] = {'cells': {}, 'shapes': {}, 'charts': {}}

            # Parse the suffix after sheet name
            suffix = block_id[len(sheet_name) + 1:]  # +1 for underscore

            if suffix.startswith("shape_"):
                # Shape: "shape_0" -> shape_idx=0
                try:
                    shape_idx = int(suffix[6:])  # len("shape_") = 6
                    result[sheet_name]['shapes'][shape_idx] = translated_text
                except ValueError:
                    logger.debug("Invalid shape block_id: %s", block_id)

            elif suffix.startswith("chart_"):
                # Chart: "chart_0_title" or "chart_0_axis_category"
                parts = suffix.split("_")
                if len(parts) >= 3:
                    try:
                        chart_idx = int(parts[1])
                        if chart_idx not in result[sheet_name]['charts']:
                            result[sheet_name]['charts'][chart_idx] = {}

                        if parts[2] == "title":
                            result[sheet_name]['charts'][chart_idx]['title'] = translated_text
                        elif parts[2] == "axis" and len(parts) >= 4:
                            axis_name = parts[3]  # "category" or "value"
                            result[sheet_name]['charts'][chart_idx][axis_name] = translated_text
                    except ValueError:
                        logger.debug("Invalid chart block_id: %s", block_id)

            else:
                # Cell reference: "A1", "AA100", etc.
                # Parse to (row, col) for consistent storage
                try:
                    col_letters, row_str = coordinate_from_string(suffix)
                    row = int(row_str)
                    col = column_index_from_string(col_letters)
                    result[sheet_name]['cells'][(row, col)] = (translated_text, suffix)
                except (ValueError, TypeError) as e:
                    logger.warning("Invalid cell reference in block_id %s: %s", block_id, e)

        return result

    def apply_translations(
        self,
        input_path: Path,
        output_path: Path,
        translations: dict[str, str],
        direction: str = "jp_to_en",
        settings=None,
        selected_sections: Optional[list[int]] = None,
        text_blocks: Optional[list[TextBlock]] = None,
    ) -> None:
        """Apply translations to Excel file.

        Args:
            input_path: Path to input Excel file
            output_path: Path to save translated file
            translations: Dict mapping block_id to translated text
            direction: "jp_to_en" or "en_to_jp"
            settings: AppSettings for font configuration
            selected_sections: List of sheet indices to process (0-indexed).
                              If None, all sheets are processed.
            text_blocks: Optional list of TextBlock objects with metadata.
                        When provided, enables precise cell positioning using
                        row/col from metadata instead of parsing block_id.
        """
        xw = _get_xlwings()

        self._ensure_xls_supported(input_path)

        if _can_use_xlwings():
            self._apply_translations_xlwings(
                input_path, output_path, translations, direction, xw, settings,
                selected_sections, text_blocks
            )
        else:
            self._apply_translations_openpyxl(
                input_path, output_path, translations, direction, settings,
                selected_sections, text_blocks
            )

    def _ensure_xls_supported(self, file_path: Path) -> None:
        """Validate that .xls files can be processed.

        On platforms without Excel/xlwings support (e.g., this Linux runtime),
        .xls files cannot be opened by openpyxl. Surface a clear error instead
        of failing with a confusing ZIP parsing exception.
        """

        # Refresh xlwings availability in case environment changes at runtime
        _get_xlwings()

        if file_path.suffix.lower() == '.xls' and not _can_use_xlwings():
            raise ValueError(
                "XLS files require Microsoft Excel via xlwings. "
                "Install xlwings with Excel or convert the file to XLSX."
            )

    def _apply_translations_xlwings(
        self,
        input_path: Path,
        output_path: Path,
        translations: dict[str, str],
        direction: str,
        xw,
        settings=None,
        selected_sections: Optional[list[int]] = None,
        text_blocks: Optional[list[TextBlock]] = None,
    ) -> None:
        """Apply translations using xlwings

        Optimized:
        - Pre-groups translations by sheet to avoid O(sheets × translations) complexity
        - Uses metadata from text_blocks for precise cell positioning
        - Batch writes contiguous row ranges to minimize COM calls
        - Disables ScreenUpdating and sets Calculation to manual for faster COM operations
        - Only processes selected sheets when selected_sections is specified
        - Restores Excel settings after completion

        Note: COM initialization is required when called from worker threads.
        """
        font_manager = FontManager(direction, settings)

        with com_initialized():
            app = _create_excel_app_with_retry(xw)

            # Store original Excel settings for restoration
            original_screen_updating = None
            original_calculation = None
            original_enable_events = None
            original_display_alerts = None

            try:
                # Optimize Excel settings for batch operations
                try:
                    original_screen_updating = app.screen_updating
                    app.screen_updating = False
                except Exception as e:
                    logger.debug("Could not disable screen updating: %s", e)

                try:
                    original_enable_events = app.api.EnableEvents
                    app.api.EnableEvents = False
                except Exception as e:
                    logger.debug("Could not disable events: %s", e)

                try:
                    original_display_alerts = app.api.DisplayAlerts
                    app.api.DisplayAlerts = False
                except Exception as e:
                    logger.debug("Could not disable display alerts: %s", e)

                wb = app.books.open(str(input_path), ignore_read_only_recommended=True)

                # SAFETY: Verify we opened the correct workbook
                _verify_workbook_path(wb, input_path, "apply_translations")

                # Set calculation mode AFTER opening workbook (more reliable)
                # Use Excel API directly with constant value
                # xlCalculationManual = -4135, xlCalculationAutomatic = -4105
                try:
                    original_calculation = app.api.Calculation
                    app.api.Calculation = -4135  # xlCalculationManual
                except Exception as e:
                    logger.debug("Could not set manual calculation (performance optimization skipped): %s", e)
                    original_calculation = None

                try:
                    # Pre-group translations by sheet name using metadata when available
                    sheet_names = {sheet.name for sheet in wb.sheets}
                    translations_by_sheet = self._group_translations_by_sheet(
                        translations, sheet_names, text_blocks
                    )

                    # Convert selected_sections to a set for O(1) lookup
                    selected_set = set(selected_sections) if selected_sections is not None else None

                    for sheet_idx, sheet in enumerate(wb.sheets):
                        # Skip sheets not in selected_sections (if specified)
                        if selected_set is not None and sheet_idx not in selected_set:
                            continue

                        sheet_name = sheet.name
                        sheet_translations = translations_by_sheet.get(sheet_name, {})

                        # === Apply to cells with batch optimization ===
                        cell_translations = sheet_translations.get('cells', {})
                        if cell_translations:
                            self._apply_cell_translations_xlwings_batch(
                                sheet, sheet_name, cell_translations, font_manager
                            )

                        # === Apply to shapes ===
                        shape_translations = sheet_translations.get('shapes', {})
                        if shape_translations:
                            try:
                                for shape_idx, shape in enumerate(sheet.shapes):
                                    if shape_idx in shape_translations:
                                        try:
                                            shape.text = shape_translations[shape_idx]
                                        except Exception as e:
                                            logger.debug("Error applying translation to shape %d in '%s': %s", shape_idx, sheet_name, e)
                            except Exception as e:
                                logger.warning("Error iterating shapes in sheet '%s': %s", sheet_name, e)

                        # === Apply to chart titles and labels ===
                        chart_translations = sheet_translations.get('charts', {})
                        if chart_translations:
                            try:
                                for chart_idx, chart in enumerate(sheet.charts):
                                    if chart_idx not in chart_translations:
                                        continue
                                    try:
                                        api_chart = chart.api[1]
                                        chart_data = chart_translations[chart_idx]

                                        # Chart title
                                        if 'title' in chart_data and api_chart.HasTitle:
                                            api_chart.ChartTitle.Text = chart_data['title']

                                        # Axis titles
                                        for axis_type, axis_name in [(1, 'category'), (2, 'value')]:
                                            if axis_name in chart_data:
                                                try:
                                                    axis = api_chart.Axes(axis_type)
                                                    if axis.HasTitle:
                                                        axis.AxisTitle.Text = chart_data[axis_name]
                                                except Exception as e:
                                                    logger.debug("Error applying translation to %s axis of chart %d: %s", axis_name, chart_idx, e)

                                    except Exception as e:
                                        logger.debug("Error applying translation to chart %d in sheet '%s': %s", chart_idx, sheet_name, e)
                            except Exception as e:
                                logger.warning("Error iterating charts in sheet '%s': %s", sheet_name, e)

                    # Clear read_only_recommended to prevent Excel dialog on open
                    try:
                        wb.api.ReadOnlyRecommended = False
                    except Exception as e:
                        logger.debug("Could not clear ReadOnlyRecommended: %s", e)

                    # Save to output path
                    try:
                        wb.save(str(output_path))
                    except Exception as e:
                        logger.error("Error saving workbook: %s", e)
                        raise

                    # Restore calculation mode BEFORE closing workbook (more reliable)
                    # Excel may reject Calculation changes after workbook is closed
                    try:
                        if original_calculation is not None:
                            app.api.Calculation = original_calculation
                            original_calculation = None  # Mark as restored
                    except Exception as e:
                        logger.debug("Could not restore calculation mode (will retry after close): %s", e)
                finally:
                    wb.close()

            finally:
                # Restore Excel settings before quitting
                try:
                    if original_display_alerts is not None:
                        app.api.DisplayAlerts = original_display_alerts
                except Exception as e:
                    logger.debug("Could not restore DisplayAlerts: %s", e)

                try:
                    if original_enable_events is not None:
                        app.api.EnableEvents = original_enable_events
                except Exception as e:
                    logger.debug("Could not restore EnableEvents: %s", e)

                try:
                    if original_calculation is not None:
                        app.api.Calculation = original_calculation
                except Exception as e:
                    logger.debug("Could not restore calculation mode: %s", e)

                try:
                    if original_screen_updating is not None:
                        app.screen_updating = original_screen_updating
                except Exception as e:
                    logger.debug("Could not restore screen updating: %s", e)

                app.quit()

    def _apply_cell_translations_xlwings_batch(
        self,
        sheet,
        sheet_name: str,
        cell_translations: dict[tuple[int, int], tuple[str, str]],
        font_manager: FontManager,
    ) -> None:
        """Apply cell translations with batch optimization for xlwings.

        Optimization strategy:
        1. Write cell values (batched by contiguous ranges)
           - Row-wise batching for row-dense tables
           - Column-wise batching for column-heavy sheets (common: 1 text column)
        2. Apply output font *name* only to translated ranges (preserves existing sizes/styles)
        3. Optionally apply per-cell font-size adjustment when configured (JP→EN only)

        Args:
            sheet: xlwings Sheet object
            sheet_name: Name of the sheet (for logging/cache lookup)
            cell_translations: Dict mapping (row, col) to (translated_text, cell_ref)
            font_manager: FontManager for determining output fonts
        """
        if not cell_translations:
            return

        debug_stats = logger.isEnabledFor(logging.DEBUG)
        if debug_stats:
            t_start = time.perf_counter()
            t_values = 0.0
            t_fonts = 0.0
            t_sizes = 0.0
            total_segments = 0
            value_range_writes = 0
            value_cell_writes = 0
            font_range_applies = 0
            font_cell_applies = 0
            font_range_fallbacks = 0

        # Get output font settings (single lookup)
        output_font_name = font_manager.get_output_font()

        # Handle merged cells:
        # - If extraction (or Excel) reports the same merged value for non-top-left cells,
        #   normalize them to the merge area's top-left so we don't write to invalid cells.
        # - Avoid batched range writes/formatting when a batch includes merged cells.
        merged_map: dict[str, tuple[int, int, int, int]] = {}
        merged_top_lefts: set[tuple[int, int]] = set()
        try:
            merged_map = self._get_merged_cells_map(sheet)
            if merged_map:
                merged_top_lefts = {(r1, c1) for (r1, c1, _r2, _c2) in merged_map.values()}
        except Exception as e:
            logger.debug("Could not build merged cells map for sheet '%s': %s", sheet_name, e)
            merged_map = {}
            merged_top_lefts = set()

        if merged_map:
            normalized: dict[tuple[int, int], tuple[str, str]] = {}
            normalized_is_top_left_source: dict[tuple[int, int], bool] = {}
            redirected = 0

            for (row, col), (translated_text, cell_ref) in cell_translations.items():
                target_row, target_col = row, col
                is_top_left_source = True

                merge_address = self._is_cell_in_merged_area(row, col, merged_map)
                if merge_address:
                    r1, c1, _r2, _c2 = merged_map[merge_address]
                    if (row, col) != (r1, c1):
                        target_row, target_col = r1, c1
                        is_top_left_source = False
                        cell_ref = f"{get_column_letter(c1)}{r1}"
                        redirected += 1

                key = (target_row, target_col)
                if key not in normalized:
                    normalized[key] = (translated_text, cell_ref)
                    normalized_is_top_left_source[key] = is_top_left_source
                    continue

                # Prefer the translation originating from the top-left cell if both exist.
                if is_top_left_source and not normalized_is_top_left_source.get(key, False):
                    normalized[key] = (translated_text, cell_ref)
                    normalized_is_top_left_source[key] = True

            if redirected:
                logger.debug(
                    "Normalized %d merged-cell translations to top-left in sheet '%s'",
                    redirected, sheet_name,
                )

            cell_translations = normalized

        # Group translations by row/col for batch value writing.
        rows_data: dict[int, list[tuple[int, str, str]]] = {}  # row -> [(col, text, cell_ref), ...]
        cols_data: dict[int, list[tuple[int, str, str]]] = {}  # col -> [(row, text, cell_ref), ...]
        for (row, col), (translated_text, cell_ref) in cell_translations.items():
            rows_data.setdefault(row, []).append((col, translated_text, cell_ref))
            cols_data.setdefault(col, []).append((row, translated_text, cell_ref))

        # Sort cells within each row/col by position (for contiguous range detection).
        for row in rows_data:
            rows_data[row].sort(key=lambda x: x[0])
        for col in cols_data:
            cols_data[col].sort(key=lambda x: x[0])

        # Font-size adjustment is only relevant for JP→EN when user configured a negative adjustment.
        # (Positive adjustments are clamped to the original size by FontSizeAdjuster.)
        needs_size_adjustment = (
            getattr(font_manager, "direction", None) == "jp_to_en"
            and getattr(getattr(font_manager, "font_size_adjuster", None), "adjustment_jp_to_en", 0.0) < 0
        )

        def _split_on_merged_cells(
            axis: str,
            fixed: int,
            contiguous_cells: list[tuple[int, str, str]],
        ) -> list[list[tuple[int, str, str]]]:
            """Split a contiguous segment so merged top-left cells become single-cell segments."""
            if not merged_top_lefts or len(contiguous_cells) <= 1:
                return [contiguous_cells]

            segments: list[list[tuple[int, str, str]]] = []
            current_segment: list[tuple[int, str, str]] = []
            for pos, text, cell_ref in contiguous_cells:
                cell_key = (fixed, pos) if axis == "row" else (pos, fixed)
                if cell_key in merged_top_lefts:
                    if current_segment:
                        segments.append(current_segment)
                        current_segment = []
                    segments.append([(pos, text, cell_ref)])
                else:
                    current_segment.append((pos, text, cell_ref))
            if current_segment:
                segments.append(current_segment)
            return segments

        row_segments: list[tuple[int, list[tuple[int, str, str]]]] = []
        col_segments: list[tuple[int, list[tuple[int, str, str]]]] = []

        for row in sorted(rows_data):
            for _start, _end, contiguous in self._find_contiguous_ranges(rows_data[row]):
                for seg in _split_on_merged_cells("row", row, contiguous):
                    row_segments.append((row, seg))

        for col in sorted(cols_data):
            for _start, _end, contiguous in self._find_contiguous_ranges(cols_data[col]):
                for seg in _split_on_merged_cells("col", col, contiguous):
                    col_segments.append((col, seg))

        # Choose the batching direction that yields fewer range operations.
        # Column-wise batching drastically reduces COM calls for the common "single text column" case.
        use_column_strategy = len(col_segments) < len(row_segments)
        strategy = "col" if use_column_strategy else "row"

        # Phase 1: Write values (batched).
        if use_column_strategy:
            for col, segment in col_segments:
                if debug_stats:
                    total_segments += 1

                seg_start = segment[0][0]
                seg_end = segment[-1][0]

                t0 = time.perf_counter() if debug_stats else None
                wrote_as_range, _wrote_range = self._write_cell_values_batch_column(
                    sheet, sheet_name, col, seg_start, seg_end, segment
                )
                if debug_stats and t0 is not None:
                    t_values += time.perf_counter() - t0
                    if wrote_as_range:
                        value_range_writes += 1
                    else:
                        value_cell_writes += len(segment)
        else:
            for row, segment in row_segments:
                if debug_stats:
                    total_segments += 1

                seg_start = segment[0][0]
                seg_end = segment[-1][0]

                t0 = time.perf_counter() if debug_stats else None
                wrote_as_range, _wrote_range = self._write_cell_values_batch(
                    sheet, sheet_name, row, seg_start, seg_end, segment
                )
                if debug_stats and t0 is not None:
                    t_values += time.perf_counter() - t0
                    if wrote_as_range:
                        value_range_writes += 1
                    else:
                        value_cell_writes += len(segment)

        # Phase 2: Apply output font name (batched union ranges when possible).
        # This avoids one COM font call per segment for large/fragmented tables.
        merged_top_left_to_address: dict[tuple[int, int], str] = {}
        if merged_map:
            merged_top_left_to_address = {
                (r1, c1): merge_address
                for merge_address, (r1, c1, _r2, _c2) in merged_map.items()
            }

        # Collect font targets from the chosen strategy segments.
        # Each entry: (axis, fixed_index, segment, address)
        font_targets: list[tuple[str, int, list[tuple[int, str, str]], str]] = []
        if use_column_strategy:
            for col, segment in col_segments:
                if not segment:
                    continue
                if len(segment) == 1:
                    row_pos = segment[0][0]
                    address = merged_top_left_to_address.get((row_pos, col))
                    if not address:
                        address = f"{get_column_letter(col)}{row_pos}"
                else:
                    start_row = segment[0][0]
                    end_row = segment[-1][0]
                    col_letter = get_column_letter(col)
                    address = f"{col_letter}{start_row}:{col_letter}{end_row}"
                font_targets.append(("col", col, segment, address))
        else:
            for row, segment in row_segments:
                if not segment:
                    continue
                if len(segment) == 1:
                    col_pos = segment[0][0]
                    address = merged_top_left_to_address.get((row, col_pos))
                    if not address:
                        address = f"{get_column_letter(col_pos)}{row}"
                else:
                    start_col = segment[0][0]
                    end_col = segment[-1][0]
                    address = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
                font_targets.append(("row", row, segment, address))

        def _iter_font_target_chunks(
            targets: list[tuple[str, int, list[tuple[int, str, str]], str]],
            max_items: int = 200,
            # Excel's Range("A1,A2,...") string argument fails around ~255 characters
            # (0x800A03EC / -2146827284). Keep a safe margin.
            max_chars: int = 250,
        ):
            chunk: list[tuple[str, int, list[tuple[int, str, str]], str]] = []
            current_chars = 0
            for target in targets:
                address = target[3]
                if not address:
                    continue
                additional = len(address) + (1 if chunk else 0)
                if chunk and (len(chunk) >= max_items or current_chars + additional > max_chars):
                    yield chunk
                    chunk = []
                    current_chars = 0
                chunk.append(target)
                current_chars += additional
            if chunk:
                yield chunk

        for chunk in _iter_font_target_chunks(font_targets):
            address_str = ",".join(target[3] for target in chunk)
            t0 = time.perf_counter() if debug_stats else None
            try:
                sheet.api.Range(address_str).Font.Name = output_font_name
                if debug_stats and t0 is not None:
                    t_fonts += time.perf_counter() - t0
                    font_range_applies += 1
                continue
            except Exception as e:
                logger.debug(
                    "Batch font apply failed for %d areas in sheet '%s': %s (falling back)",
                    len(chunk), sheet_name, e,
                )
                if debug_stats:
                    font_range_fallbacks += 1

            # Fallback: apply per target (and per cell when range apply fails).
            for axis, fixed, segment, _address in chunk:
                if len(segment) == 1:
                    pos = segment[0][0]
                    row_pos = pos if axis == "col" else fixed
                    col_pos = fixed if axis == "col" else pos
                    cell = sheet.range(row_pos, col_pos)
                    t0 = time.perf_counter() if debug_stats else None
                    try:
                        if (row_pos, col_pos) in merged_top_lefts:
                            try:
                                cell.api.MergeArea.Font.Name = output_font_name
                            except Exception:
                                cell.api.Font.Name = output_font_name
                        else:
                            cell.api.Font.Name = output_font_name
                        if debug_stats and t0 is not None:
                            t_fonts += time.perf_counter() - t0
                            font_cell_applies += 1
                    except Exception as inner_e:
                        logger.debug(
                            "Cell font apply failed for row %d col %d in '%s': %s",
                            row_pos, col_pos, sheet_name, inner_e,
                        )
                    continue

                if axis == "col":
                    col = fixed
                    start_row = segment[0][0]
                    end_row = segment[-1][0]
                    rng = sheet.range((start_row, col), (end_row, col))
                    t0 = time.perf_counter() if debug_stats else None
                    try:
                        rng.api.Font.Name = output_font_name
                        if debug_stats and t0 is not None:
                            t_fonts += time.perf_counter() - t0
                            font_range_applies += 1
                    except Exception as inner_e:
                        logger.debug(
                            "Range font apply failed for col %d rows %d-%d in '%s': %s (falling back to per-cell)",
                            col, start_row, end_row, sheet_name, inner_e,
                        )
                        if debug_stats:
                            font_range_fallbacks += 1
                        for row_pos, _text, _cell_ref in segment:
                            try:
                                t0 = time.perf_counter() if debug_stats else None
                                sheet.range(row_pos, col).api.Font.Name = output_font_name
                                if debug_stats and t0 is not None:
                                    t_fonts += time.perf_counter() - t0
                                    font_cell_applies += 1
                            except Exception as cell_e:
                                logger.debug(
                                    "Cell font apply failed for row %d col %d in '%s': %s",
                                    row_pos, col, sheet_name, cell_e,
                                )
                else:
                    row = fixed
                    start_col = segment[0][0]
                    end_col = segment[-1][0]
                    rng = sheet.range((row, start_col), (row, end_col))
                    t0 = time.perf_counter() if debug_stats else None
                    try:
                        rng.api.Font.Name = output_font_name
                        if debug_stats and t0 is not None:
                            t_fonts += time.perf_counter() - t0
                            font_range_applies += 1
                    except Exception as inner_e:
                        logger.debug(
                            "Range font apply failed for row %d cols %d-%d in '%s': %s (falling back to per-cell)",
                            row, start_col, end_col, sheet_name, inner_e,
                        )
                        if debug_stats:
                            font_range_fallbacks += 1
                        for col_pos, _text, _cell_ref in segment:
                            try:
                                t0 = time.perf_counter() if debug_stats else None
                                sheet.range(row, col_pos).api.Font.Name = output_font_name
                                if debug_stats and t0 is not None:
                                    t_fonts += time.perf_counter() - t0
                                    font_cell_applies += 1
                            except Exception as cell_e:
                                logger.debug(
                                    "Cell font apply failed for row %d col %d in '%s': %s",
                                    row, col_pos, sheet_name, cell_e,
                                )

        # Phase 3 (optional): apply font-size adjustment per translated cell.
        if needs_size_adjustment:
            for (row, col), (_translated_text, cell_ref) in cell_translations.items():
                t0 = time.perf_counter() if debug_stats else None
                try:
                    cell = sheet.range(row, col)
                    try:
                        original_size = float(cell.api.Font.Size)
                    except Exception:
                        continue

                    _, adjusted_size = font_manager.select_font(None, original_size)
                    if adjusted_size == original_size:
                        continue

                    try:
                        # Merged cells: apply to the whole merge area for consistency
                        if getattr(cell.api, "MergeCells", False):
                            cell.api.MergeArea.Font.Size = adjusted_size
                        else:
                            cell.api.Font.Size = adjusted_size
                    except Exception:
                        cell.api.Font.Size = adjusted_size
                except Exception as e:
                    logger.debug(
                        "Font size adjustment failed for %s_%s: %s",
                        sheet_name, cell_ref, e,
                    )
                finally:
                    if debug_stats and t0 is not None:
                        t_sizes += time.perf_counter() - t0

        if debug_stats:
            total_seconds = time.perf_counter() - t_start
            logger.debug(
                "[TIMING] Excel apply (cells) '%s': cells=%d strategy=%s lines=%d segments=%d "
                "values(range=%d cell=%d, %.2fs) fonts(range=%d cell=%d fallback=%d, %.2fs) "
                "sizes(%.2fs) total=%.2fs",
                sheet_name,
                len(cell_translations),
                strategy,
                len(cols_data) if use_column_strategy else len(rows_data),
                total_segments,
                value_range_writes,
                value_cell_writes,
                t_values,
                font_range_applies,
                font_cell_applies,
                font_range_fallbacks,
                t_fonts,
                t_sizes,
                total_seconds,
            )

    def _find_contiguous_ranges(
        self, cells: list[tuple[int, str, str]]
    ) -> list[tuple[int, int, list[tuple[int, str, str]]]]:
        """Find contiguous ranges in a sorted 1D list of cells.

        Args:
            cells: List of (pos, text, cell_ref) sorted by pos

        Returns:
            List of (start_pos, end_pos, cells_in_range)
        """
        if not cells:
            return []

        ranges = []
        current_range = [cells[0]]
        current_start = cells[0][0]

        for i in range(1, len(cells)):
            col = cells[i][0]
            prev_col = cells[i - 1][0]

            if col == prev_col + 1:
                # Contiguous
                current_range.append(cells[i])
            else:
                # Gap - save current range and start new one
                ranges.append((current_start, cells[i - 1][0], current_range))
                current_range = [cells[i]]
                current_start = col

        # Don't forget the last range
        ranges.append((current_start, cells[-1][0], current_range))
        return ranges

    def _write_cell_values_batch(
        self,
        sheet,
        sheet_name: str,
        row: int,
        start_col: int,
        end_col: int,
        cells: list[tuple[int, str, str]],
    ) -> tuple[bool, object | None]:
        """Write cell values in batch (values only, no font).

        Args:
            sheet: xlwings Sheet object
            sheet_name: Name of the sheet (for logging)
            row: Row number
            start_col: Starting column
            end_col: Ending column
            cells: List of (col, text, cell_ref) sorted by col

        Returns:
            (written_as_range, range_obj)

            - written_as_range: True if written as a multi-cell range write.
            - range_obj: 互換性のために残している戻り値（現在は常に None）。
        """
        try:
            if len(cells) == 1:
                # Single cell
                col, text, cell_ref = cells[0]
                final_text = text
                if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                    final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    logger.warning(
                        "Translation truncated for cell %s_%s: %d -> %d chars",
                        sheet_name, cell_ref, len(text), len(final_text)
                    )
                # Use Excel COM Value2 directly (faster than xlwings .value for many small writes)
                sheet.api.Cells(row, col).Value2 = final_text
                return False, None
            else:
                # Multiple contiguous cells - batch write
                values = []
                for col, text, cell_ref in cells:
                    final_text = text
                    if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                        final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                        logger.warning(
                            "Translation truncated for cell %s_%s: %d -> %d chars",
                            sheet_name, cell_ref, len(text), len(final_text)
                        )
                    values.append(final_text)
                # Use Excel COM Value2 directly (faster and avoids unnecessary conversions)
                rng = sheet.api.Range(sheet.api.Cells(row, start_col), sheet.api.Cells(row, end_col))
                rng.Value2 = values
                return True, None
        except Exception as e:
            logger.warning(
                "Error writing values to row %d cols %d-%d in '%s': %s",
                row, start_col, end_col, sheet_name, e
            )
            # Fallback to individual writes
            for col, text, cell_ref in cells:
                try:
                    final_text = text
                    if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                        final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    try:
                        sheet.api.Cells(row, col).Value2 = final_text
                    except Exception as write_e:
                        # Merged cells: Excel may reject writes to non-top-left cells.
                        # Attempt to write to the merge area's top-left instead.
                        try:
                            cell = sheet.api.Cells(row, col)
                            if getattr(cell, "MergeCells", False):
                                merge_area = cell.MergeArea
                                sheet.api.Cells(merge_area.Row, merge_area.Column).Value2 = final_text
                            else:
                                raise write_e
                        except Exception:
                            raise write_e
                except Exception as inner_e:
                    logger.debug("Error writing cell %s_%s: %s", sheet_name, cell_ref, inner_e)
            return False, None

    def _write_cell_values_batch_column(
        self,
        sheet,
        sheet_name: str,
        col: int,
        start_row: int,
        end_row: int,
        cells: list[tuple[int, str, str]],
    ) -> tuple[bool, object | None]:
        """Write cell values in batch for a single column segment (values only, no font).

        Args:
            sheet: xlwings Sheet object
            sheet_name: Name of the sheet (for logging)
            col: Column number
            start_row: Starting row
            end_row: Ending row
            cells: List of (row, text, cell_ref) sorted by row

        Returns:
            (written_as_range, range_obj)

            - written_as_range: True if written as a multi-cell range write.
            - range_obj: 互換性のために残している戻り値（現在は常に None）。
        """
        try:
            if len(cells) == 1:
                row, text, cell_ref = cells[0]
                final_text = text
                if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                    final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    logger.warning(
                        "Translation truncated for cell %s_%s: %d -> %d chars",
                        sheet_name, cell_ref, len(text), len(final_text)
                    )
                # Use Excel COM Value2 directly (faster than xlwings .value for many small writes)
                sheet.api.Cells(row, col).Value2 = final_text
                return False, None

            values: list[list[str]] = []
            for row, text, cell_ref in cells:
                final_text = text
                if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                    final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    logger.warning(
                        "Translation truncated for cell %s_%s: %d -> %d chars",
                        sheet_name, cell_ref, len(text), len(final_text)
                    )
                values.append([final_text])

            rng = sheet.api.Range(sheet.api.Cells(start_row, col), sheet.api.Cells(end_row, col))
            rng.Value2 = values
            return True, None
        except Exception as e:
            logger.warning(
                "Error writing values to col %d rows %d-%d in '%s': %s",
                col, start_row, end_row, sheet_name, e,
            )
            for row, text, cell_ref in cells:
                try:
                    final_text = text
                    if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                        final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    try:
                        sheet.api.Cells(row, col).Value2 = final_text
                    except Exception as write_e:
                        try:
                            cell = sheet.api.Cells(row, col)
                            if getattr(cell, "MergeCells", False):
                                merge_area = cell.MergeArea
                                sheet.api.Cells(merge_area.Row, merge_area.Column).Value2 = final_text
                            else:
                                raise write_e
                        except Exception:
                            raise write_e
                except Exception as inner_e:
                    logger.debug("Error writing cell %s_%s: %s", sheet_name, cell_ref, inner_e)
            return False, None

    def _apply_single_cell_xlwings(
        self,
        sheet,
        sheet_name: str,
        row: int,
        col: int,
        translated_text: str,
        cell_ref: str,
        font_manager: FontManager,
    ) -> None:
        """Apply translation to a single cell."""
        try:
            cell = sheet.range(row, col)
            block_id = f"{sheet_name}_{cell_ref}"

            # Try to get font info from cache
            cached_font = self._font_cache.get(block_id)
            if cached_font:
                original_font_name, original_font_size = cached_font
            else:
                original_font_name = None
                original_font_size = 11.0
                try:
                    original_font_name = cell.font.name
                    original_font_size = cell.font.size or 11.0
                except Exception as e:
                    logger.debug("Error reading font for cell %s: %s", block_id, e)

            new_font_name, new_font_size = font_manager.select_font(
                original_font_name, original_font_size
            )

            # Truncate if needed
            final_text = translated_text
            if translated_text and len(translated_text) > EXCEL_CELL_CHAR_LIMIT:
                final_text = translated_text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                logger.warning(
                    "Translation truncated for cell %s: %d -> %d chars",
                    block_id, len(translated_text), len(final_text)
                )

            cell.value = final_text

            try:
                # Check if cell is part of a merged range
                try:
                    is_merged = cell.api.MergeCells
                except Exception:
                    is_merged = False

                if is_merged:
                    # Apply font to the entire merge area
                    try:
                        merge_area = cell.api.MergeArea
                        merge_area.Font.Name = new_font_name
                        merge_area.Font.Size = new_font_size
                    except Exception as merge_e:
                        logger.debug("Error applying font to merged cell %s: %s", block_id, merge_e)
                else:
                    cell.font.name = new_font_name
                    cell.font.size = new_font_size
            except Exception as e:
                logger.debug("Error applying font to cell %s: %s", block_id, e)

        except Exception as e:
            logger.warning("Error applying translation to cell %s_%s: %s", sheet_name, cell_ref, e)

    def _apply_range_batch_xlwings(
        self,
        sheet,
        sheet_name: str,
        row: int,
        start_col: int,
        end_col: int,
        cells: list[tuple[int, str, str]],
        font_manager: FontManager,
    ) -> None:
        """Apply translations to a contiguous range of cells in one batch.

        This reduces COM calls by writing multiple values at once.
        """
        try:
            # Build the value list for the range
            values = []
            font_info_list = []  # To track font changes needed

            for col, text, cell_ref in cells:
                block_id = f"{sheet_name}_{cell_ref}"

                # Truncate if needed
                final_text = text
                if text and len(text) > EXCEL_CELL_CHAR_LIMIT:
                    final_text = text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                    logger.warning(
                        "Translation truncated for cell %s: %d -> %d chars",
                        block_id, len(text), len(final_text)
                    )
                values.append(final_text)

                # Get font info
                cached_font = self._font_cache.get(block_id)
                if cached_font:
                    original_font_name, original_font_size = cached_font
                else:
                    original_font_name = None
                    original_font_size = 11.0
                font_info_list.append((original_font_name, original_font_size))

            # Write values in batch (single COM call for all values)
            rng = sheet.range((row, start_col), (row, end_col))
            rng.value = values

            # Apply fonts (still need individual font settings, but value write is batched)
            # Get the new font once (same for all cells in batch)
            # Note: font may vary by original font, so we apply individually
            for i, (col, _, cell_ref) in enumerate(cells):
                try:
                    cell = sheet.range(row, col)
                    orig_name, orig_size = font_info_list[i]
                    new_name, new_size = font_manager.select_font(orig_name, orig_size)
                    cell.font.name = new_name
                    cell.font.size = new_size
                except Exception as e:
                    logger.debug("Error applying font to cell %s_%s: %s", sheet_name, cell_ref, e)

        except Exception as e:
            logger.warning(
                "Error applying batch translation to row %d cols %d-%d in '%s': %s",
                row, start_col, end_col, sheet_name, e
            )
            # Fallback to individual cell writes
            for col, text, cell_ref in cells:
                self._apply_single_cell_xlwings(
                    sheet, sheet_name, row, col, text, cell_ref, font_manager
                )

    def _apply_translations_openpyxl(
        self,
        input_path: Path,
        output_path: Path,
        translations: dict[str, str],
        direction: str,
        settings=None,
        selected_sections: Optional[list[int]] = None,
        text_blocks: Optional[list[TextBlock]] = None,
    ) -> None:
        """Apply translations using openpyxl (fallback - cells only)

        Optimized:
        - Uses metadata from text_blocks for precise cell positioning
        - Direct cell access using (row, col) tuples instead of string parsing
        - Font object caching to avoid creating duplicate Font objects
        - Only accesses cells that need translation
        - Only processes selected sheets when selected_sections is specified
        """
        font_manager = FontManager(direction, settings)
        wb = openpyxl.load_workbook(input_path)

        # Font object cache: (name, size, bold, italic, underline, strike, color_rgb) -> Font
        # openpyxl Font objects are immutable, so we can safely reuse them
        font_cache: dict[tuple, Font] = {}

        def get_cached_font(
            name: str,
            size: float,
            bold: bool = False,
            italic: bool = False,
            underline: str | None = None,
            strike: bool = False,
            color=None,
        ) -> Font:
            """Get or create a cached Font object."""
            # Create cache key (color needs special handling)
            color_key = None
            if color is not None:
                # Extract color RGB value for cache key
                color_key = getattr(color, 'rgb', None)

            cache_key = (name, size, bold, italic, underline, strike, color_key)

            if cache_key not in font_cache:
                font_cache[cache_key] = Font(
                    name=name,
                    size=size,
                    bold=bold,
                    italic=italic,
                    underline=underline,
                    strike=strike,
                    color=color,
                )
            return font_cache[cache_key]

        try:
            # Pre-group translations by sheet using metadata when available
            sheet_names = set(wb.sheetnames)
            translations_by_sheet = self._group_translations_by_sheet(
                translations, sheet_names, text_blocks
            )

            # Convert selected_sections to a set for O(1) lookup
            selected_set = set(selected_sections) if selected_sections is not None else None

            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                # Skip sheets not in selected_sections (if specified)
                if selected_set is not None and sheet_idx not in selected_set:
                    continue

                sheet = wb[sheet_name]
                sheet_translations = translations_by_sheet.get(sheet_name, {})
                cell_translations = sheet_translations.get('cells', {})

                # Direct cell access using (row, col) tuples
                for (row_idx, col_idx), (translated_text, cell_ref) in cell_translations.items():
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx)

                        original_font_name = cell.font.name if cell.font else None
                        original_font_size = cell.font.size if cell.font and cell.font.size else 11.0

                        new_font_name, new_font_size = font_manager.select_font(
                            original_font_name,
                            original_font_size
                        )

                        # Check and truncate if exceeds Excel cell limit
                        final_text = translated_text
                        if translated_text and len(translated_text) > EXCEL_CELL_CHAR_LIMIT:
                            final_text = translated_text[:EXCEL_CELL_CHAR_LIMIT - 3] + "..."
                            logger.warning(
                                "Translation truncated for cell %s_%s: %d -> %d chars (Excel limit: %d)",
                                sheet_name, cell_ref, len(translated_text),
                                len(final_text), EXCEL_CELL_CHAR_LIMIT
                            )

                        cell.value = final_text

                        # Use cached font object to avoid creating duplicates
                        if cell.font:
                            cell.font = get_cached_font(
                                name=new_font_name,
                                size=new_font_size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color=cell.font.color,
                            )
                        else:
                            cell.font = get_cached_font(name=new_font_name, size=new_font_size)

                    except Exception as e:
                        logger.warning("Error applying translation to cell %s_%s: %s", sheet_name, cell_ref, e)

            # Clear read_only_recommended to prevent Excel dialog on open
            if hasattr(wb, 'properties') and wb.properties is not None:
                wb.properties.read_only_recommended = False

            wb.save(output_path)
        finally:
            wb.close()

    def create_bilingual_workbook(
        self,
        original_path: Path,
        translated_path: Path,
        output_path: Path,
    ) -> dict[str, int]:
        """
        Create a bilingual workbook with original and translated sheets interleaved.

        Output format:
            Sheet1 (original), Sheet1_translated, Sheet2 (original), Sheet2_translated, ...

        Uses xlwings when available to preserve shapes, charts, and images.
        Falls back to openpyxl (cells only) when xlwings/Excel is not available.

        Args:
            original_path: Path to the original workbook
            translated_path: Path to the translated workbook
            output_path: Path to save the bilingual workbook

        Returns:
            dict with original_sheets, translated_sheets, total_sheets counts
        """
        xw = _get_xlwings()

        if _can_use_xlwings():
            return self._create_bilingual_workbook_xlwings(
                original_path, translated_path, output_path, xw
            )
        else:
            return self._create_bilingual_workbook_openpyxl(
                original_path, translated_path, output_path
            )

    def _create_bilingual_workbook_xlwings(
        self,
        original_path: Path,
        translated_path: Path,
        output_path: Path,
        xw,
    ) -> dict[str, int]:
        """
        Create bilingual workbook using xlwings (preserves shapes, charts, images).

        Uses COM Sheet.Copy() to copy sheets with all content including:
        - Cell values and formatting
        - Shapes (TextBox, etc.)
        - Charts
        - Images
        - Merged cells
        """
        with com_initialized():
            app = _create_excel_app_with_retry(xw)

            # Track opened workbooks for proper cleanup
            original_wb = None
            translated_wb = None
            bilingual_wb = None

            try:
                # Open source workbooks (track each for cleanup)
                original_wb = app.books.open(str(original_path), ignore_read_only_recommended=True)
                translated_wb = app.books.open(str(translated_path), ignore_read_only_recommended=True)

                # SAFETY: Verify we opened the correct workbooks
                _verify_workbook_path(original_wb, original_path, "bilingual_original")
                _verify_workbook_path(translated_wb, translated_path, "bilingual_translated")

                # Create new workbook for bilingual output
                bilingual_wb = app.books.add()

                # Remove default sheets from new workbook (with retry limit to prevent infinite loop)
                max_delete_attempts = 10
                delete_attempts = 0
                while len(bilingual_wb.sheets) > 1 and delete_attempts < max_delete_attempts:
                    try:
                        bilingual_wb.sheets[-1].delete()
                    except Exception as e:
                        logger.debug("Error deleting default sheet: %s", e)
                        break
                    delete_attempts += 1

                existing_names: set[str] = set()
                original_sheets = len(original_wb.sheets)
                translated_sheets = len(translated_wb.sheets)

                # Interleave sheets: original, translated, original, translated, ...
                for i, original_sheet in enumerate(original_wb.sheets):
                    sheet_name = original_sheet.name

                    # Copy original sheet to bilingual workbook
                    safe_orig_name = sanitize_sheet_name(sheet_name)
                    unique_orig_name = _ensure_unique_sheet_name(safe_orig_name, existing_names)

                    # Use COM API to copy sheet (preserves all content) with retry for RPC errors
                    _copy_sheet_with_retry(
                        source_sheet=original_sheet,
                        target_wb=bilingual_wb,
                        position="Before",
                        ref_sheet_index=0,
                        new_name=unique_orig_name,
                    )

                    # Copy translated sheet if exists
                    if i < len(translated_wb.sheets):
                        translated_sheet = translated_wb.sheets[i]
                        trans_title = sanitize_sheet_name(f"{sheet_name}_translated")
                        unique_trans_title = _ensure_unique_sheet_name(trans_title, existing_names)

                        # Copy translated sheet after the original with retry for RPC errors
                        _copy_sheet_with_retry(
                            source_sheet=translated_sheet,
                            target_wb=bilingual_wb,
                            position="After",
                            ref_sheet_index=0,
                            new_name=unique_trans_title,
                        )

                # Handle extra translated sheets if any
                if translated_sheets > original_sheets:
                    for i in range(original_sheets, translated_sheets):
                        translated_sheet = translated_wb.sheets[i]
                        trans_title = sanitize_sheet_name(f"{translated_sheet.name}_translated")
                        unique_trans_title = _ensure_unique_sheet_name(trans_title, existing_names)

                        _copy_sheet_with_retry(
                            source_sheet=translated_sheet,
                            target_wb=bilingual_wb,
                            position="Before",
                            ref_sheet_index=0,
                            new_name=unique_trans_title,
                        )

                # Remove the initial empty sheet if it still exists
                # Check for default sheet names in multiple locales
                default_sheet_prefixes = ("Sheet", "シート", "Feuil", "Hoja", "Blatt", "Foglio")
                for sheet in bilingual_wb.sheets:
                    is_default_name = any(sheet.name.startswith(prefix) for prefix in default_sheet_prefixes)
                    if is_default_name:
                        try:
                            # Check if sheet is empty (no values and no shapes)
                            has_content = sheet.used_range.value is not None
                            if not has_content:
                                sheet.delete()
                                break
                        except Exception as e:
                            logger.debug("Error checking/deleting default sheet '%s': %s", sheet.name, e)
                            break

                # Reorder sheets to interleave correctly
                # Current order might be mixed, need to sort by original index
                self._reorder_bilingual_sheets(bilingual_wb, original_wb.sheets, existing_names)

                # Clear read_only_recommended to prevent Excel dialog on open
                try:
                    bilingual_wb.api.ReadOnlyRecommended = False
                except Exception as e:
                    logger.debug("Could not clear ReadOnlyRecommended: %s", e)

                # Save to output path
                bilingual_wb.save(str(output_path))

                return {
                    'original_sheets': original_sheets,
                    'translated_sheets': translated_sheets,
                    'total_sheets': len(bilingual_wb.sheets),
                }

            finally:
                # Close workbooks in reverse order of opening (safest)
                for wb, name in [
                    (bilingual_wb, "bilingual"),
                    (translated_wb, "translated"),
                    (original_wb, "original"),
                ]:
                    if wb is not None:
                        try:
                            wb.close()
                        except Exception as e:
                            logger.debug("Error closing %s workbook: %s", name, e)

                # Always quit the app
                try:
                    app.quit()
                except Exception as e:
                    logger.debug("Error quitting Excel app: %s", e)

    def _reorder_bilingual_sheets(self, bilingual_wb, original_sheets, existing_names: set[str]) -> None:
        """Reorder sheets in bilingual workbook to interleave original and translated."""
        # Build expected order: Sheet1, Sheet1_translated, Sheet2, Sheet2_translated, ...
        expected_order = []
        for original_sheet in original_sheets:
            sheet_name = original_sheet.name
            safe_orig_name = sanitize_sheet_name(sheet_name)
            trans_name = sanitize_sheet_name(f"{sheet_name}_translated")

            # Find actual names (may have suffix like _1 for uniqueness)
            # Original sheet: must NOT contain '_translated'
            for name in existing_names:
                if '_translated' in name:
                    continue
                # Match exact name or name with uniqueness suffix (_1, _2, etc.)
                if name == safe_orig_name or name.startswith(f"{safe_orig_name}_"):
                    expected_order.append(name)
                    break

            # Translated sheet: must contain '_translated'
            # Handle truncated names: "VeryLong..._translated" may become "VeryLong..."
            # Use removesuffix for proper suffix removal (not rstrip which removes char set)
            trans_base = trans_name[:-3] if trans_name.endswith('...') else trans_name
            for name in existing_names:
                if '_translated' not in name:
                    continue
                # Match exact name or name with uniqueness suffix or truncated base
                if name == trans_name or name.startswith(f"{trans_name}_") or name.startswith(trans_base):
                    expected_order.append(name)
                    break

        # Move sheets to correct positions
        for i, expected_name in enumerate(expected_order):
            for sheet in bilingual_wb.sheets:
                if sheet.name == expected_name:
                    if bilingual_wb.sheets.index(sheet) != i:
                        try:
                            # Move sheet to correct position
                            if i == 0:
                                sheet.api.Move(Before=bilingual_wb.sheets[0].api)
                            else:
                                sheet.api.Move(After=bilingual_wb.sheets[i - 1].api)
                        except Exception as e:
                            logger.debug("Error reordering sheet %s: %s", expected_name, e)
                    break

    def _create_bilingual_workbook_openpyxl(
        self,
        original_path: Path,
        translated_path: Path,
        output_path: Path,
    ) -> dict[str, int]:
        """
        Create bilingual workbook using openpyxl (cells only, no shapes/charts).

        Note: This fallback does not preserve shapes, charts, or images.
        Use xlwings version for full content preservation.
        """
        from copy import copy

        original_wb = openpyxl.load_workbook(original_path)
        translated_wb = openpyxl.load_workbook(translated_path)

        try:
            # Create a new workbook and remove default sheet
            bilingual_wb = openpyxl.Workbook()
            default_sheet = bilingual_wb.active
            bilingual_wb.remove(default_sheet)

            existing_names: set[str] = set()

            original_sheets = len(original_wb.sheetnames)
            translated_sheets = len(translated_wb.sheetnames)

            # Interleave sheets
            for i, sheet_name in enumerate(original_wb.sheetnames):
                # Copy original sheet (sanitize in case original has forbidden chars)
                original_sheet = original_wb[sheet_name]
                safe_orig_name = sanitize_sheet_name(sheet_name)
                unique_orig_name = _ensure_unique_sheet_name(safe_orig_name, existing_names)
                orig_copy = bilingual_wb.create_sheet(title=unique_orig_name)
                self._copy_sheet_content(original_sheet, orig_copy)

                # Copy translated sheet if exists
                if i < len(translated_wb.sheetnames):
                    trans_sheet_name = translated_wb.sheetnames[i]
                    translated_sheet = translated_wb[trans_sheet_name]
                    # Create translated sheet with suffix (sanitize for forbidden chars)
                    trans_title = sanitize_sheet_name(f"{sheet_name}_translated")
                    unique_trans_title = _ensure_unique_sheet_name(trans_title, existing_names)
                    trans_copy = bilingual_wb.create_sheet(title=unique_trans_title)
                    self._copy_sheet_content(translated_sheet, trans_copy)

            # Handle extra translated sheets if any
            if translated_sheets > original_sheets:
                for i in range(original_sheets, translated_sheets):
                    trans_sheet_name = translated_wb.sheetnames[i]
                    translated_sheet = translated_wb[trans_sheet_name]
                    # Sanitize for forbidden chars and length
                    trans_title = sanitize_sheet_name(f"{trans_sheet_name}_translated")
                    unique_trans_title = _ensure_unique_sheet_name(trans_title, existing_names)
                    trans_copy = bilingual_wb.create_sheet(title=unique_trans_title)
                    self._copy_sheet_content(translated_sheet, trans_copy)

            # Clear read_only_recommended to prevent Excel dialog on open
            if hasattr(bilingual_wb, 'properties') and bilingual_wb.properties is not None:
                bilingual_wb.properties.read_only_recommended = False

            bilingual_wb.save(output_path)

            return {
                'original_sheets': original_sheets,
                'translated_sheets': translated_sheets,
                'total_sheets': len(bilingual_wb.sheetnames),
            }
        finally:
            original_wb.close()
            translated_wb.close()

    def _copy_sheet_content(self, source_sheet, target_sheet):
        """Copy content from source sheet to target sheet.

        Optimized:
        - Only processes cells that have values or styles (skips empty cells)
        - Uses direct dimension access to limit iteration range
        """
        from copy import copy

        # Get actual data range to avoid iterating empty cells
        # source_sheet.dimensions returns e.g., "A1:Z100" or None for empty sheets
        dimensions = source_sheet.dimensions
        if not dimensions or dimensions == "A1:A1":
            # Empty or minimal sheet - check if A1 has content
            cell = source_sheet.cell(1, 1)
            if cell.value is None and not cell.has_style:
                # Truly empty sheet, skip cell iteration
                pass
            else:
                # Copy single cell
                target_cell = target_sheet.cell(row=1, column=1)
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.fill = copy(cell.fill)
                    target_cell.border = copy(cell.border)
                    target_cell.alignment = copy(cell.alignment)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
        else:
            # Copy cell values and styles - only cells with content or style
            for row in source_sheet.iter_rows():
                for cell in row:
                    # Skip completely empty cells (no value and no style)
                    if cell.value is None and not cell.has_style:
                        continue

                    target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                    target_cell.value = cell.value

                    # Copy style only if present
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.fill = copy(cell.fill)
                        target_cell.border = copy(cell.border)
                        target_cell.alignment = copy(cell.alignment)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)

        # Copy column widths (only explicitly set widths)
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            if col_dim.width is not None:
                target_sheet.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights (only explicitly set heights)
        for row_num, row_dim in source_sheet.row_dimensions.items():
            if row_dim.height is not None:
                target_sheet.row_dimensions[row_num].height = row_dim.height

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        # Copy conditional formatting rules
        try:
            for cf_range, cf_rules in source_sheet.conditional_formatting._cf_rules.items():
                for rule in cf_rules:
                    target_sheet.conditional_formatting.add(cf_range, rule)
        except (AttributeError, Exception) as e:
            logger.debug("Error copying conditional formatting: %s", e)

        # Copy data validations
        try:
            for dv in source_sheet.data_validations.dataValidation:
                target_sheet.add_data_validation(dv)
        except (AttributeError, Exception) as e:
            logger.debug("Error copying data validations: %s", e)

        # Copy hyperlinks (per-cell)
        # Note: Hyperlinks must be accessed through cells, not via private _hyperlinks dict
        # because _hyperlinks is an internal API that may change between openpyxl versions
        try:
            for row in source_sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                        # Copy hyperlink properties (target, tooltip, etc.)
                        from openpyxl.worksheet.hyperlink import Hyperlink
                        target_cell.hyperlink = Hyperlink(
                            ref=target_cell.coordinate,
                            target=cell.hyperlink.target,
                            tooltip=cell.hyperlink.tooltip,
                            display=cell.hyperlink.display,
                            location=cell.hyperlink.location,
                        )
        except (AttributeError, Exception) as e:
            logger.debug("Error copying hyperlinks: %s", e)

        # Copy comments (iterate cells to find comments)
        # Note: Comments must be copied after cell iteration to avoid issues with
        # cells that haven't been created yet in the target sheet
        try:
            from openpyxl.comments import Comment
            for row in source_sheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                        # Create a new Comment object (comments are mutable and need copying)
                        target_cell.comment = Comment(cell.comment.text, cell.comment.author)
        except (AttributeError, Exception) as e:
            logger.debug("Error copying comments: %s", e)

    def export_glossary_csv(
        self,
        translations: dict[str, str],
        text_blocks: list[TextBlock],
        output_path: Path,
    ) -> dict[str, int]:
        """
        Export source/translation pairs as glossary CSV.

        This method exports original text and translated text pairs in CSV format,
        suitable for use as a reference file in future translations.

        Format:
            原文,訳文,シート,セル
            日本語テキスト,English text,Sheet1,A1
            ...

        Args:
            translations: Mapping of block_id to translated text
            text_blocks: List of TextBlock objects containing original texts
            output_path: Output CSV file path

        Returns:
            Dictionary with export statistics:
            - 'total': Total translation pairs
            - 'exported': Successfully exported pairs
            - 'skipped': Pairs skipped (empty or not found)
        """
        import csv

        # Build lookup from block_id to original text and metadata
        block_lookup: dict[str, tuple[str, str, str]] = {}
        for block in text_blocks:
            sheet_name = block.metadata.get('sheet', '')
            cell_ref = ''
            if block.metadata.get('type') == 'cell':
                row = block.metadata.get('row', 0)
                col = block.metadata.get('col', 0)
                if row and col:
                    cell_ref = f"{get_column_letter(col)}{row}"
            elif block.metadata.get('type') == 'shape':
                shape_name = block.metadata.get('shape_name', '')
                cell_ref = f"Shape:{shape_name}"
            elif block.metadata.get('type') in ('chart_title', 'chart_axis_title'):
                chart_idx = block.metadata.get('chart', 0)
                cell_ref = f"Chart:{chart_idx}"
            block_lookup[block.id] = (block.text, sheet_name, cell_ref)

        stats = {'total': 0, 'exported': 0, 'skipped': 0}

        with output_path.open('w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['原文', '訳文', 'シート', 'セル'])

            for block_id, translated_text in translations.items():
                stats['total'] += 1

                if block_id not in block_lookup:
                    stats['skipped'] += 1
                    continue

                original_text, sheet_name, cell_ref = block_lookup[block_id]

                # Skip empty translations
                if not translated_text or not translated_text.strip():
                    stats['skipped'] += 1
                    continue

                writer.writerow([original_text, translated_text, sheet_name, cell_ref])
                stats['exported'] += 1

        logger.info(
            "Glossary CSV exported: %s (total=%d, exported=%d, skipped=%d)",
            output_path, stats['total'], stats['exported'], stats['skipped']
        )

        return stats
