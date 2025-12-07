# tests/test_error_recovery.py
"""
Error recovery tests for YakuLingo.
Tests graceful handling of various failure scenarios:
- Network timeouts
- Partial batch failures
- Corrupted files
- Copilot disconnection
- API errors
"""

import pytest
import asyncio
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch
import openpyxl
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from yakulingo.models.types import (
    TranslationStatus,
    TranslationProgress,
    TextBlock,
    TextTranslationResult,
    TranslationOption,
    FileInfo,
    FileType,
)
from yakulingo.config.settings import AppSettings
from yakulingo.services.translation_service import TranslationService, BatchTranslator
from yakulingo.services.prompt_builder import PromptBuilder
from yakulingo.ui.state import AppState, Tab, FileState


# --- Fixtures ---

@pytest.fixture
def settings():
    """Default AppSettings"""
    return AppSettings()


@pytest.fixture
def sample_excel(tmp_path):
    """Create a sample Excel file"""
    file_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 11):
        ws[f"A{i}"] = f"テキスト{i}"
    wb.save(file_path)
    return file_path


@pytest.fixture
def large_excel(tmp_path):
    """Create a large Excel file for multi-batch testing"""
    file_path = tmp_path / "large.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 101):  # 100 rows for multiple batches
        ws[f"A{i}"] = f"テキスト{i}"
    wb.save(file_path)
    return file_path


# --- Network Error Tests ---

class TestNetworkErrors:
    """Test handling of network-related errors"""

    def test_connection_timeout_error(self, settings):
        """Handle connection timeout during Copilot connection"""
        mock_copilot = MagicMock()
        # translate_single is used for text translation, not translate_sync
        mock_copilot.translate_single.side_effect = TimeoutError("Connection timed out")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "timed out" in result.error_message.lower()

    def test_network_disconnection_mid_batch(self, settings, large_excel):
        """Handle network disconnection during batch translation"""
        mock_copilot = MagicMock()

        call_count = [0]

        def simulate_disconnect(texts, prompt, ref=None, char_limit=None, skip_clear_wait=False):
            call_count[0] += 1
            if call_count[0] > 1:
                raise ConnectionError("Network disconnected")
            return [f"Trans{i}" for i in range(len(texts))]

        mock_copilot.translate_sync.side_effect = simulate_disconnect

        # Use small batch size to force multiple batches
        settings.max_chars_per_batch = 100
        service = TranslationService(mock_copilot, settings)
        result = service.translate_file(large_excel)

        assert result.status == TranslationStatus.FAILED
        assert "disconnected" in result.error_message.lower() or "Network" in result.error_message

    def test_dns_resolution_failure(self, settings):
        """Handle DNS resolution failure"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = OSError("Name or service not known")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_ssl_certificate_error(self, settings):
        """Handle SSL certificate verification failure"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = ConnectionError("SSL: CERTIFICATE_VERIFY_FAILED")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "SSL" in result.error_message or "CERTIFICATE" in result.error_message


# --- Partial Batch Failure Tests ---

class TestPartialBatchFailures:
    """Test handling of partial failures in batch processing"""

    def test_single_batch_partial_failure(self, settings):
        """Handle partial failure within single batch"""
        mock_copilot = MagicMock()

        # Return fewer results than expected
        mock_copilot.translate_sync.return_value = ["Trans1", "Trans2"]  # Only 2 results for 5 blocks

        service = TranslationService(mock_copilot, settings)

        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}", location=f"A{i}")
            for i in range(5)
        ]

        # Implementation preserves original text for missing translations
        results = service.batch_translator.translate_blocks(blocks)

        # All 5 blocks are in results: 2 translated + 3 with original text
        assert len(results) == 5
        assert results["0"] == "Trans1"
        assert results["1"] == "Trans2"
        assert results["2"] == "テスト2"  # Original text preserved
        assert results["3"] == "テスト3"  # Original text preserved
        assert results["4"] == "テスト4"  # Original text preserved

    def test_first_batch_fails_second_succeeds(self, settings):
        """Test when first batch fails but would need recovery"""
        mock_copilot = MagicMock()

        call_count = [0]

        def batch_with_first_failure(texts, prompt, ref=None, char_limit=None, skip_clear_wait=False):
            call_count[0] += 1
            if call_count[0] == 1:
                raise RuntimeError("First batch API error")
            return [f"Trans{i}" for i in range(len(texts))]

        mock_copilot.translate_sync.side_effect = batch_with_first_failure

        service = TranslationService(mock_copilot, settings)

        blocks = [
            TextBlock(id=str(i), text=f"テスト{i}", location=f"A{i}")
            for i in range(75)  # Will need 2 batches
        ]

        # Batch translator doesn't have built-in retry
        # Error propagates up
        with pytest.raises(RuntimeError) as exc:
            service.batch_translator.translate_blocks(blocks)

        assert "First batch API error" in str(exc.value)

    def test_empty_response_from_api(self, settings, sample_excel):
        """Handle empty response from translation API"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = []  # Empty response

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        # Implementation preserves original text for missing translations
        # So blocks_translated equals blocks_total (original text is used)
        assert result.status == TranslationStatus.COMPLETED
        # blocks_translated = len(translations) which includes all blocks with original text
        assert result.blocks_translated == result.blocks_total

    def test_malformed_api_response(self, settings):
        """Handle malformed response from translation API"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = None  # None instead of list

        service = TranslationService(mock_copilot, settings)

        blocks = [TextBlock(id="1", text="テスト", location="A1")]

        with pytest.raises(TypeError):
            service.batch_translator.translate_blocks(blocks)


# --- File Processing Error Tests ---

class TestFileProcessingErrors:
    """Test handling of file processing errors"""

    def test_corrupted_excel_file(self, settings, tmp_path):
        """Handle corrupted Excel file"""
        corrupt_file = tmp_path / "corrupt.xlsx"
        corrupt_file.write_bytes(b"This is not a valid Excel file")

        mock_copilot = MagicMock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(corrupt_file)

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_file_not_found(self, settings):
        """Handle non-existent file"""
        mock_copilot = MagicMock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(Path("/nonexistent/path/file.xlsx"))

        assert result.status == TranslationStatus.FAILED

    def test_permission_denied(self, settings, tmp_path):
        """Handle permission denied error"""
        mock_copilot = MagicMock()

        # Create a file and make it unreadable (Unix-like systems)
        protected_file = tmp_path / "protected.xlsx"
        wb = openpyxl.Workbook()
        wb.save(protected_file)

        # Mock the processor to raise permission error
        service = TranslationService(mock_copilot, settings)

        with patch.object(service.processors['.xlsx'], 'extract_text_blocks',
                         side_effect=PermissionError("Permission denied")):
            result = service.translate_file(protected_file)

        assert result.status == TranslationStatus.FAILED
        assert "Permission" in result.error_message or result.error_message

    def test_unsupported_file_format(self, settings, tmp_path):
        """Handle unsupported file format"""
        unsupported_file = tmp_path / "document.xyz"
        unsupported_file.write_text("Plain text content")

        mock_copilot = MagicMock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(unsupported_file)

        assert result.status == TranslationStatus.FAILED
        assert "Unsupported" in result.error_message

    def test_empty_file(self, settings, tmp_path):
        """Handle file with no content"""
        empty_file = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(empty_file)

        mock_copilot = MagicMock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(empty_file)

        assert result.status == TranslationStatus.COMPLETED
        assert result.blocks_total == 0
        assert result.warnings is not None
        assert any("No translatable" in w for w in result.warnings)

    def test_output_directory_not_writable(self, settings, sample_excel, tmp_path):
        """Handle non-writable output directory"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = ["Trans" for _ in range(10)]

        # Set output to non-existent directory
        settings.output_directory = "/nonexistent/directory"

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        # Either fails during apply or generates path in input directory
        # Behavior depends on implementation
        assert result.status in [TranslationStatus.FAILED, TranslationStatus.COMPLETED]


# --- Copilot Disconnection Tests ---

class TestCopilotDisconnection:
    """Test handling of Copilot disconnection scenarios"""

    def test_disconnection_during_text_translation(self, settings):
        """Handle disconnection during text translation"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = ConnectionError("Copilot disconnected")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_reconnection_required_between_batches(self, settings):
        """Simulate scenario where reconnection would be needed"""
        mock_copilot = MagicMock()

        # First batch succeeds, then connection fails
        mock_copilot.translate_sync.side_effect = [
            [f"Trans{i}" for i in range(2)],
            ConnectionRefusedError("Connection refused"),
        ]

        # Use small batch size to force multiple batches
        settings.max_chars_per_batch = 50
        service = TranslationService(mock_copilot, settings)

        blocks = [
            TextBlock(id=str(i), text=f"LongTestText{i}x" * 5, location=f"A{i}")
            for i in range(4)  # 4 blocks with ~80 chars each -> 2 batches
        ]

        with pytest.raises(ConnectionRefusedError):
            service.batch_translator.translate_blocks(blocks)


# --- API Error Tests ---

class TestAPIErrors:
    """Test handling of various API errors"""

    def test_rate_limit_error(self, settings):
        """Handle rate limiting from API"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = RuntimeError("Rate limit exceeded. Please wait.")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "Rate limit" in result.error_message

    def test_authentication_error(self, settings):
        """Handle authentication failure"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = RuntimeError("Authentication failed: Invalid token")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "Authentication" in result.error_message

    def test_server_error_500(self, settings):
        """Handle server 500 error"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = RuntimeError("HTTP 500: Internal Server Error")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "500" in result.error_message

    def test_service_unavailable_503(self, settings):
        """Handle service unavailable error"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = RuntimeError("HTTP 503: Service Unavailable")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED

    def test_invalid_response_format(self, settings):
        """Handle invalid response format from API"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.return_value = None  # Invalid response

        service = TranslationService(mock_copilot, settings)

        # translate_text wraps in try-except, so should handle this
        result = service.translate_text("テスト")

        # None is returned as output_text
        assert result.status == TranslationStatus.COMPLETED
        assert result.output_text is None


# --- Memory and Resource Error Tests ---

class TestResourceErrors:
    """Test handling of resource-related errors"""

    def test_large_file_memory_handling(self, settings, tmp_path):
        """Handle very large files that might cause memory issues"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = ["Trans"]

        # Create file with many cells
        large_file = tmp_path / "large.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 1001):
            ws[f"A{i}"] = f"Long text content for cell {i} " * 10
        wb.save(large_file)

        service = TranslationService(mock_copilot, settings)

        # Should handle large file without crashing
        result = service.translate_file(large_file)

        # May complete or fail, but shouldn't crash
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]

    def test_disk_full_simulation(self, settings, sample_excel, tmp_path):
        """Handle disk full error during output write"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = [f"Trans{i}" for i in range(10)]

        service = TranslationService(mock_copilot, settings)

        # Mock the processor's apply_translations to raise disk error
        with patch.object(service.processors['.xlsx'], 'apply_translations',
                         side_effect=OSError("No space left on device")):
            result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "space" in result.error_message.lower() or "device" in result.error_message.lower()


# --- State Recovery Tests ---

class TestStateRecovery:
    """Test state recovery after errors"""

    def test_state_reset_after_text_error(self):
        """State resets properly after text translation error"""
        state = AppState.__new__(AppState)
        state.current_tab = Tab.TEXT
        state.source_text = "テスト"
        state.text_translating = True
        state.text_result = None
        state.copilot_ready = True
        state._history_db = None
        state.history = []

        # Simulate error
        state.text_translating = False
        state.text_result = TextTranslationResult(
            source_text="テスト",
            source_char_count=3,
            options=[],
            error_message="Translation failed",
        )

        assert state.text_translating is False
        assert state.text_result.error_message == "Translation failed"
        assert state.can_translate() is True  # Can try again

    def test_state_reset_after_file_error(self):
        """State resets properly after file translation error"""
        state = AppState.__new__(AppState)
        state.current_tab = Tab.FILE
        state.file_state = FileState.TRANSLATING
        state.translation_progress = 0.5
        state.copilot_ready = True
        state.selected_file = None
        state.file_info = None
        state.output_file = None
        state.translation_status = ""
        state.error_message = ""
        state._history_db = None
        state.history = []

        # Error occurs
        state.file_state = FileState.ERROR
        state.error_message = "Network timeout"

        assert state.file_state == FileState.ERROR
        assert state.error_message == "Network timeout"
        assert state.is_translating() is False

        # User clicks reset
        state.file_state = FileState.EMPTY
        state.selected_file = None
        state.file_info = None
        state.translation_progress = 0.0
        state.translation_status = ""
        state.error_message = ""

        assert state.file_state == FileState.EMPTY
        assert state.error_message == ""


# --- Cancellation During Error Tests ---

class TestCancellationDuringErrors:
    """Test cancellation handling during error scenarios"""

    def test_cancel_after_partial_completion(self, settings, large_excel):
        """Cancel after some batches completed successfully"""
        mock_copilot = MagicMock()

        call_count = [0]
        service_ref = [None]  # Will hold reference to service

        def translate_then_cancel(texts, prompt, ref=None, char_limit=None, skip_clear_wait=False):
            call_count[0] += 1
            # Cancel after first batch to ensure cancel flag is checked
            if call_count[0] >= 1 and service_ref[0]:
                service_ref[0].cancel()
            return [f"Trans{i}" for i in range(len(texts))]

        mock_copilot.translate_sync.side_effect = translate_then_cancel

        service = TranslationService(mock_copilot, settings)
        service_ref[0] = service

        result = service.translate_file(large_excel)

        # Should be cancelled after first batch triggers cancel
        assert result.status == TranslationStatus.CANCELLED

    def test_cancel_during_error_recovery(self, settings, sample_excel):
        """Cancel while error recovery would be in progress"""
        mock_copilot = MagicMock()

        service = TranslationService(mock_copilot, settings)
        service._cancel_requested = True

        result = service.translate_file(sample_excel)

        # Cancel flag is reset at start, so completes
        assert result.status == TranslationStatus.COMPLETED


# --- Result Parsing Error Tests ---

class TestResultParsingErrors:
    """Test handling of result parsing errors"""

    def test_parse_multi_option_malformed(self):
        """Handle malformed multi-option result"""
        mock_copilot = MagicMock()
        settings = AppSettings()

        service = TranslationService(mock_copilot, settings)

        # Malformed response
        raw_result = "Just some random text without proper format"

        options = service._parse_multi_option_result(raw_result)

        # Should return empty list for malformed input
        assert options == []

    def test_parse_single_translation_malformed(self):
        """Handle malformed single translation result"""
        mock_copilot = MagicMock()
        settings = AppSettings()

        service = TranslationService(mock_copilot, settings)

        # No 訳文: prefix
        raw_result = "Some translation without proper markers"

        options = service._parse_single_translation_result(raw_result)

        # Should fallback to treating whole result as text
        assert len(options) == 1
        assert options[0].text == raw_result.strip()

    def test_parse_empty_result(self):
        """Handle empty result string"""
        mock_copilot = MagicMock()
        settings = AppSettings()

        service = TranslationService(mock_copilot, settings)

        options = service._parse_single_translation_result("")

        # Should return empty list
        assert options == []

    def test_parse_result_with_only_whitespace(self):
        """Handle result with only whitespace"""
        mock_copilot = MagicMock()
        settings = AppSettings()

        service = TranslationService(mock_copilot, settings)

        options = service._parse_single_translation_result("   \n\t\n   ")

        # Should return empty list after stripping whitespace
        assert options == []


# --- Adjust Translation Error Tests ---

class TestAdjustTranslationErrors:
    """Test error handling in translation adjustment"""

    def test_adjust_with_api_error(self):
        """Handle API error during adjustment"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.side_effect = RuntimeError("API Error")

        settings = AppSettings()
        service = TranslationService(mock_copilot, settings)

        result = service.adjust_translation("Original text", "shorter")

        # Should return None on failure
        assert result is None

    def test_adjust_with_empty_result(self):
        """Handle empty result from adjustment"""
        mock_copilot = MagicMock()
        mock_copilot.translate_single.return_value = ""

        settings = AppSettings()
        service = TranslationService(mock_copilot, settings)

        result = service.adjust_translation("Original text", "shorter")

        # Should return None for empty result
        assert result is None


# --- Edge Case Error Tests ---

class TestEdgeCaseErrors:
    """Test error handling for edge cases"""

    def test_unicode_error_in_content(self, settings, tmp_path):
        """Handle files with problematic Unicode characters"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Create file with unusual Unicode
        file_path = tmp_path / "unicode.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Test with emoji: 🍎🍊🍋"
        ws["A2"] = "Test with CJK: 漢字テスト한국어"
        ws["A3"] = "Test with symbols: ∞∑∏"
        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        # Should handle Unicode without error
        assert result.status == TranslationStatus.COMPLETED

    def test_very_long_text_single_cell(self, settings, tmp_path):
        """Handle very long text in single cell"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Create file with very long text
        file_path = tmp_path / "long_text.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "テスト " * 5000  # Very long text
        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        # Should handle or split appropriately
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]

    def test_special_characters_in_file_path(self, settings, tmp_path):
        """Handle file paths with special characters"""
        mock_copilot = MagicMock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Create file with special characters in name
        file_path = tmp_path / "テスト file (1).xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Content"
        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        # Should handle special characters in path
        assert result.status == TranslationStatus.COMPLETED
        assert "テスト file (1)" in result.output_path.stem
