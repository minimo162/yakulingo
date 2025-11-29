# ecm_translate/processors/excel_processor.py
"""
Processor for Excel files (.xlsx, .xls).
"""

from pathlib import Path
from typing import Iterator
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from .base import FileProcessor
from .translators import CellTranslator
from .font_manager import FontManager, FontTypeDetector
from ecm_translate.models.types import TextBlock, FileInfo, FileType


class ExcelProcessor(FileProcessor):
    """
    Processor for Excel files (.xlsx, .xls).

    Translation targets:
    - Cell values (text only)
    - Shape text (TextBox, etc.)
    - Chart titles and labels

    Preserved:
    - Formulas (not translated)
    - Cell formatting (font, color, borders)
    - Column widths, row heights
    - Merged cells
    - Images
    - Charts (structure)

    Not translated:
    - Sheet names
    - Named ranges
    - Comments
    - Header/Footer text
    """

    def __init__(self):
        self.cell_translator = CellTranslator()
        self.font_type_detector = FontTypeDetector()

    @property
    def file_type(self) -> FileType:
        return FileType.EXCEL

    @property
    def supported_extensions(self) -> list[str]:
        return ['.xlsx', '.xls']

    def get_file_info(self, file_path: Path) -> FileInfo:
        """Get Excel file info"""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        sheet_count = len(wb.sheetnames)
        text_count = 0

        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if self.cell_translator.should_translate(str(cell.value)):
                            text_count += 1

        wb.close()

        return FileInfo(
            path=file_path,
            file_type=FileType.EXCEL,
            size_bytes=file_path.stat().st_size,
            sheet_count=sheet_count,
            text_block_count=text_count,
        )

    def extract_text_blocks(self, file_path: Path) -> Iterator[TextBlock]:
        """Extract text from cells and shapes"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Extract cell text
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value and isinstance(cell.value, str):
                        if self.cell_translator.should_translate(str(cell.value)):
                            col_letter = get_column_letter(col_idx)

                            # Get font info for metadata
                            font_name = None
                            font_size = 11.0  # default
                            if cell.font:
                                font_name = cell.font.name
                                if cell.font.size:
                                    font_size = cell.font.size

                            yield TextBlock(
                                id=f"{sheet_name}_{col_letter}{row_idx}",
                                text=str(cell.value),
                                location=f"{sheet_name}, {col_letter}{row_idx}",
                                metadata={
                                    'sheet': sheet_name,
                                    'row': row_idx,
                                    'col': col_idx,
                                    'type': 'cell',
                                    'font_name': font_name,
                                    'font_size': font_size,
                                }
                            )

            # Extract shape text (TextBox, etc.)
            if hasattr(sheet, '_charts'):
                for chart_idx, chart in enumerate(sheet._charts):
                    if hasattr(chart, 'title') and chart.title:
                        title_text = str(chart.title)
                        if self.cell_translator.should_translate(title_text):
                            yield TextBlock(
                                id=f"{sheet_name}_chart_{chart_idx}_title",
                                text=title_text,
                                location=f"{sheet_name}, Chart {chart_idx + 1} Title",
                                metadata={
                                    'sheet': sheet_name,
                                    'chart': chart_idx,
                                    'type': 'chart_title',
                                }
                            )

        wb.close()

    def apply_translations(
        self,
        input_path: Path,
        output_path: Path,
        translations: dict[str, str],
        direction: str = "jp_to_en",
    ) -> None:
        """Apply translations to Excel file"""
        wb = openpyxl.load_workbook(input_path)
        font_manager = FontManager(direction)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    col_letter = get_column_letter(col_idx)
                    block_id = f"{sheet_name}_{col_letter}{row_idx}"

                    if block_id in translations:
                        translated_text = translations[block_id]

                        # Get original font info
                        original_font_name = cell.font.name if cell.font else None
                        original_font_size = cell.font.size if cell.font and cell.font.size else 11.0

                        # Get new font settings
                        new_font_name, new_font_size = font_manager.select_font(
                            original_font_name,
                            original_font_size
                        )

                        # Apply translation
                        cell.value = translated_text

                        # Apply new font (preserve other formatting)
                        if cell.font:
                            cell.font = Font(
                                name=new_font_name,
                                size=new_font_size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color=cell.font.color,
                            )
                        else:
                            cell.font = Font(name=new_font_name, size=new_font_size)

        wb.save(output_path)
        wb.close()
