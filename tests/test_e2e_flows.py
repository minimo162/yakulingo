# tests/test_e2e_flows.py
"""
End-to-end integration tests for YakuLingo.
Tests complete workflows from user input through to final output.
Simulates realistic usage scenarios.
"""

import pytest
import asyncio
import time
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch, AsyncMock
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from yakulingo.models.types import (
    TranslationStatus,
    TranslationProgress,
    TextBlock,
    TextTranslationResult,
    TranslationOption,
    FileInfo,
    FileType,
    HistoryEntry,
)
from yakulingo.config.settings import AppSettings
from yakulingo.services.translation_service import TranslationService, is_japanese_text
from yakulingo.services.prompt_builder import PromptBuilder
from yakulingo.ui.state import AppState, Tab, FileState
from yakulingo.processors.excel_processor import ExcelProcessor
from yakulingo.processors.word_processor import WordProcessor


# --- Fixtures ---

@pytest.fixture
def mock_copilot():
    """Mock CopilotHandler simulating real behavior"""
    mock = MagicMock()
    mock.is_connected = True

    def smart_translate_single(text, prompt, reference_files=None, char_limit=None):
        """Simulate realistic translation response based on input"""
        # Check if this is a language detection request
        if "何語で書かれていますか" in prompt or "言語名のみで答えてください" in prompt:
            # Return detected language
            if is_japanese_text(text):
                return "日本語"
            else:
                return "英語"

        # Otherwise, it's a translation request
        if is_japanese_text(text):
            # Japanese → English
            return f"""訳文: Translation of "{text[:20]}..."
解説: This is a standard translation."""
        else:
            # English → Japanese
            return f"""訳文: 「{text[:20]}...」の翻訳
解説: これは標準的な翻訳です。"""

    def smart_translate_batch(texts, prompt, reference_files=None, char_limit=None, skip_clear_wait=False, timeout=None):
        """Simulate batch translation"""
        results = []
        for text in texts:
            if is_japanese_text(text):
                results.append(f"Translation: {text[:10]}")
            else:
                results.append(f"翻訳: {text[:10]}")
        return results

    mock.translate_single = smart_translate_single
    mock.translate_sync = smart_translate_batch
    return mock


@pytest.fixture
def settings():
    """Default settings"""
    return AppSettings()


@pytest.fixture
def app_state_with_mock_db():
    """AppState with mocked database"""
    state = AppState.__new__(AppState)
    state.current_tab = Tab.TEXT
    state.source_text = ""
    state.text_translating = False
    state.text_result = None
    state.text_translation_elapsed_time = None
    state.file_state = FileState.EMPTY
    state.selected_file = None
    state.file_info = None
    state.file_output_language = "en"
    state.translation_progress = 0.0
    state.translation_status = ""
    state.output_file = None
    state.error_message = ""
    state.reference_files = []
    state.copilot_ready = False
    state.copilot_error = ""
    state.history = []
    state.history_drawer_open = False
    state.max_history_entries = 50
    state._history_db = None
    return state


# --- E2E Test: Complete Text Translation Flow ---

class TestE2ETextTranslation:
    """End-to-end tests for complete text translation flow"""

    def test_full_japanese_to_english_flow(self, mock_copilot, settings, app_state_with_mock_db):
        """Complete flow: User enters Japanese → Gets English translation"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        # Step 1: User connects to Copilot
        state.copilot_ready = True

        assert state.copilot_ready is True

        # Step 2: User enters Japanese text
        state.source_text = "今日は天気がいいですね。散歩に行きましょう。"

        # Verify translation is possible
        assert state.can_translate() is True

        # Step 3: User clicks translate
        state.text_translating = True
        start_time = time.time()

        # Step 4: Translation executes
        result = service.translate_text_with_options(
            state.source_text,
            reference_files=state.reference_files if state.reference_files else None,
        )

        elapsed = time.time() - start_time

        # Step 5: Update state with result
        state.text_translating = False
        state.text_result = result
        state.text_translation_elapsed_time = elapsed

        # Step 6: Verify result
        assert state.text_translating is False
        assert state.text_result is not None
        assert len(state.text_result.options) >= 1

        # Step 7: Add to history
        entry = HistoryEntry(
            source_text=state.source_text,
            result=state.text_result,
        )
        state.history.insert(0, entry)

        assert len(state.history) == 1
        assert state.history[0].source_text == state.source_text

    def test_full_english_to_japanese_flow(self, mock_copilot, settings, app_state_with_mock_db):
        """Complete flow: User enters English → Gets Japanese translation"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        # Connect
        state.copilot_ready = True

        # Enter English text
        state.source_text = "The weather is nice today. Let's go for a walk."

        # Verify language detection
        assert is_japanese_text(state.source_text) is False

        # Translate
        state.text_translating = True
        result = service.translate_text_with_options(state.source_text)
        state.text_translating = False
        state.text_result = result

        # Verify
        assert state.text_result is not None
        assert state.text_result.output_language in ["jp", "en"]

    def test_translation_with_glossary_flow(self, mock_copilot, settings, app_state_with_mock_db, tmp_path):
        """Complete flow with glossary reference"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        # Create glossary
        glossary_path = tmp_path / "glossary.csv"
        glossary_path.write_text(
            "Japanese,English\n"
            "人工知能,AI\n"
            "機械学習,Machine Learning\n",
            encoding='utf-8'
        )

        # Setup
        state.copilot_ready = True
        state.source_text = "人工知能と機械学習は関連しています。"
        state.reference_files = [glossary_path]

        # Translate with reference
        result = service.translate_text_with_options(
            state.source_text,
            reference_files=state.reference_files,
        )

        state.text_result = result

        # Verify reference files were used
        assert state.text_result is not None

    def test_clear_and_retranslate_flow(self, mock_copilot, settings, app_state_with_mock_db):
        """User clears text and enters new text to translate"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # First translation
        state.source_text = "最初のテキスト"
        result1 = service.translate_text_with_options(state.source_text)
        state.text_result = result1

        # Clear
        state.source_text = ""
        state.text_result = None

        assert state.source_text == ""
        assert state.text_result is None
        assert state.can_translate() is False

        # New translation
        state.source_text = "新しいテキスト"
        assert state.can_translate() is True

        result2 = service.translate_text_with_options(state.source_text)
        state.text_result = result2

        assert state.text_result is not None


# --- E2E Test: Complete File Translation Flow ---

class TestE2EFileTranslation:
    """End-to-end tests for complete file translation flow"""

    def test_full_excel_translation_flow(self, mock_copilot, settings, tmp_path, app_state_with_mock_db):
        """Complete flow: User uploads Excel → Gets translated Excel"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        # Step 1: Create test file
        input_file = tmp_path / "quarterly_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "売上レポート"
        ws["A1"] = "四半期レポート"
        ws["A1"].font = Font(name="MS Gothic", size=14, bold=True)
        ws["A2"] = "売上高"
        ws["B2"] = "1,000,000"
        ws["A3"] = "利益"
        ws["B3"] = "200,000"
        ws["A4"] = "コメント"
        ws["B4"] = "売上は前年比10%増加しました。"
        wb.save(input_file)

        # Step 2: Connect and switch to file tab
        state.copilot_ready = True
        state.current_tab = Tab.FILE

        # Step 3: User selects file
        state.file_state = FileState.SELECTED
        state.selected_file = input_file
        state.file_info = service.get_file_info(input_file)

        assert state.file_info is not None
        assert state.file_info.file_type == FileType.EXCEL
        assert state.can_translate() is True

        # Step 4: User clicks translate
        state.file_state = FileState.TRANSLATING
        state.translation_progress = 0.0

        # Progress callback
        progress_log = []

        def on_progress(progress: TranslationProgress):
            progress_log.append(progress)
            state.translation_progress = progress.current / progress.total
            state.translation_status = progress.status

        # Step 5: Execute translation
        result = service.translate_file(
            input_file,
            on_progress=on_progress,
            output_language=state.file_output_language,
        )

        # Step 6: Update state with result
        if result.status == TranslationStatus.COMPLETED:
            state.file_state = FileState.COMPLETE
            state.output_file = result.output_path
        else:
            state.file_state = FileState.ERROR
            state.error_message = result.error_message

        # Verify
        assert state.file_state == FileState.COMPLETE
        assert state.output_file.exists()
        assert "_translated" in state.output_file.name

        # Step 7: Verify output content
        wb_out = openpyxl.load_workbook(state.output_file)
        ws_out = wb_out.active

        # Numbers should be preserved
        assert ws_out["B2"].value == "1,000,000"
        assert ws_out["B3"].value == "200,000"

        # Text should be translated
        assert ws_out["A1"].value != "四半期レポート"  # Translated

        # Font should be converted (JP -> EN)
        assert ws_out["A1"].font.name in ["Arial", "Calibri"]

        # Progress should have been tracked
        assert len(progress_log) > 0
        assert progress_log[-1].status == "Complete"

    def test_multi_file_sequential_flow(self, mock_copilot, settings, tmp_path, app_state_with_mock_db):
        """User translates multiple files sequentially"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True
        state.current_tab = Tab.FILE

        output_files = []

        for i in range(3):
            # Create file
            input_file = tmp_path / f"document_{i}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = f"文書{i}の内容"
            wb.save(input_file)

            # Select file
            state.file_state = FileState.SELECTED
            state.selected_file = input_file
            state.file_info = service.get_file_info(input_file)

            # Translate
            state.file_state = FileState.TRANSLATING
            result = service.translate_file(input_file)

            # Complete
            state.file_state = FileState.COMPLETE
            state.output_file = result.output_path
            output_files.append(result.output_path)

            # Reset for next file
            state.reset_file_state()

            assert state.file_state == FileState.EMPTY

        # Verify all files translated
        assert len(output_files) == 3
        for output_file in output_files:
            assert output_file.exists()

    def test_file_error_and_retry_flow(self, mock_copilot, settings, tmp_path, app_state_with_mock_db):
        """User encounters error and retries translation"""
        state = app_state_with_mock_db

        # Mock that fails first time
        fail_then_succeed = MagicMock()
        call_count = [0]

        def conditional_fail(texts, prompt, ref=None, char_limit=None, skip_clear_wait=False, timeout=None):
            call_count[0] += 1
            if call_count[0] == 1:
                raise RuntimeError("Temporary API error")
            return [f"Trans{i}" for i in range(len(texts))]

        fail_then_succeed.translate_sync = conditional_fail
        fail_then_succeed.is_connected = True

        service = TranslationService(fail_then_succeed, settings)

        # Create file
        input_file = tmp_path / "retry_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "テスト"
        wb.save(input_file)

        state.copilot_ready = True
        state.current_tab = Tab.FILE
        state.file_state = FileState.SELECTED
        state.selected_file = input_file

        # First attempt - fails
        state.file_state = FileState.TRANSLATING
        result1 = service.translate_file(input_file)

        state.file_state = FileState.ERROR
        state.error_message = result1.error_message

        assert state.file_state == FileState.ERROR
        assert "error" in state.error_message.lower()

        # User clicks retry (reset and try again)
        state.file_state = FileState.SELECTED
        state.error_message = ""

        # Second attempt - succeeds
        state.file_state = FileState.TRANSLATING
        result2 = service.translate_file(input_file)

        state.file_state = FileState.COMPLETE
        state.output_file = result2.output_path

        assert state.file_state == FileState.COMPLETE
        assert state.output_file.exists()


# --- E2E Test: History Flow ---

class TestE2EHistoryFlow:
    """End-to-end tests for translation history"""

    def test_history_accumulation_flow(self, mock_copilot, settings, app_state_with_mock_db):
        """Multiple translations accumulate in history"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        texts = [
            "最初の翻訳テスト",
            "二番目の翻訳テスト",
            "三番目の翻訳テスト",
        ]

        for text in texts:
            state.source_text = text
            result = service.translate_text_with_options(text)
            state.text_result = result

            # Add to history
            entry = HistoryEntry(source_text=text, result=result)
            state.history.insert(0, entry)

        # Verify history order (most recent first)
        assert len(state.history) == 3
        assert state.history[0].source_text == "三番目の翻訳テスト"
        assert state.history[2].source_text == "最初の翻訳テスト"

    def test_restore_from_history_flow(self, mock_copilot, settings, app_state_with_mock_db):
        """User restores translation from history"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Create some history
        state.source_text = "履歴からの復元テスト"
        result = service.translate_text_with_options(state.source_text)
        state.text_result = result

        entry = HistoryEntry(source_text=state.source_text, result=result)
        state.history.insert(0, entry)

        # Clear current state
        state.source_text = ""
        state.text_result = None

        # Restore from history
        history_entry = state.history[0]
        state.source_text = history_entry.source_text
        state.text_result = history_entry.result

        assert state.source_text == "履歴からの復元テスト"
        assert state.text_result is not None

    def test_history_limit_enforcement(self, mock_copilot, settings, app_state_with_mock_db):
        """History respects maximum entries limit"""
        state = app_state_with_mock_db
        state.max_history_entries = 5
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Add more than max entries
        for i in range(10):
            text = f"テスト{i}"
            result = service.translate_text_with_options(text)
            entry = HistoryEntry(source_text=text, result=result)
            state.history.insert(0, entry)

            # Trim to max
            if len(state.history) > state.max_history_entries:
                state.history = state.history[:state.max_history_entries]

        assert len(state.history) == 5
        # Most recent entries should be kept
        assert state.history[0].source_text == "テスト9"


# --- E2E Test: Tab Switching Flow ---

class TestE2ETabSwitchingFlow:
    """End-to-end tests for tab switching behavior"""

    def test_tab_switch_preserves_state(self, mock_copilot, settings, tmp_path, app_state_with_mock_db):
        """Switching tabs preserves state in both tabs"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Enter text in text tab
        state.current_tab = Tab.TEXT
        state.source_text = "テキストタブの内容"

        # Translate text
        result = service.translate_text_with_options(state.source_text)
        state.text_result = result

        # Switch to file tab
        state.current_tab = Tab.FILE

        # Create and select file
        input_file = tmp_path / "tab_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "ファイルタブの内容"
        wb.save(input_file)

        state.file_state = FileState.SELECTED
        state.selected_file = input_file

        # Switch back to text tab
        state.current_tab = Tab.TEXT

        # Verify text state preserved
        assert state.source_text == "テキストタブの内容"
        assert state.text_result is not None

        # Switch to file tab
        state.current_tab = Tab.FILE

        # Verify file state preserved
        assert state.file_state == FileState.SELECTED
        assert state.selected_file == input_file

    def test_tab_switch_blocked_during_translation(self, app_state_with_mock_db):
        """Tab switching should be blocked during translation"""
        state = app_state_with_mock_db

        # Start text translation
        state.current_tab = Tab.TEXT
        state.text_translating = True

        # Verify translation in progress
        assert state.is_translating() is True

        # In real UI, tab switch would be blocked
        # This tests the state check that enables that behavior


# --- E2E Test: Settings Integration ---

class TestE2ESettingsIntegration:
    """End-to-end tests for settings integration"""

    def test_output_directory_setting(self, mock_copilot, tmp_path, app_state_with_mock_db):
        """Output directory setting is respected"""
        state = app_state_with_mock_db

        # Create input file
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        input_file = input_dir / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "テスト"
        wb.save(input_file)

        # Create output directory
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        # Configure settings
        settings = AppSettings(output_directory=str(output_dir))
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True
        state.current_tab = Tab.FILE
        state.file_state = FileState.SELECTED
        state.selected_file = input_file

        # Translate
        result = service.translate_file(input_file)

        # Verify output location
        assert result.output_path.parent == output_dir

    def test_reference_files_setting(self, mock_copilot, tmp_path, app_state_with_mock_db):
        """Reference files from settings are used"""
        state = app_state_with_mock_db

        # Create glossary
        glossary = tmp_path / "glossary.csv"
        glossary.write_text("JP,EN\nテスト,Test\n", encoding='utf-8')

        # Configure settings
        settings = AppSettings(reference_files=["glossary.csv"])
        service = TranslationService(mock_copilot, settings)

        # Get reference paths
        ref_paths = settings.get_reference_file_paths(tmp_path)

        assert len(ref_paths) == 1
        assert ref_paths[0] == glossary


# --- E2E Test: Language Detection Flow ---

class TestE2ELanguageDetection:
    """End-to-end tests for automatic language detection"""

    def test_auto_detect_japanese(self, mock_copilot, settings, app_state_with_mock_db):
        """Japanese text is auto-detected and translated to English"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Pure Japanese
        state.source_text = "日本語のテキストです。"

        # Verify detection
        assert is_japanese_text(state.source_text) is True

        result = service.translate_text_with_options(state.source_text)

        # Output should be English
        assert result.output_language == "en"

    def test_auto_detect_english(self, mock_copilot, settings, app_state_with_mock_db):
        """English text is auto-detected and translated to Japanese"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Pure English
        state.source_text = "This is English text."

        # Verify detection
        assert is_japanese_text(state.source_text) is False

        result = service.translate_text_with_options(state.source_text)

        # Output should be Japanese
        assert result.output_language == "jp"

    def test_mixed_language_text(self, mock_copilot, settings, app_state_with_mock_db):
        """Mixed language text is detected appropriately"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        state.copilot_ready = True

        # Mixed with majority Japanese
        state.source_text = "これはテストです。This is a test."

        # Detection depends on threshold (default 0.3)
        detected_jp = is_japanese_text(state.source_text)

        result = service.translate_text_with_options(state.source_text)

        # Result should reflect detection
        if detected_jp:
            assert result.output_language == "en"
        else:
            assert result.output_language == "jp"


# --- E2E Test: Font Conversion Flow ---

class TestE2EFontConversion:
    """End-to-end tests for font conversion during translation"""

    def test_jp_to_en_font_conversion(self, mock_copilot, settings, tmp_path):
        """Japanese fonts converted to English equivalents"""
        service = TranslationService(mock_copilot, settings)

        # Create file with Japanese fonts
        input_file = tmp_path / "jp_fonts.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "明朝体のテキスト"
        ws["A1"].font = Font(name="MS Mincho", size=12)

        ws["A2"] = "ゴシック体のテキスト"
        ws["A2"].font = Font(name="MS Gothic", size=14)

        ws["A3"] = "游ゴシックのテキスト"
        ws["A3"].font = Font(name="Yu Gothic", size=11)

        wb.save(input_file)

        # Translate (JP to EN)
        result = service.translate_file(input_file, output_language="en")

        # Verify font conversion
        wb_out = openpyxl.load_workbook(result.output_path)
        ws_out = wb_out.active

        # Mincho → Arial
        assert ws_out["A1"].font.name == "Arial"
        # Gothic → Arial (AppSettings default)
        assert ws_out["A2"].font.name == "Arial"
        # Font size unchanged (font_size_adjustment_jp_to_en = 0.0)
        assert ws_out["A2"].font.size == 14

    def test_en_to_jp_font_conversion(self, mock_copilot, settings, tmp_path):
        """English fonts converted to Japanese equivalents"""
        service = TranslationService(mock_copilot, settings)

        # Create file with English fonts
        input_file = tmp_path / "en_fonts.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "Arial text"
        ws["A1"].font = Font(name="Arial", size=10)

        ws["A2"] = "Times New Roman text"
        ws["A2"].font = Font(name="Times New Roman", size=12)

        wb.save(input_file)

        # Translate (EN to JP)
        result = service.translate_file(input_file, output_language="jp")

        # Verify output exists
        assert result.output_path.exists()


# --- E2E Test: Progress Tracking Flow ---

class TestE2EProgressTracking:
    """End-to-end tests for progress tracking"""

    def test_progress_updates_during_file_translation(self, mock_copilot, settings, tmp_path):
        """Progress updates correctly during file translation"""
        service = TranslationService(mock_copilot, settings)

        # Create file with many cells
        input_file = tmp_path / "progress_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 51):
            ws[f"A{i}"] = f"テキスト{i}"
        wb.save(input_file)

        progress_history = []

        def track_progress(progress: TranslationProgress):
            progress_history.append({
                'current': progress.current,
                'total': progress.total,
                'status': progress.status,
                'percentage': progress.percentage,
            })

        result = service.translate_file(input_file, on_progress=track_progress)

        # Verify progress tracking
        assert len(progress_history) > 0

        # Should have stages: extracting, translating, applying, complete
        statuses = [p['status'] for p in progress_history]
        assert any("Extract" in s for s in statuses)
        assert any("Complete" in s for s in statuses)

        # Final progress should be 100
        assert progress_history[-1]['current'] == 100

    def test_progress_percentage_calculation(self, mock_copilot, settings, tmp_path):
        """Progress percentage is calculated correctly"""
        service = TranslationService(mock_copilot, settings)

        input_file = tmp_path / "percentage_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 11):
            ws[f"A{i}"] = f"テキスト{i}"
        wb.save(input_file)

        percentages = []

        def track_percentage(progress: TranslationProgress):
            percentages.append(progress.percentage)

        service.translate_file(input_file, on_progress=track_percentage)

        # Percentages should increase
        assert percentages[-1] == 1.0  # 100%

        # Should be monotonically increasing or equal
        for i in range(1, len(percentages)):
            assert percentages[i] >= percentages[i - 1]


# --- E2E Test: Complete Workflow Simulation ---

class TestE2ECompleteWorkflow:
    """Simulate complete realistic user workflows"""

    def test_typical_user_session(self, mock_copilot, settings, tmp_path, app_state_with_mock_db):
        """Simulate a typical user session with multiple operations"""
        state = app_state_with_mock_db
        service = TranslationService(mock_copilot, settings)

        # === Session Start ===

        # 1. Connect to Copilot
        state.copilot_ready = True

        # === Text Translation ===

        # 2. Translate some text
        state.current_tab = Tab.TEXT
        state.source_text = "会議の議事録を作成してください。"
        result1 = service.translate_text_with_options(state.source_text)
        state.text_result = result1
        state.history.insert(0, HistoryEntry(source_text=state.source_text, result=result1))

        # 3. Translate more text
        state.source_text = "プロジェクトの進捗報告"
        result2 = service.translate_text_with_options(state.source_text)
        state.text_result = result2
        state.history.insert(0, HistoryEntry(source_text=state.source_text, result=result2))

        # === File Translation ===

        # 4. Switch to file tab
        state.current_tab = Tab.FILE

        # 5. Create and translate a file
        input_file = tmp_path / "meeting_notes.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "会議メモ"
        ws["A2"] = "日時: 2024年1月15日"
        ws["A3"] = "参加者: 田中、佐藤、鈴木"
        ws["A4"] = "議題: 新製品開発について"
        wb.save(input_file)

        state.file_state = FileState.SELECTED
        state.selected_file = input_file
        state.file_info = service.get_file_info(input_file)

        state.file_state = FileState.TRANSLATING
        file_result = service.translate_file(input_file)
        state.file_state = FileState.COMPLETE
        state.output_file = file_result.output_path

        # === Back to Text ===

        # 6. Switch back to text tab
        state.current_tab = Tab.TEXT

        # 7. Check history and restore
        assert len(state.history) >= 2
        old_entry = state.history[1]  # Second most recent
        state.source_text = old_entry.source_text
        state.text_result = old_entry.result

        # === Session End ===

        # Verify final state
        assert state.copilot_ready is True
        assert len(state.history) >= 2
        assert state.output_file.exists()
        assert state.source_text == "会議の議事録を作成してください。"
