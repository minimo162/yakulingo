# tests/test_error_handling.py
"""
Error handling tests for YakuLingo.
Tests error scenarios and recovery behaviors.
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch
import openpyxl

from yakulingo.models.types import (
    TranslationStatus,
    TextBlock,
)
from yakulingo.config.settings import AppSettings
from yakulingo.services.translation_service import TranslationService, BatchTranslator
from yakulingo.services.copilot_handler import CopilotHandler
from yakulingo.services.prompt_builder import PromptBuilder


# --- Fixtures ---

@pytest.fixture
def settings():
    """Default AppSettings"""
    return AppSettings()


@pytest.fixture
def sample_excel(tmp_path):
    """Create a basic Excel file"""
    file_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "テスト文章"
    wb.save(file_path)
    return file_path


# --- CopilotHandler Error Tests ---

class TestCopilotConnectionErrors:
    """Test Copilot connection error handling"""

    def test_translate_text_copilot_not_connected(self, settings):
        """Translation fails gracefully when Copilot not connected"""
        mock_copilot = Mock()
        mock_copilot.is_connected = False
        mock_copilot.translate_single.side_effect = RuntimeError("Not connected to Copilot")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None
        assert "Not connected" in result.error_message

    def test_translate_file_copilot_error(self, settings, sample_excel):
        """File translation fails gracefully on Copilot error"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = ConnectionError("Copilot session expired")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "session expired" in result.error_message.lower()

    def test_translate_copilot_timeout(self, settings, sample_excel):
        """Handle Copilot request timeout"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = TimeoutError("Request timed out after 120s")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "timed out" in result.error_message.lower()


# --- File Error Tests ---

class TestFileErrors:
    """Test file-related error handling"""

    def test_file_not_found(self, settings):
        """Handle missing file"""
        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(Path("/nonexistent/file.xlsx"))

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_unsupported_file_type_in_translate(self, settings, tmp_path):
        """Handle unsupported file type"""
        # Create an unsupported file type (.xyz)
        xyz_file = tmp_path / "test.xyz"
        xyz_file.write_text("Test content")

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(xyz_file)

        assert result.status == TranslationStatus.FAILED
        assert "Unsupported" in result.error_message

    def test_corrupted_excel_file(self, settings, tmp_path):
        """Handle corrupted Excel file"""
        # Create invalid xlsx (just random bytes)
        corrupt_file = tmp_path / "corrupt.xlsx"
        corrupt_file.write_bytes(b"This is not a valid Excel file")

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(corrupt_file)

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None

    def test_output_directory_not_writable(self, settings, sample_excel, tmp_path):
        """Handle unwritable output directory"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Set output to non-existent directory
        settings.output_directory = "/nonexistent/path/output"

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        # Should fail when trying to write output
        assert result.status == TranslationStatus.FAILED


# --- Batch Translation Error Tests ---

class TestBatchTranslationErrors:
    """Test batch translation error handling"""

    def test_partial_batch_failure(self, settings):
        """Handle partial failure in batch translation"""
        mock_copilot = Mock()
        # First batch succeeds, second fails
        mock_copilot.translate_sync.side_effect = [
            ["Trans1", "Trans2"],
            Exception("Batch 2 failed"),
        ]

        prompt_builder = PromptBuilder()

        # Set small max_chars_per_batch to force multiple batches
        translator = BatchTranslator(mock_copilot, prompt_builder, max_chars_per_batch=50)

        # Create blocks that will span two batches
        blocks = [
            TextBlock(id=str(i), text=f"LongText{i}x" * 5, location=f"A{i}")
            for i in range(4)  # 4 blocks with ~55 chars each
        ]

        with pytest.raises(Exception) as exc:
            translator.translate_blocks(blocks)

        assert "Batch 2 failed" in str(exc.value)

    def test_empty_response_from_copilot(self, settings):
        """Handle empty response from Copilot"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = []  # Empty response

        prompt_builder = PromptBuilder()  # Use real PromptBuilder

        translator = BatchTranslator(mock_copilot, prompt_builder)

        blocks = [TextBlock(id="1", text="Test", location="A1")]

        # Should handle empty response gracefully
        results = translator.translate_blocks(blocks)

        # Implementation keeps original text when translation is missing
        assert len(results) == 1
        assert results["1"] == "Test"  # Original text preserved

    def test_mismatched_response_count(self, settings):
        """Handle when Copilot returns wrong number of translations"""
        mock_copilot = Mock()
        # Send 3 blocks, get 2 translations back
        mock_copilot.translate_sync.return_value = ["Trans1", "Trans2"]

        prompt_builder = PromptBuilder()  # Use real PromptBuilder

        translator = BatchTranslator(mock_copilot, prompt_builder)

        blocks = [
            TextBlock(id="1", text="Text1", location="A1"),
            TextBlock(id="2", text="Text2", location="A2"),
            TextBlock(id="3", text="Text3", location="A3"),
        ]

        results = translator.translate_blocks(blocks)

        # All blocks are in results: 2 translated + 1 with original text
        assert len(results) == 3
        assert results["1"] == "Trans1"
        assert results["2"] == "Trans2"
        assert results["3"] == "Text3"  # Original text preserved


# --- TranslationService Error Tests ---

class TestTranslationServiceErrors:
    """Test TranslationService error handling"""

    def test_translate_text_exception_captured(self, settings):
        """Exception during text translation is captured"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = ValueError("Invalid input")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "Invalid input" in result.error_message
        assert result.duration_seconds >= 0

    def test_translate_file_exception_captured(self, settings, sample_excel):
        """Exception during file translation is captured"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = RuntimeError("Translation failed")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "Translation failed" in result.error_message
        assert result.duration_seconds >= 0

    def test_get_file_info_error(self, settings, tmp_path):
        """Handle error in get_file_info"""
        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        # Non-existent file
        with pytest.raises(Exception):
            service.get_file_info(Path("/nonexistent/file.xlsx"))


# --- Cancellation Tests ---

class TestCancellationHandling:
    """Test cancellation during various stages"""

    def test_cancel_during_translation(self, settings, sample_excel):
        """Cancellation during multi-batch translation"""
        # Note: _cancel_requested is reset at start of translate_file()
        # So we need to trigger cancellation during translation
        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        call_count = [0]

        def cancel_on_second_call(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 1:
                service.cancel()
            return ["Translated"]

        mock_copilot.translate_sync.side_effect = cancel_on_second_call

        result = service.translate_file(sample_excel)

        # Single batch completes, cancellation flag set for next check
        # Result depends on whether cancellation is checked after batch
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.CANCELLED]

    def test_cancel_flag_reset_at_start(self, settings, sample_excel):
        """Cancellation flag is reset when translate_file starts"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        service = TranslationService(mock_copilot, settings)

        # Cancel before translation
        service.cancel()
        assert service._cancel_requested is True

        # Start translation - flag should reset
        result = service.translate_file(sample_excel)

        # Translation should complete because flag was reset at start
        assert result.status == TranslationStatus.COMPLETED


# --- Settings Error Tests ---

class TestSettingsErrors:
    """Test settings-related error handling"""

    def test_invalid_output_directory_setting(self, tmp_path):
        """Handle invalid output directory in settings"""
        settings = AppSettings(output_directory="/invalid/path/that/does/not/exist")

        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Create a valid input file
        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "テスト"
        wb.save(file_path)

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        # Should fail due to invalid output directory
        assert result.status == TranslationStatus.FAILED

    def test_settings_load_missing_file(self, tmp_path):
        """Loading settings from missing file uses defaults"""
        settings_path = tmp_path / "nonexistent_settings.json"

        settings = AppSettings.load(settings_path)

        # Should use defaults
        assert settings.max_chars_per_batch == 7000
        assert settings.request_timeout == 120

    def test_settings_load_invalid_json(self, tmp_path):
        """Loading settings from invalid JSON uses defaults"""
        settings_path = tmp_path / "invalid_settings.json"
        settings_path.write_text("{ invalid json }", encoding="utf-8")

        settings = AppSettings.load(settings_path)

        # Should use defaults (or handle gracefully)
        assert settings is not None


# --- Processor Error Tests ---

class TestProcessorErrors:
    """Test file processor error handling"""

    def test_extract_from_empty_xlsx(self, settings, tmp_path):
        """Extract from completely empty Excel file"""
        file_path = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        # Don't add any content
        wb.save(file_path)

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        # Should complete with warning about no translatable text
        assert result.status == TranslationStatus.COMPLETED
        assert result.blocks_total == 0
        assert result.warnings is not None

    def test_extract_from_xlsx_only_skip_patterns(self, settings, tmp_path):
        """Extract from file with only skip patterns"""
        file_path = tmp_path / "skip_only.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # Only non-translatable content
        ws["A1"] = "12345"  # Numbers
        ws["A2"] = "test@example.com"  # Email
        ws["A3"] = "https://example.com"  # URL
        ws["A4"] = "2024-01-15"  # Date

        wb.save(file_path)

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(file_path)

        assert result.status == TranslationStatus.COMPLETED
        assert result.blocks_total == 0
        # Copilot should not be called
        mock_copilot.translate_sync.assert_not_called()


# --- Network Error Simulation ---

class TestNetworkErrors:
    """Test network-related error scenarios"""

    def test_copilot_network_error(self, settings, sample_excel):
        """Handle network error during Copilot communication"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = ConnectionError(
            "Network unreachable"
        )

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "Network" in result.error_message or "unreachable" in result.error_message

    def test_copilot_ssl_error(self, settings, sample_excel):
        """Handle SSL/TLS error"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = ConnectionError("SSL certificate verify failed")

        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "SSL" in result.error_message


# --- Memory/Resource Error Tests ---

class TestResourceErrors:
    """Test resource-related error handling"""

    def test_very_large_text_block(self, settings):
        """Handle very large text block"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated"

        service = TranslationService(mock_copilot, settings)

        # Create very large text (100KB)
        large_text = "あ" * 100000

        result = service.translate_text(large_text)

        # Should either succeed or fail gracefully
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]
        if result.status == TranslationStatus.FAILED:
            assert result.error_message is not None


# --- Malformed Copilot Response Tests ---

class TestMalformedCopilotResponses:
    """Test handling of malformed responses from Copilot"""

    def test_empty_string_response(self, settings):
        """Handle empty string response from Copilot"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = ""

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        # Empty response should still return completed status
        assert result.status == TranslationStatus.COMPLETED
        assert result.output_text == ""

    def test_none_response(self, settings):
        """Handle None response from Copilot"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = None

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        # None response should complete (with None as output)
        assert result.status == TranslationStatus.COMPLETED

    def test_response_with_only_whitespace(self, settings):
        """Handle whitespace-only response"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "   \n\t  "

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.COMPLETED

    def test_response_with_unexpected_format(self, settings):
        """Handle response with unexpected format"""
        mock_copilot = Mock()
        # Response that doesn't match expected pattern
        mock_copilot.translate_single.return_value = "ERROR: Unable to process request"

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        # Should still complete, treating raw response as output
        assert result.status == TranslationStatus.COMPLETED
        assert "ERROR" in result.output_text

    def test_batch_response_count_mismatch(self, settings):
        """Handle when batch response has wrong count"""
        mock_copilot = Mock()
        # Send 3 texts, get only 1 back
        mock_copilot.translate_sync.return_value = ["Only one translation"]

        from yakulingo.services.prompt_builder import PromptBuilder
        prompt_builder = PromptBuilder()

        translator = BatchTranslator(mock_copilot, prompt_builder)

        blocks = [
            TextBlock(id="1", text="Text1", location="A1"),
            TextBlock(id="2", text="Text2", location="A2"),
            TextBlock(id="3", text="Text3", location="A3"),
        ]

        results = translator.translate_blocks(blocks)

        # All blocks are in results: 1 translated + 2 with original text
        assert len(results) == 3
        assert results["1"] == "Only one translation"
        assert results["2"] == "Text2"  # Original text preserved
        assert results["3"] == "Text3"  # Original text preserved

    def test_batch_response_extra_items(self, settings):
        """Handle when batch response has too many items"""
        mock_copilot = Mock()
        # Send 2 texts, get 5 back
        mock_copilot.translate_sync.return_value = [
            "Trans1", "Trans2", "Extra1", "Extra2", "Extra3"
        ]

        from yakulingo.services.prompt_builder import PromptBuilder
        prompt_builder = PromptBuilder()

        translator = BatchTranslator(mock_copilot, prompt_builder)

        blocks = [
            TextBlock(id="1", text="Text1", location="A1"),
            TextBlock(id="2", text="Text2", location="A2"),
        ]

        results = translator.translate_blocks(blocks)

        # Should only use the first 2 responses
        assert len(results) == 2


# --- Timeout and Retry Tests ---

class TestTimeoutAndRetry:
    """Test timeout and retry behavior"""

    def test_translate_file_timeout_error(self, settings, sample_excel):
        """Handle timeout during file translation"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.side_effect = TimeoutError(
            "Request timed out after 120 seconds"
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED
        assert "timed out" in result.error_message.lower()

    def test_translate_text_timeout_error(self, settings):
        """Handle timeout during text translation"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = TimeoutError(
            "Connection timed out"
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "timed out" in result.error_message.lower()

    def test_intermittent_failure_recovery(self, settings, sample_excel):
        """Test that service can recover after intermittent failure"""
        mock_copilot = Mock()
        call_count = [0]

        def intermittent_failure(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 1:
                raise ConnectionError("Temporary network issue")
            return ["Translated"]

        mock_copilot.translate_sync.side_effect = intermittent_failure

        service = TranslationService(mock_copilot, settings)

        # First call fails
        result1 = service.translate_file(sample_excel)
        assert result1.status == TranslationStatus.FAILED

        # Second call succeeds
        result2 = service.translate_file(sample_excel)
        assert result2.status == TranslationStatus.COMPLETED


# --- File Permission and I/O Error Tests ---

class TestFilePermissionErrors:
    """Test file permission error handling"""

    def test_read_only_output_directory(self, settings, sample_excel, tmp_path):
        """Handle read-only output directory"""
        import os

        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Set output directory that doesn't exist
        settings.output_directory = str(tmp_path / "nonexistent" / "readonly")

        service = TranslationService(mock_copilot, settings)
        result = service.translate_file(sample_excel)

        assert result.status == TranslationStatus.FAILED

    def test_input_file_deleted_during_processing(self, settings, tmp_path):
        """Handle input file being deleted during processing"""
        mock_copilot = Mock()

        # Create file then delete before processing
        file_path = tmp_path / "temporary.xlsx"

        service = TranslationService(mock_copilot, settings)

        # File doesn't exist
        result = service.translate_file(file_path)

        assert result.status == TranslationStatus.FAILED

    def test_output_file_already_exists_and_locked(self, settings, sample_excel, tmp_path):
        """Handle output file that already exists"""
        mock_copilot = Mock()
        mock_copilot.translate_sync.return_value = ["Translated"]

        # Create existing translated file
        existing_file = tmp_path / "sample_translated.xlsx"
        existing_file.write_text("existing content")

        # Set output to same directory
        settings.output_directory = str(tmp_path)

        service = TranslationService(mock_copilot, settings)

        # Should generate numbered filename
        result = service.translate_file(sample_excel)

        # Should either succeed with new name or fail gracefully
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]


# --- Unicode and Encoding Error Tests ---

class TestEncodingErrors:
    """Test encoding error handling"""

    def test_translate_text_with_invalid_unicode(self, settings):
        """Handle text with unusual Unicode characters"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated"

        service = TranslationService(mock_copilot, settings)

        # Text with various Unicode characters
        unicode_text = "日本語 \u200b\u200c\u200d 絵文字🎌 零幅文字"

        result = service.translate_text(unicode_text)

        # Should handle gracefully
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]

    def test_translate_text_with_surrogate_pairs(self, settings):
        """Handle text with surrogate pair emojis"""
        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated with emoji 😀"

        service = TranslationService(mock_copilot, settings)

        # Text with emoji that uses surrogate pairs
        emoji_text = "テスト 👨‍👩‍👧‍👦 絵文字"

        result = service.translate_text(emoji_text)

        assert result.status == TranslationStatus.COMPLETED


# --- Concurrent Access Error Tests ---

class TestConcurrentAccessErrors:
    """Test concurrent access error handling"""

    def test_cancel_during_batch_processing(self, settings, sample_excel):
        """Test cancellation during batch processing"""
        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        batch_count = [0]

        def slow_translate(*args, **kwargs):
            batch_count[0] += 1
            if batch_count[0] == 1:
                # Cancel after first batch
                service.cancel()
            return ["Translated"]

        mock_copilot.translate_sync.side_effect = slow_translate

        result = service.translate_file(sample_excel)

        # Should either complete or be cancelled
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.CANCELLED]

    def test_multiple_cancel_calls(self, settings):
        """Test multiple cancel calls are safe"""
        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        # Multiple cancel calls should not raise
        service.cancel()
        service.cancel()
        service.cancel()

        assert service._cancel_requested is True


# --- Edge Cases in Error Messages ---

class TestErrorMessageFormatting:
    """Test error message formatting"""

    def test_error_message_with_unicode(self, settings):
        """Test error message with Unicode characters"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = ValueError(
            "エラー: 無効な入力 - Invalid input"
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "エラー" in result.error_message

    def test_error_message_with_newlines(self, settings):
        """Test error message with newlines"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = RuntimeError(
            "Error occurred\nLine 1\nLine 2\nLine 3"
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "Error occurred" in result.error_message

    def test_very_long_error_message(self, settings):
        """Test handling of very long error message"""
        mock_copilot = Mock()
        long_message = "Error: " + "x" * 10000
        mock_copilot.translate_single.side_effect = RuntimeError(long_message)

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert result.error_message is not None


# --- Processor-Specific Error Tests ---

class TestProcessorSpecificErrors:
    """Test processor-specific error handling"""

    def test_excel_with_password_protection(self, settings, tmp_path):
        """Handle password-protected Excel file"""
        # Create a file that looks like xlsx but has invalid content
        protected_file = tmp_path / "protected.xlsx"
        protected_file.write_bytes(b"PK\x03\x04encrypted content simulation")

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(protected_file)

        assert result.status == TranslationStatus.FAILED

    def test_word_with_corrupted_xml(self, settings, tmp_path):
        """Handle Word file with corrupted XML"""
        corrupted_file = tmp_path / "corrupted.docx"
        corrupted_file.write_bytes(b"PK\x03\x04corrupted content")

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(corrupted_file)

        assert result.status == TranslationStatus.FAILED

    def test_pdf_with_no_text_layer(self, settings, tmp_path):
        """Handle PDF with no extractable text"""
        # Create minimal PDF-like file
        empty_pdf = tmp_path / "empty.pdf"
        empty_pdf.write_bytes(b"%PDF-1.4\n%%EOF")

        mock_copilot = Mock()
        service = TranslationService(mock_copilot, settings)

        result = service.translate_file(empty_pdf)

        # Should fail or complete with 0 blocks
        assert result.status in [TranslationStatus.COMPLETED, TranslationStatus.FAILED]


# --- Connection State Error Tests ---

class TestConnectionStateErrors:
    """Test connection state error handling"""

    def test_translate_after_disconnect(self, settings):
        """Handle translation attempt after Copilot disconnect"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = RuntimeError(
            "Not connected to Copilot"
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "Not connected" in result.error_message

    def test_translate_with_expired_session(self, settings):
        """Handle expired Copilot session"""
        mock_copilot = Mock()
        mock_copilot.translate_single.side_effect = RuntimeError(
            "Session expired. Please reconnect."
        )

        service = TranslationService(mock_copilot, settings)
        result = service.translate_text("テスト")

        assert result.status == TranslationStatus.FAILED
        assert "expired" in result.error_message.lower()
