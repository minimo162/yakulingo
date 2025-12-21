# tests/test_prompt_builder.py
"""Tests for yakulingo.services.prompt_builder"""

import tempfile
from pathlib import Path

from yakulingo.services.prompt_builder import (
    PromptBuilder,
    REFERENCE_INSTRUCTION,
    DEFAULT_TO_EN_TEMPLATE,
    DEFAULT_TO_JP_TEMPLATE,
    DEFAULT_TEXT_TO_JP_TEMPLATE,
    DEFAULT_TEXT_TO_EN_CLIPBOARD_TEMPLATE,
    DEFAULT_TEXT_TO_JP_CLIPBOARD_TEMPLATE,
)


class TestPromptBuilder:
    """Tests for PromptBuilder class"""

    def test_default_templates_used_when_no_dir(self):
        """Default templates used when prompts_dir is None"""
        builder = PromptBuilder(prompts_dir=None)

        # Check internal templates are loaded (style-specific templates)
        # Each style should have templates for en and jp
        assert ("en", "concise") in builder._templates
        assert ("jp", "concise") in builder._templates
        assert builder._templates[("en", "concise")] == DEFAULT_TO_EN_TEMPLATE
        assert builder._templates[("jp", "concise")] == DEFAULT_TO_JP_TEMPLATE

    def test_default_text_templates_used_when_no_dir(self):
        """Default text templates are loaded when prompts_dir is None"""
        builder = PromptBuilder(prompts_dir=None)

        assert builder._text_templates[("jp", "concise")] == DEFAULT_TEXT_TO_JP_TEMPLATE
        assert builder.get_text_template("en", "concise") is None
        assert builder.get_text_clipboard_template("en") == DEFAULT_TEXT_TO_EN_CLIPBOARD_TEMPLATE
        assert builder.get_text_clipboard_template("jp") == DEFAULT_TEXT_TO_JP_CLIPBOARD_TEMPLATE

    def test_build_includes_input_text(self):
        """Build includes input text in prompt"""
        builder = PromptBuilder()
        prompt = builder.build("こんにちは")

        assert "こんにちは" in prompt
        # Input text appears after "---" separator
        assert "---" in prompt

    def test_build_for_english_output(self):
        """Build prompt for English output (JP→EN)"""
        builder = PromptBuilder()
        prompt = builder.build("テスト", output_language="en")

        assert "テスト" in prompt
        assert "英語" in prompt or "English" in prompt.lower()

    def test_build_for_japanese_output(self):
        """Build prompt for Japanese output (EN→JP)"""
        builder = PromptBuilder()
        prompt = builder.build("Test", output_language="jp")

        assert "Test" in prompt
        assert "日本語" in prompt

    def test_build_without_reference(self):
        """Build without reference files"""
        builder = PromptBuilder()
        prompt = builder.build("テスト", has_reference_files=False)

        assert "添付の参考ファイル" not in prompt
        assert REFERENCE_INSTRUCTION.strip() not in prompt

    def test_build_with_reference(self):
        """Build with reference files"""
        builder = PromptBuilder()
        prompt = builder.build("テスト", has_reference_files=True)

        assert "添付の参考ファイル" in prompt
        assert "用語集" in prompt

    def test_build_batch_numbered_format(self):
        """Build batch creates numbered format with end markers"""
        builder = PromptBuilder()
        texts = ["こんにちは", "さようなら", "ありがとう"]
        prompt = builder.build_batch(texts)

        # Each item should be numbered
        assert "1. こんにちは" in prompt
        assert "2. さようなら" in prompt
        assert "3. ありがとう" in prompt

    def test_build_batch_with_reference(self):
        """Build batch with reference files"""
        builder = PromptBuilder()
        texts = ["Text1", "Text2"]
        prompt = builder.build_batch(texts, has_reference_files=True)

        assert "1. Text1" in prompt
        assert "2. Text2" in prompt
        assert "添付の参考ファイル" in prompt

    def test_build_batch_for_japanese_output(self):
        """Build batch for Japanese output"""
        builder = PromptBuilder()
        texts = ["Hello", "World"]
        prompt = builder.build_batch(texts, output_language="jp")

        assert "1. Hello" in prompt
        assert "2. World" in prompt
        assert "日本語" in prompt

    def test_missing_text_prompts_fall_back_to_defaults(self, tmp_path):
        """Text/clipboard templates fall back to defaults when files are missing"""
        prompts_dir = tmp_path / "prompts"
        prompts_dir.mkdir()

        custom_jp = prompts_dir / "text_translate_to_jp.txt"
        custom_jp.write_text("custom jp", encoding="utf-8")

        custom_clipboard_en = prompts_dir / "text_translate_to_en_clipboard.txt"
        custom_clipboard_en.write_text("custom clipboard en", encoding="utf-8")

        builder = PromptBuilder(prompts_dir=prompts_dir)

        assert builder._text_templates[("jp", "standard")] == "custom jp"
        assert builder._text_templates[("jp", "minimal")] == "custom jp"
        assert builder.get_text_template("en", "concise") is None
        assert builder.get_text_clipboard_template("en") == "custom clipboard en"
        assert builder.get_text_clipboard_template("jp") == DEFAULT_TEXT_TO_JP_CLIPBOARD_TEMPLATE


class TestPromptBuilderParseBatchResult:
    """Tests for PromptBuilder.parse_batch_result()"""

    def test_parse_numbered_results(self):
        """Parse results with numbered format"""
        builder = PromptBuilder()
        result = """1. Hello
2. Goodbye
3. Thank you"""
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert parsed[0] == "Hello"
        assert parsed[1] == "Goodbye"
        assert parsed[2] == "Thank you"

    def test_parse_without_numbers(self):
        """Parse results without numbered format"""
        builder = PromptBuilder()
        result = """Hello
Goodbye
Thank you"""
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert parsed[0] == "Hello"
        assert parsed[1] == "Goodbye"
        assert parsed[2] == "Thank you"

    def test_parse_fewer_results_pads_empty(self):
        """Parse fewer results than expected pads with empty strings"""
        builder = PromptBuilder()
        result = "1. Hello"
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert parsed[0] == "Hello"
        assert parsed[1] == ""
        assert parsed[2] == ""

    def test_parse_more_results_truncates(self):
        """Parse more results than expected truncates"""
        builder = PromptBuilder()
        result = """1. One
2. Two
3. Three
4. Four"""
        parsed = builder.parse_batch_result(result, expected_count=2)

        assert len(parsed) == 2
        assert parsed[0] == "One"
        assert parsed[1] == "Two"

    def test_parse_skips_empty_lines(self):
        """Parse skips empty lines"""
        builder = PromptBuilder()
        result = """1. Hello

2. World

"""
        parsed = builder.parse_batch_result(result, expected_count=2)

        assert len(parsed) == 2
        assert parsed[0] == "Hello"
        assert parsed[1] == "World"

    def test_parse_handles_whitespace(self):
        """Parse handles leading/trailing whitespace"""
        builder = PromptBuilder()
        result = """  1. Hello
  2.   World   """
        parsed = builder.parse_batch_result(result, expected_count=2)

        assert parsed[0] == "Hello"
        assert parsed[1] == "World"

    def test_parse_empty_result(self):
        """Parse empty result returns empty strings"""
        builder = PromptBuilder()
        result = ""
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert all(p == "" for p in parsed)

    def test_parse_mixed_numbered_and_unnumbered(self):
        """Parse handles mixed numbered and unnumbered lines"""
        builder = PromptBuilder()
        result = """1. First
Second line
3. Third"""
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert parsed[0] == "First"
        assert parsed[1] == "Second line"
        assert parsed[2] == "Third"

    def test_parse_large_batch(self):
        """Parse large batch result"""
        builder = PromptBuilder()
        lines = [f"{i+1}. Translation {i+1}" for i in range(100)]
        result = "\n".join(lines)
        parsed = builder.parse_batch_result(result, expected_count=100)

        assert len(parsed) == 100
        assert parsed[0] == "Translation 1"
        assert parsed[99] == "Translation 100"

    def test_parse_with_special_characters(self):
        """Parse results with special characters"""
        builder = PromptBuilder()
        result = """1. Hello! @#$%
2. 日本語テスト
3. Mixed: 英語と日本語"""
        parsed = builder.parse_batch_result(result, expected_count=3)

        assert len(parsed) == 3
        assert parsed[0] == "Hello! @#$%"
        assert parsed[1] == "日本語テスト"
        assert parsed[2] == "Mixed: 英語と日本語"

    def test_parse_multiline_entries(self):
        """Parse may not handle multiline entries (single line per entry)"""
        builder = PromptBuilder()
        # This tests current behavior - each line is a separate result
        result = """1. First line
continuation of first
2. Second"""
        parsed = builder.parse_batch_result(result, expected_count=3)

        # Current implementation treats each non-empty line as result
        assert len(parsed) == 3
        assert parsed[0] == "First line"


class TestPromptBuilderTemplateLoading:
    """Tests for template loading from files"""

    def test_load_to_en_template_from_file(self):
        """Load JP→EN template from file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            prompts_dir = Path(tmpdir)

            # Create custom to_en template
            to_en_file = prompts_dir / "file_translate_to_en.txt"
            to_en_file.write_text("Custom EN template: {input_text}\n{reference_section}")

            builder = PromptBuilder(prompts_dir=prompts_dir)
            prompt = builder.build("test", output_language="en")

            assert "Custom EN template: test" in prompt

    def test_load_to_jp_template_from_file(self):
        """Load EN→JP template from file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            prompts_dir = Path(tmpdir)

            # Create custom to_jp template
            to_jp_file = prompts_dir / "file_translate_to_jp.txt"
            to_jp_file.write_text("Custom JP template: {input_text}\n{reference_section}")

            builder = PromptBuilder(prompts_dir=prompts_dir)
            prompt = builder.build("test", output_language="jp")

            assert "Custom JP template: test" in prompt

    def test_fallback_to_default_when_to_en_missing(self):
        """Fallback to default when file_translate_to_en.txt missing"""
        with tempfile.TemporaryDirectory() as tmpdir:
            prompts_dir = Path(tmpdir)
            # Don't create file_translate_to_en.txt

            builder = PromptBuilder(prompts_dir=prompts_dir)
            prompt = builder.build("test", output_language="en")

            # Should use default template
            assert "ファイル翻訳リクエスト" in prompt
            assert "英語" in prompt

    def test_fallback_to_default_when_to_jp_missing(self):
        """Fallback to default when file_translate_to_jp.txt missing"""
        with tempfile.TemporaryDirectory() as tmpdir:
            prompts_dir = Path(tmpdir)
            # Don't create file_translate_to_jp.txt

            builder = PromptBuilder(prompts_dir=prompts_dir)
            prompt = builder.build("test", output_language="jp")

            # Should use default template
            assert "ファイル翻訳リクエスト" in prompt
            assert "日本語" in prompt

    def test_mixed_custom_and_default_templates(self):
        """Load one custom template, use default for other"""
        with tempfile.TemporaryDirectory() as tmpdir:
            prompts_dir = Path(tmpdir)

            # Only create to_en template
            to_en_file = prompts_dir / "file_translate_to_en.txt"
            to_en_file.write_text("Custom EN: {input_text}\n{reference_section}\n{translation_rules}")

            builder = PromptBuilder(prompts_dir=prompts_dir)

            # EN output uses custom
            prompt_en = builder.build("test", output_language="en")
            assert "Custom EN: test" in prompt_en

            # JP output uses default
            prompt_jp = builder.build("test", output_language="jp")
            assert "ファイル翻訳リクエスト" in prompt_jp
            assert "日本語" in prompt_jp


class TestDefaultTemplates:
    """Tests for default prompt template content"""

    def test_to_en_template_has_placeholders(self):
        """TO_EN template has required placeholders"""
        assert "{input_text}" in DEFAULT_TO_EN_TEMPLATE
        assert "{reference_section}" in DEFAULT_TO_EN_TEMPLATE

    def test_to_jp_template_has_placeholders(self):
        """TO_JP template has required placeholders"""
        assert "{input_text}" in DEFAULT_TO_JP_TEMPLATE
        assert "{reference_section}" in DEFAULT_TO_JP_TEMPLATE

    def test_to_en_template_contains_rules(self):
        """TO_EN template contains translation rules placeholders and structure"""
        assert "出力形式（最優先ルール）" in DEFAULT_TO_EN_TEMPLATE
        assert "翻訳スタイル" in DEFAULT_TO_EN_TEMPLATE
        assert "{translation_rules}" in DEFAULT_TO_EN_TEMPLATE  # Placeholder for common rules

    def test_to_jp_template_contains_rules(self):
        """TO_JP template contains translation rules and structure"""
        assert "出力形式（最優先ルール）" in DEFAULT_TO_JP_TEMPLATE
        assert "翻訳ガイドライン" in DEFAULT_TO_JP_TEMPLATE
        assert "数値表記ルール" in DEFAULT_TO_JP_TEMPLATE
        assert "億" in DEFAULT_TO_JP_TEMPLATE  # Number notation rule

    def test_to_en_template_targets_english(self):
        """TO_EN template targets English output"""
        assert "英語" in DEFAULT_TO_EN_TEMPLATE

    def test_to_jp_template_targets_japanese(self):
        """TO_JP template targets Japanese output"""
        assert "日本語" in DEFAULT_TO_JP_TEMPLATE

    def test_reference_instruction_content(self):
        """REFERENCE_INSTRUCTION has correct content"""
        assert "添付の参考ファイル" in REFERENCE_INSTRUCTION
        assert "用語集" in REFERENCE_INSTRUCTION
        assert "訳語" in REFERENCE_INSTRUCTION


class TestPromptBuilderEdgeCases:
    """Edge case tests for PromptBuilder"""

    def test_empty_input_text(self):
        """Handle empty input text"""
        builder = PromptBuilder()
        prompt = builder.build("")

        assert "---" in prompt
        # Empty input should still produce valid prompt structure

    def test_very_long_input_text(self):
        """Handle very long input text"""
        builder = PromptBuilder()
        long_text = "あ" * 10000
        prompt = builder.build(long_text)

        assert long_text in prompt

    def test_special_characters_in_input(self):
        """Handle special characters in input"""
        builder = PromptBuilder()
        special_text = "Test {with} [brackets] and $pecial ch@rs!"
        prompt = builder.build(special_text)

        assert special_text in prompt

    def test_build_batch_empty_list(self):
        """Handle empty text list"""
        builder = PromptBuilder()
        prompt = builder.build_batch([])

        # Should produce valid prompt with empty input section
        assert "---" in prompt

    def test_build_batch_single_item(self):
        """Handle single item in batch"""
        builder = PromptBuilder()
        prompt = builder.build_batch(["Single"])

        assert "1. Single" in prompt

    def test_get_template_defaults_to_en(self):
        """_get_template defaults to English output"""
        builder = PromptBuilder()

        # Default behavior (defaults to en, concise)
        template = builder._get_template()
        assert template == builder._templates[("en", "concise")]

        # Explicit en
        template_en = builder._get_template("en")
        assert template_en == builder._templates[("en", "concise")]

    def test_get_template_for_jp(self):
        """_get_template returns JP template for jp language"""
        builder = PromptBuilder()
        template = builder._get_template("jp")
        assert template == builder._templates[("jp", "concise")]

    def test_get_template_unknown_language_falls_back(self):
        """_get_template returns fallback for unknown language"""
        builder = PromptBuilder()
        # Unknown language falls back to JP template (since output_language != "en")
        template = builder._get_template("unknown")
        # Should return JP template as fallback (unknown != "en")
        assert template == DEFAULT_TO_JP_TEMPLATE


# =============================================================================
# Tests: Glossary and Reference File Edge Cases
# =============================================================================

class TestGlossaryEdgeCases:
    """Edge case tests for glossary/reference file handling"""

    def test_prompt_with_reference_includes_instruction(self):
        """Prompt with reference files includes reference instruction"""
        builder = PromptBuilder()
        prompt = builder.build("テスト", has_reference_files=True)

        assert "Reference Files" in prompt
        assert "添付の参考ファイル" in prompt
        assert "用語集" in prompt

    def test_prompt_without_reference_excludes_instruction(self):
        """Prompt without reference files excludes reference instruction"""
        builder = PromptBuilder()
        prompt = builder.build("テスト", has_reference_files=False)

        assert "Reference Files" not in prompt
        assert "添付の参考ファイル" not in prompt

    def test_batch_prompt_with_reference(self):
        """Batch prompt includes reference instruction when enabled"""
        builder = PromptBuilder()
        texts = ["Text1", "Text2"]
        prompt = builder.build_batch(texts, has_reference_files=True)

        assert "Reference Files" in prompt
        assert "添付の参考ファイル" in prompt

    def test_batch_prompt_without_reference(self):
        """Batch prompt excludes reference instruction when disabled"""
        builder = PromptBuilder()
        texts = ["Text1", "Text2"]
        prompt = builder.build_batch(texts, has_reference_files=False)

        assert "Reference Files" not in prompt
        assert "添付の参考ファイル" not in prompt

    def test_reference_instruction_format(self):
        """Reference instruction has proper format"""
        # Verify the reference instruction constant
        assert "Reference Files" in REFERENCE_INSTRUCTION
        assert "用語集がある場合" in REFERENCE_INSTRUCTION
        assert "訳語を使用" in REFERENCE_INSTRUCTION

    def test_reference_section_placeholder_replaced(self):
        """Reference section placeholder is properly replaced"""
        builder = PromptBuilder()

        prompt_with_ref = builder.build("test", has_reference_files=True)
        prompt_without_ref = builder.build("test", has_reference_files=False)

        # No placeholder remains
        assert "{reference_section}" not in prompt_with_ref
        assert "{reference_section}" not in prompt_without_ref

    def test_input_text_placeholder_replaced(self):
        """Input text placeholder is properly replaced"""
        builder = PromptBuilder()
        prompt = builder.build("特定のテキスト")

        assert "{input_text}" not in prompt
        assert "特定のテキスト" in prompt


class TestReferenceFileIntegration:
    """Tests for reference file integration with translation service"""

    def test_translation_service_accepts_reference_files(self):
        """TranslationService accepts reference file paths"""
        from unittest.mock import Mock
        from yakulingo.config.settings import AppSettings
        from yakulingo.services.translation_service import TranslationService

        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated"

        service = TranslationService(mock_copilot, AppSettings())

        with tempfile.TemporaryDirectory() as tmpdir:
            ref_file = Path(tmpdir) / "glossary.csv"
            ref_file.write_text("term,translation\nテスト,Test")

            result = service.translate_text("テスト", reference_files=[ref_file])

            # Translation should succeed
            from yakulingo.models.types import TranslationStatus
            assert result.status == TranslationStatus.COMPLETED

    def test_translation_service_handles_none_reference(self):
        """TranslationService handles None reference files"""
        from unittest.mock import Mock
        from yakulingo.config.settings import AppSettings
        from yakulingo.services.translation_service import TranslationService

        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated"

        service = TranslationService(mock_copilot, AppSettings())

        result = service.translate_text("テスト", reference_files=None)

        from yakulingo.models.types import TranslationStatus
        assert result.status == TranslationStatus.COMPLETED

    def test_translation_service_handles_empty_reference_list(self):
        """TranslationService handles empty reference file list"""
        from unittest.mock import Mock
        from yakulingo.config.settings import AppSettings
        from yakulingo.services.translation_service import TranslationService

        mock_copilot = Mock()
        mock_copilot.translate_single.return_value = "Translated"

        service = TranslationService(mock_copilot, AppSettings())

        result = service.translate_text("テスト", reference_files=[])

        from yakulingo.models.types import TranslationStatus
        assert result.status == TranslationStatus.COMPLETED


class TestGlossaryFileFormats:
    """Tests for different glossary file format handling"""

    def test_csv_glossary_reference(self):
        """CSV glossary file as reference"""
        with tempfile.TemporaryDirectory() as tmpdir:
            glossary = Path(tmpdir) / "glossary.csv"
            glossary.write_text("Japanese,English\n会議,meeting\n報告書,report")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["glossary.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1
            assert paths[0] == glossary

    def test_excel_glossary_reference(self):
        """Excel glossary file as reference"""
        with tempfile.TemporaryDirectory() as tmpdir:
            import openpyxl
            glossary = Path(tmpdir) / "glossary.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Japanese'
            ws['B1'] = 'English'
            ws['A2'] = '会議'
            ws['B2'] = 'meeting'
            wb.save(glossary)

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["glossary.xlsx"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1
            assert paths[0] == glossary

    def test_mixed_reference_files(self):
        """Multiple reference files of different types"""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "terms.csv"
            csv_file.write_text("term,translation")

            txt_file = Path(tmpdir) / "notes.txt"
            txt_file.write_text("Translation notes")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["terms.csv", "notes.txt"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 2

    def test_missing_reference_file_excluded(self):
        """Missing reference files are excluded from list"""
        with tempfile.TemporaryDirectory() as tmpdir:
            existing = Path(tmpdir) / "exists.csv"
            existing.write_text("data")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["exists.csv", "missing.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            # Only existing file should be included
            assert len(paths) == 1
            assert paths[0] == existing


class TestGlossaryEdgeCasesAdvanced:
    """Advanced edge case tests for glossary handling"""

    def test_empty_glossary_file(self):
        """Handle empty glossary file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            empty_glossary = Path(tmpdir) / "empty.csv"
            empty_glossary.write_text("")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["empty.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            # Empty file should still be included
            assert len(paths) == 1

    def test_glossary_with_unicode_filename(self):
        """Handle glossary with Unicode filename"""
        with tempfile.TemporaryDirectory() as tmpdir:
            unicode_file = Path(tmpdir) / "用語集.csv"
            unicode_file.write_text("term,translation")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["用語集.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1
            assert paths[0] == unicode_file

    def test_glossary_with_special_characters(self):
        """Handle glossary with special content"""
        with tempfile.TemporaryDirectory() as tmpdir:
            special_glossary = Path(tmpdir) / "special.csv"
            special_glossary.write_text(
                "term,translation\n"
                "Ω,omega\n"
                "∞,infinity\n"
                "❤️,heart\n"
            )

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["special.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1

    def test_large_glossary_file(self):
        """Handle large glossary file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            large_glossary = Path(tmpdir) / "large.csv"

            # Create large glossary with many entries
            lines = ["term,translation"]
            for i in range(10000):
                lines.append(f"term{i},translation{i}")
            large_glossary.write_text("\n".join(lines))

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["large.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1
            assert paths[0].stat().st_size > 100000  # Should be a large file

    def test_glossary_path_normalization(self):
        """Handle different path formats for glossary"""
        with tempfile.TemporaryDirectory() as tmpdir:
            glossary = Path(tmpdir) / "glossary.csv"
            glossary.write_text("term,translation")

            from yakulingo.config.settings import AppSettings

            # Relative path
            settings1 = AppSettings(reference_files=["glossary.csv"])
            paths1 = settings1.get_reference_file_paths(Path(tmpdir))
            assert len(paths1) == 1

            # Absolute path
            settings2 = AppSettings(reference_files=[str(glossary)])
            paths2 = settings2.get_reference_file_paths(Path(tmpdir))
            assert len(paths2) == 1

    def test_glossary_with_subdirectory(self):
        """Handle glossary in subdirectory"""
        with tempfile.TemporaryDirectory() as tmpdir:
            subdir = Path(tmpdir) / "data" / "glossaries"
            subdir.mkdir(parents=True)

            glossary = subdir / "terms.csv"
            glossary.write_text("term,translation")

            from yakulingo.config.settings import AppSettings
            settings = AppSettings(reference_files=["data/glossaries/terms.csv"])

            paths = settings.get_reference_file_paths(Path(tmpdir))

            assert len(paths) == 1
            assert paths[0] == glossary
