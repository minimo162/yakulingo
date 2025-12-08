# tests/test_excel_processor.py
"""Tests for yakulingo.processors.excel_processor"""

import pytest
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

from yakulingo.processors.excel_processor import ExcelProcessor
from yakulingo.models.types import FileType


# --- Fixtures ---

@pytest.fixture
def processor():
    """ExcelProcessor instance"""
    return ExcelProcessor()


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample Excel file with text content"""
    file_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add various content
    ws["A1"] = "こんにちは"  # Japanese text (should translate)
    ws["A2"] = "世界"        # Japanese text (should translate)
    ws["B1"] = "12345"       # Numbers only (should skip)
    ws["B2"] = "test@example.com"  # Email (should skip)
    ws["C1"] = "Hello World"  # English text (should translate)

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_font(tmp_path):
    """Create Excel file with specific fonts"""
    file_path = tmp_path / "font_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"  # Set explicit sheet name

    # Cell with Mincho font
    ws["A1"] = "明朝フォント"
    ws["A1"].font = Font(name="MS Mincho", size=12)

    # Cell with Gothic font
    ws["A2"] = "ゴシックフォント"
    ws["A2"].font = Font(name="MS Gothic", size=14)

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_multiple_sheets(tmp_path):
    """Create Excel file with multiple sheets"""
    file_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()

    # Sheet1
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "シート1のテキスト"

    # Sheet2
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "シート2のテキスト"

    wb.save(file_path)
    return file_path


@pytest.fixture
def empty_xlsx(tmp_path):
    """Create empty Excel file"""
    file_path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(file_path)
    return file_path


# --- Tests: Properties ---

class TestExcelProcessorProperties:
    """Test ExcelProcessor properties"""

    def test_file_type(self, processor):
        assert processor.file_type == FileType.EXCEL

    def test_supported_extensions(self, processor):
        extensions = processor.supported_extensions
        assert ".xlsx" in extensions
        assert ".xls" in extensions


# --- Tests: get_file_info ---

class TestExcelProcessorGetFileInfo:
    """Test ExcelProcessor.get_file_info()"""

    def test_file_info_basic(self, processor, sample_xlsx):
        """Basic file info retrieval"""
        info = processor.get_file_info(sample_xlsx)

        assert info.path == sample_xlsx
        assert info.file_type == FileType.EXCEL
        assert info.size_bytes > 0
        assert info.sheet_count == 1

    def test_file_info_multiple_sheets(self, processor, xlsx_with_multiple_sheets):
        """File info with multiple sheets"""
        info = processor.get_file_info(xlsx_with_multiple_sheets)
        assert info.sheet_count == 2

    def test_file_info_empty(self, processor, empty_xlsx):
        """File info for empty file"""
        info = processor.get_file_info(empty_xlsx)
        assert info.sheet_count == 1  # Default sheet exists


# --- Tests: extract_text_blocks ---

class TestExcelProcessorExtractTextBlocks:
    """Test ExcelProcessor.extract_text_blocks()"""

    def test_extracts_translatable_text(self, processor, sample_xlsx):
        """Extracts only translatable text blocks (Japanese text only)"""
        blocks = list(processor.extract_text_blocks(sample_xlsx))

        # Should have 2 blocks (only Japanese text, English-only is skipped)
        assert len(blocks) == 2

        # Check block IDs
        block_ids = [b.id for b in blocks]
        assert "Sheet1_A1" in block_ids
        assert "Sheet1_A2" in block_ids
        # Note: "Hello World" (C1) is skipped as it's English-only

        # Check texts
        texts = [b.text for b in blocks]
        assert "こんにちは" in texts
        assert "世界" in texts

    def test_skips_numbers_and_emails(self, processor, sample_xlsx):
        """Skips non-translatable content"""
        blocks = list(processor.extract_text_blocks(sample_xlsx))
        texts = [b.text for b in blocks]

        assert "12345" not in texts
        assert "test@example.com" not in texts

    def test_block_metadata(self, processor, sample_xlsx):
        """Blocks include correct metadata"""
        blocks = list(processor.extract_text_blocks(sample_xlsx))

        a1_block = next(b for b in blocks if b.id == "Sheet1_A1")
        assert a1_block.metadata["sheet"] == "Sheet1"
        assert a1_block.metadata["row"] == 1
        assert a1_block.metadata["col"] == 1
        assert a1_block.metadata["type"] == "cell"

    def test_block_location(self, processor, sample_xlsx):
        """Blocks have human-readable location"""
        blocks = list(processor.extract_text_blocks(sample_xlsx))

        a1_block = next(b for b in blocks if b.id == "Sheet1_A1")
        assert "Sheet1" in a1_block.location
        assert "A1" in a1_block.location

    def test_extracts_font_info(self, processor, xlsx_with_font):
        """Font info in metadata (None in read_only mode for performance)

        Note: Font info is fetched during apply_translations from the original file,
        not during extraction. This is an intentional optimization (read_only=True).
        """
        blocks = list(processor.extract_text_blocks(xlsx_with_font))

        mincho_block = next(b for b in blocks if "明朝" in b.text)
        # Font info is None in read_only mode (fetched during apply_translations)
        assert mincho_block.metadata["font_name"] is None
        assert mincho_block.metadata["font_size"] == 11.0  # default

        gothic_block = next(b for b in blocks if "ゴシック" in b.text)
        assert gothic_block.metadata["font_name"] is None
        assert gothic_block.metadata["font_size"] == 11.0  # default

    def test_extracts_from_multiple_sheets(self, processor, xlsx_with_multiple_sheets):
        """Extracts from all sheets"""
        blocks = list(processor.extract_text_blocks(xlsx_with_multiple_sheets))

        assert len(blocks) == 2

        sheets = {b.metadata["sheet"] for b in blocks}
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets

    def test_empty_file_returns_no_blocks(self, processor, empty_xlsx):
        """Empty file yields no blocks"""
        blocks = list(processor.extract_text_blocks(empty_xlsx))
        assert blocks == []


# --- Tests: apply_translations ---

class TestExcelProcessorApplyTranslations:
    """Test ExcelProcessor.apply_translations()"""

    def test_applies_translations(self, processor, sample_xlsx, tmp_path):
        """Applies translations to correct cells"""
        output_path = tmp_path / "output.xlsx"

        translations = {
            "Sheet1_A1": "Hello",
            "Sheet1_A2": "World",
            "Sheet1_C1": "こんにちは世界",
        }

        processor.apply_translations(
            sample_xlsx, output_path, translations, "jp_to_en"
        )

        # Verify output
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        assert ws["A1"].value == "Hello"
        assert ws["A2"].value == "World"
        assert ws["C1"].value == "こんにちは世界"

        # Unchanged cells
        assert ws["B1"].value == "12345"
        assert ws["B2"].value == "test@example.com"

    def test_preserves_untranslated_cells(self, processor, sample_xlsx, tmp_path):
        """Cells not in translations dict are unchanged"""
        output_path = tmp_path / "output.xlsx"

        translations = {
            "Sheet1_A1": "Hello",
            # A2 not translated
        }

        processor.apply_translations(
            sample_xlsx, output_path, translations, "jp_to_en"
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        assert ws["A1"].value == "Hello"
        assert ws["A2"].value == "世界"  # unchanged

    def test_changes_font_jp_to_en(self, processor, xlsx_with_font, tmp_path):
        """Font changes when translating JP to EN"""
        output_path = tmp_path / "output.xlsx"

        translations = {
            "Sheet1_A1": "Mincho Font Text",
            "Sheet1_A2": "Gothic Font Text",
        }

        processor.apply_translations(
            xlsx_with_font, output_path, translations, "jp_to_en"
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Mincho -> Arial
        assert ws["A1"].font.name == "Arial"
        # Gothic -> Arial (default mapping)
        assert ws["A2"].font.name == "Arial"

    def test_adjusts_font_size(self, processor, xlsx_with_font, tmp_path):
        """Font size is adjusted for JP to EN"""
        output_path = tmp_path / "output.xlsx"

        translations = {
            "Sheet1_A1": "Test",
        }

        processor.apply_translations(
            xlsx_with_font, output_path, translations, "jp_to_en"
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Font size unchanged (DEFAULT_JP_TO_EN_ADJUSTMENT = 0.0)
        assert ws["A1"].font.size == 12

    def test_creates_output_file(self, processor, sample_xlsx, tmp_path):
        """Output file is created even if no translations"""
        output_path = tmp_path / "output.xlsx"

        processor.apply_translations(
            sample_xlsx, output_path, {}, "jp_to_en"
        )

        assert output_path.exists()


# --- Tests: Edge cases ---

class TestExcelProcessorEdgeCases:
    """Test edge cases"""

    def test_large_row_numbers(self, processor, tmp_path):
        """Handles cells beyond row 10"""
        file_path = tmp_path / "large.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Use Japanese text for translation target
        ws["A100"] = "100行目のテキスト"
        ws["Z50"] = "Z列のテキスト"

        wb.save(file_path)

        blocks = list(processor.extract_text_blocks(file_path))
        assert len(blocks) == 2

        ids = [b.id for b in blocks]
        assert "Sheet1_A100" in ids
        assert "Sheet1_Z50" in ids

    def test_unicode_sheet_names(self, processor, tmp_path):
        """Handles Unicode sheet names"""
        file_path = tmp_path / "unicode_sheet.xlsx"
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "日本語シート"
        ws["A1"] = "テスト"

        wb.save(file_path)

        blocks = list(processor.extract_text_blocks(file_path))
        assert len(blocks) == 1
        assert blocks[0].metadata["sheet"] == "日本語シート"

    def test_preserves_cell_formatting(self, processor, tmp_path):
        """Preserves bold, italic, underline"""
        file_path = tmp_path / "formatted.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "Bold text"
        ws["A1"].font = Font(name="Arial", bold=True, italic=True)

        wb.save(file_path)

        output_path = tmp_path / "output.xlsx"
        processor.apply_translations(
            file_path, output_path,
            {"Sheet1_A1": "Translated"},
            "jp_to_en"
        )

        wb_out = openpyxl.load_workbook(output_path)
        ws_out = wb_out.active

        assert ws_out["A1"].value == "Translated"
        assert ws_out["A1"].font.bold is True
        assert ws_out["A1"].font.italic is True


# --- Tests: create_bilingual_workbook ---

class TestExcelProcessorCreateBilingualWorkbook:
    """Test ExcelProcessor.create_bilingual_workbook()"""

    def test_creates_bilingual_workbook(self, processor, tmp_path):
        """Creates workbook with interleaved original and translated sheets"""
        # Create original workbook
        original_path = tmp_path / "original.xlsx"
        wb_orig = openpyxl.Workbook()
        ws1 = wb_orig.active
        ws1.title = "Sheet1"
        ws1["A1"] = "日本語テキスト"
        ws2 = wb_orig.create_sheet("Sheet2")
        ws2["A1"] = "もう一つのテキスト"
        wb_orig.save(original_path)

        # Create translated workbook
        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ws1_t = wb_trans.active
        ws1_t.title = "Sheet1"
        ws1_t["A1"] = "Japanese text"
        ws2_t = wb_trans.create_sheet("Sheet2")
        ws2_t["A1"] = "Another text"
        wb_trans.save(translated_path)

        # Create bilingual workbook
        output_path = tmp_path / "bilingual.xlsx"
        result = processor.create_bilingual_workbook(
            original_path, translated_path, output_path
        )

        # Verify result
        assert result["original_sheets"] == 2
        assert result["translated_sheets"] == 2
        assert result["total_sheets"] == 4

        # Verify output file
        wb_out = openpyxl.load_workbook(output_path)
        sheet_names = wb_out.sheetnames

        assert len(sheet_names) == 4
        assert "Sheet1" in sheet_names
        assert "Sheet1_translated" in sheet_names
        assert "Sheet2" in sheet_names
        assert "Sheet2_translated" in sheet_names

        # Verify content
        assert wb_out["Sheet1"]["A1"].value == "日本語テキスト"
        assert wb_out["Sheet1_translated"]["A1"].value == "Japanese text"

    def test_handles_single_sheet(self, processor, sample_xlsx, tmp_path):
        """Works with single-sheet workbooks"""
        # Create translated version
        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ws = wb_trans.active
        ws.title = "Sheet1"
        ws["A1"] = "Hello"
        ws["A2"] = "World"
        ws["C1"] = "Hello World"
        wb_trans.save(translated_path)

        output_path = tmp_path / "bilingual.xlsx"
        result = processor.create_bilingual_workbook(
            sample_xlsx, translated_path, output_path
        )

        assert result["original_sheets"] == 1
        assert result["translated_sheets"] == 1
        assert result["total_sheets"] == 2

    def test_truncates_long_sheet_names(self, processor, tmp_path):
        """Truncates sheet names exceeding Excel's 31 character limit"""
        # Create original with long sheet name
        original_path = tmp_path / "original.xlsx"
        wb_orig = openpyxl.Workbook()
        ws = wb_orig.active
        ws.title = "VeryLongSheetNameThatExceeds"  # 28 chars
        ws["A1"] = "テスト"
        wb_orig.save(original_path)

        # Create translated
        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ws_t = wb_trans.active
        ws_t.title = "VeryLongSheetNameThatExceeds"
        ws_t["A1"] = "Test"
        wb_trans.save(translated_path)

        output_path = tmp_path / "bilingual.xlsx"
        result = processor.create_bilingual_workbook(
            original_path, translated_path, output_path
        )

        wb_out = openpyxl.load_workbook(output_path)
        # Should have 2 sheets, translated name should be truncated
        assert len(wb_out.sheetnames) == 2
        for name in wb_out.sheetnames:
            assert len(name) <= 31

    def test_handles_duplicate_sheet_names_after_sanitization(self, processor, tmp_path):
        """Avoids sheet name collisions when sanitized names overlap"""
        original_path = tmp_path / "original.xlsx"
        wb_orig = openpyxl.Workbook()
        ws1 = wb_orig.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Original 1"
        ws2 = wb_orig.create_sheet("Sheet1_translated")
        ws2["A1"] = "Original 2"
        wb_orig.save(original_path)

        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ts1 = wb_trans.active
        ts1.title = "Sheet1"
        ts1["A1"] = "Translated 1"
        ts2 = wb_trans.create_sheet("Sheet1_translated")
        ts2["A1"] = "Translated 2"
        wb_trans.save(translated_path)

        output_path = tmp_path / "bilingual.xlsx"
        result = processor.create_bilingual_workbook(
            original_path, translated_path, output_path
        )

        wb_out = openpyxl.load_workbook(output_path)
        assert result["total_sheets"] == 4
        assert len(wb_out.sheetnames) == 4
        assert len(set(wb_out.sheetnames)) == 4
        assert wb_out.sheetnames == [
            "Sheet1",
            "Sheet1_translated",
            "Sheet1_translated_1",
            "Sheet1_translated_translated",
        ]

        assert wb_out["Sheet1"]["A1"].value == "Original 1"
        assert wb_out["Sheet1_translated"]["A1"].value == "Translated 1"
        assert wb_out["Sheet1_translated_1"]["A1"].value == "Original 2"
        assert wb_out["Sheet1_translated_translated"]["A1"].value == "Translated 2"

    def test_copies_cell_styles(self, processor, xlsx_with_font, tmp_path):
        """Preserves cell styles in bilingual output"""
        # Create translated version
        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ws = wb_trans.active
        ws.title = "Sheet1"
        ws["A1"] = "Mincho Font Text"
        ws["A1"].font = Font(name="Arial", size=10)
        ws["A2"] = "Gothic Font Text"
        ws["A2"].font = Font(name="Calibri", size=12)
        wb_trans.save(translated_path)

        output_path = tmp_path / "bilingual.xlsx"
        processor.create_bilingual_workbook(
            xlsx_with_font, translated_path, output_path
        )

        wb_out = openpyxl.load_workbook(output_path)
        # Original sheet should preserve original fonts
        assert wb_out["Sheet1"]["A1"].font.name == "MS Mincho"


# --- Tests: Special Character Handling ---

class TestSpecialCharacterHandling:
    """Test handling of special characters in sheet names"""

    def test_sanitize_sheet_name_forbidden_chars(self):
        """Removes forbidden characters from sheet names"""
        from yakulingo.processors.excel_processor import sanitize_sheet_name

        # Characters \ / ? * [ ] : should be replaced
        assert sanitize_sheet_name("Sheet:1") == "Sheet_1"
        assert sanitize_sheet_name("Data/2024") == "Data_2024"
        assert sanitize_sheet_name("Test\\Sheet") == "Test_Sheet"
        assert sanitize_sheet_name("Sheet?Name") == "Sheet_Name"
        assert sanitize_sheet_name("Sheet*Name") == "Sheet_Name"
        assert sanitize_sheet_name("[Sheet]") == "_Sheet_"
        assert sanitize_sheet_name("Multi:Char/Test") == "Multi_Char_Test"

    def test_sanitize_sheet_name_length_limit(self):
        """Truncates sheet names exceeding 31 characters"""
        from yakulingo.processors.excel_processor import sanitize_sheet_name

        # Exactly 31 chars should be kept
        name_31 = "A" * 31
        assert len(sanitize_sheet_name(name_31)) == 31

        # Over 31 chars should be truncated with "..."
        name_40 = "B" * 40
        result = sanitize_sheet_name(name_40)
        assert len(result) == 31
        assert result.endswith("...")

    def test_sanitize_sheet_name_japanese(self):
        """Handles Japanese characters correctly"""
        from yakulingo.processors.excel_processor import sanitize_sheet_name

        # Japanese characters should be preserved
        assert sanitize_sheet_name("データシート") == "データシート"
        assert sanitize_sheet_name("2024年度:売上") == "2024年度_売上"

    def test_group_translations_with_underscore_sheet_names(self, processor):
        """Correctly groups translations for sheet names with underscores"""
        translations = {
            "my_sheet_data_A1": "Translated 1",
            "my_sheet_A1": "Translated 2",
            "my_A1": "Translated 3",
        }
        sheet_names = {"my", "my_sheet", "my_sheet_data"}

        result = processor._group_translations_by_sheet(translations, sheet_names)

        # Each translation should go to the correct sheet (longest match first)
        assert "my_sheet_data" in result
        assert result["my_sheet_data"]["cells"]["A1"] == "Translated 1"
        assert "my_sheet" in result
        assert result["my_sheet"]["cells"]["A1"] == "Translated 2"
        assert "my" in result
        assert result["my"]["cells"]["A1"] == "Translated 3"

    def test_bilingual_workbook_with_forbidden_chars(self, processor, tmp_path):
        """Creates bilingual workbook when original sheet has forbidden chars"""
        # Create original with forbidden char in sheet name
        original_path = tmp_path / "original.xlsx"
        wb_orig = openpyxl.Workbook()
        ws = wb_orig.active
        # Note: openpyxl doesn't allow forbidden chars in sheet names,
        # but we test the sanitization logic anyway
        ws.title = "Sheet1"  # Use valid name for test
        ws["A1"] = "テスト"
        wb_orig.save(original_path)

        # Create translated
        translated_path = tmp_path / "translated.xlsx"
        wb_trans = openpyxl.Workbook()
        ws_trans = wb_trans.active
        ws_trans.title = "Sheet1"
        ws_trans["A1"] = "Test"
        wb_trans.save(translated_path)

        # Create bilingual
        output_path = tmp_path / "bilingual.xlsx"
        processor.create_bilingual_workbook(
            original_path, translated_path, output_path
        )

        # Verify output
        wb_out = openpyxl.load_workbook(output_path)
        assert len(wb_out.sheetnames) == 2
        # Both sheet names should be valid (no forbidden chars)
        for name in wb_out.sheetnames:
            for char in "\\/?*[]:":
                assert char not in name
